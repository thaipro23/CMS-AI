from __future__ import annotations

from fastapi import APIRouter, Depends, File, Header, HTTPException, Query, UploadFile
from io import BytesIO
from typing import Any

from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

from app.core.errors import public_http_exception
from app.core.config import is_production, settings
from app.core.rbac import UserContext, get_user_context, require_permission
from app.db.session import get_db
from app.models.rbac import UserRoleAssignment
from app.schemas.rbac import (
    EffectiveRBACOut,
    RBACBootstrapOut,
    RBACPermissionOut,
    RBACRoleOut,
    RoleAssignmentBatchCreate,
    RoleAssignmentBatchOut,
    RoleAssignmentCreate,
    RoleAssignmentImportOut,
    RoleAssignmentListOut,
    RoleAssignmentOut,
    RoleAssignmentRevoke,
)
from app.services.audit_log import AuditErrorType, log_audit
from app.services.business_rbac import BusinessRBACService

router = APIRouter()


IMPORT_HEADERS = ['user_id', 'email', 'role_code', 'scope_type', 'scope_id', 'grant_reason', 'sync_openedx']
ROLE_HINTS = [
    ['SYSTEM_ADMIN', 'SYSTEM', '*', 'Quản trị web toàn hệ thống. Chỉ admin kỹ thuật dùng.'],
    ['DEPARTMENT_HEAD', 'DEPARTMENT', 'DEPARTMENT_ID', 'Trưởng bộ môn. VD: CNTT'],
    ['SUBJECT_OWNER', 'SUBJECT', 'SUBJECT_ID', 'Chủ môn theo môn. VD: WEB107'],
    ['SUBJECT_OWNER', 'SUBJECT_VERSION', 'SUBJECT_VERSION_ID', 'Chủ môn theo version/kỳ cụ thể.'],
    ['QUESTION_REVIEWER', 'SUBJECT', 'SUBJECT_ID', 'Người duyệt toàn môn.'],
    ['QUESTION_REVIEWER', 'SUBJECT_VERSION', 'SUBJECT_VERSION_ID', 'Người duyệt trong một version/kỳ.'],
    ['QUESTION_REVIEWER', 'CHAPTER', 'CHAPTER_ID', 'Người duyệt đúng một bài/chapter.'],
    ['CAMPUS_OWNER', 'CAMPUS', 'PH', 'Chủ cơ sở PH, được vận hành sinh viên/lớp trong cơ sở.'],
    ['CAMPUS_OWNER', 'CAMPUS', '*', 'Chủ cơ sở tất cả cơ sở.'],
    ['TEACHER_ASSIGNED', 'CLASS', 'CLASS_ID', 'Giáo viên chỉ xem lớp AP được phân công; CLASS scope là ràng buộc phụ.'],
]


def _truthy(value: Any) -> bool:
    return str(value or '').strip().lower() in {'1', 'true', 'yes', 'y', 'co', 'có', 'x'}


def _cell(value: Any) -> str:
    if value is None:
        return ''
    return str(value).strip()


def _build_import_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'PhanQuyen'
    header_fill = PatternFill('solid', fgColor='111827')
    guide_fill = PatternFill('solid', fgColor='EEF2FF')
    for col, name in enumerate(IMPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = Font(color='FFFFFF', bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    samples = [
        ['head_cntt', 'head_cntt@fpt.edu.vn', 'DEPARTMENT_HEAD', 'DEPARTMENT', 'DEPARTMENT_ID', 'Phụ trách bộ môn CNTT', 'false'],
        ['owner_web107', 'owner_web107@fpt.edu.vn', 'SUBJECT_OWNER', 'SUBJECT', 'SUBJECT_ID', 'Chủ môn WEB107', 'false'],
        ['reviewer_bai1', 'reviewer_bai1@fpt.edu.vn', 'QUESTION_REVIEWER', 'CHAPTER', 'CHAPTER_ID', 'Duyệt câu hỏi Bài 1', 'false'],
    ]
    for row_idx, row in enumerate(samples, 2):
        for col, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col, value=value)
    widths = [24, 32, 28, 22, 32, 42, 16]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width
    role_validation = DataValidation(type='list', formula1='"SYSTEM_ADMIN,DEPARTMENT_HEAD,SUBJECT_OWNER,QUESTION_REVIEWER,CAMPUS_OWNER,CAMPUS_MANAGER,TEACHER_ASSIGNED"', allow_blank=False)
    scope_validation = DataValidation(type='list', formula1='"SYSTEM,DEPARTMENT,SUBJECT,SUBJECT_VERSION,CHAPTER,COURSE,CAMPUS,CLASS"', allow_blank=False)
    bool_validation = DataValidation(type='list', formula1='"false,true"', allow_blank=True)
    ws.add_data_validation(role_validation)
    ws.add_data_validation(scope_validation)
    ws.add_data_validation(bool_validation)
    role_validation.add('C2:C5000')
    scope_validation.add('D2:D5000')
    bool_validation.add('G2:G5000')
    ws.freeze_panes = 'A2'

    guide = wb.create_sheet('HuongDan')
    guide['A1'] = 'Hướng dẫn import phân quyền AI Question Bank'
    guide['A1'].font = Font(size=15, bold=True)
    guide['A3'] = 'Nguyên tắc'
    guide['A3'].font = Font(bold=True)
    notes = [
        'Không tạo role nghiệp vụ trong Open edX. AI Server là nguồn sự thật về phân quyền.',
        'Mỗi dòng là một assignment. Không xóa quyền cũ; hệ thống chỉ thêm hoặc bỏ qua dòng đã tồn tại.',
        'scope_id phải là ID thật trong hệ thống, không phải tên hiển thị. Có thể lấy ID từ màn Phân quyền bằng nút copy.',
        'sync_openedx hiện chỉ lưu yêu cầu vào metadata; chưa tự cấp quyền kỹ thuật Open edX nếu policy sync chưa bật.',
    ]
    for i, note in enumerate(notes, 4):
        guide.cell(row=i, column=1, value=f'- {note}')
    guide['A10'] = 'Role hợp lệ'
    guide['A10'].font = Font(bold=True)
    for c, v in enumerate(['role_code', 'scope_type', 'scope_id mẫu', 'ghi chú'], 1):
        cell = guide.cell(row=11, column=c, value=v)
        cell.fill = guide_fill
        cell.font = Font(bold=True)
    for r, row in enumerate(ROLE_HINTS, 12):
        for c, v in enumerate(row, 1):
            guide.cell(row=r, column=c, value=v)
    for c in range(1, 5):
        guide.column_dimensions[chr(64 + c)].width = [28, 22, 32, 70][c-1]

    raw = BytesIO()
    wb.save(raw)
    return raw.getvalue()


def _read_import_rows(raw: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    ws = wb['PhanQuyen'] if 'PhanQuyen' in wb.sheetnames else wb.active
    header = [_cell(cell.value).lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    index = {name: idx for idx, name in enumerate(header) if name}
    missing = [name for name in IMPORT_HEADERS if name not in index]
    if missing:
        raise HTTPException(status_code=400, detail=f'File Excel thiếu cột: {", ".join(missing)}')
    rows: list[dict[str, Any]] = []
    for row_index, row in enumerate(ws.iter_rows(min_row=2), 2):
        values = {name: row[index[name]].value if index[name] < len(row) else None for name in IMPORT_HEADERS}
        if not any(_cell(v) for v in values.values()):
            continue
        rows.append({
            'row_index': row_index,
            'user_id': _cell(values.get('user_id')),
            'email': _cell(values.get('email')) or None,
            'role_code': _cell(values.get('role_code')).upper(),
            'scope_type': _cell(values.get('scope_type')).upper(),
            'scope_id': _cell(values.get('scope_id')) or '*',
            'grant_reason': _cell(values.get('grant_reason')),
            'sync_openedx': _truthy(values.get('sync_openedx')),
        })
    return rows



@router.get('/me', response_model=EffectiveRBACOut)
def effective_me(user: UserContext = Depends(get_user_context), db: Session = Depends(get_db)):
    service = BusinessRBACService(db)
    assignments = service.active_assignments_for_actor(user)
    raw_claims = user.raw_claims or {}
    permissions = sorted(service.effective_permissions_for_user(user))
    return {
        'user_id': user.user_id,
        'legacy_role': user.role,
        'effective_legacy_role': service.effective_legacy_role_for_user(
            user.user_id,
            user.role,
            email=user.email or raw_claims.get('email'),
            username=user.username or raw_claims.get('username'),
        ),
        'is_system_admin': service.is_system_admin(user),
        'permissions': permissions,
        'business_permissions': permissions,
        'assignments': [service.serialize_assignment(item) for item in assignments],
    }



@router.get('/scope-audit')
def scope_audit(user: UserContext = Depends(get_user_context), db: Session = Depends(get_db)):
    """Explain effective backend scope without exposing cross-campus data.

    v25.9.16.7.2.50 uses this as an operator/debug endpoint so admins can see
    whether a token is unrestricted, campus-scoped, subject-scoped, or only AP
    teacher-assigned before opening /student-management, /teacher-management,
    /analytics/learning, /jobs or /audit.
    """
    service = BusinessRBACService(db)
    assignments = service.active_assignments_for_actor(user)
    try:
        from app.services.academic_service import AcademicService
        decision = AcademicService(db).access_decision(user)
        academic_scope = {
            'unrestricted': bool(decision.unrestricted),
            'teacher_ids': sorted(str(item) for item in (decision.teacher_ids or set())),
            'subject_codes': sorted(str(item) for item in (decision.subject_codes or set())),
            'campus_codes': sorted(str(item) for item in (decision.campus_codes or set())),
        }
    except Exception:
        academic_scope = {'unrestricted': False, 'teacher_ids': [], 'subject_codes': [], 'campus_codes': [], 'error': 'academic_scope_unavailable'}
    visibility = service.visibility_for_user(user)
    return {
        'user_id': user.user_id,
        'email': user.email,
        'username': user.username,
        'legacy_role': user.role,
        'is_system_admin': service.is_system_admin(user),
        'permissions': sorted(service.effective_permissions_for_user(user)),
        'campus_scope': service.campus_scope_for_user(user),
        'academic_scope': academic_scope,
        'bank_scope': {
            'unrestricted': bool(visibility.unrestricted),
            'parent_department_ids': sorted(visibility.parent_department_ids),
            'parent_subject_ids': sorted(visibility.parent_subject_ids),
            'parent_offering_ids': sorted(visibility.parent_offering_ids),
            'broad_department_ids': sorted(visibility.broad_department_ids),
            'broad_subject_ids': sorted(visibility.broad_subject_ids),
            'broad_offering_ids': sorted(visibility.broad_offering_ids),
            'exact_chapter_ids': sorted(visibility.exact_chapter_ids),
        },
        'assignments': [service.serialize_assignment(item) for item in assignments],
        'backend_enforced': True,
    }


@router.get('/roles', response_model=list[RBACRoleOut])
def list_roles(user: UserContext = Depends(require_permission('view_rbac')), db: Session = Depends(get_db)):
    service = BusinessRBACService(db)
    service.ensure_default_catalog()
    return service.list_roles()


@router.get('/permissions', response_model=list[RBACPermissionOut])
def list_permissions(user: UserContext = Depends(require_permission('view_rbac')), db: Session = Depends(get_db)):
    service = BusinessRBACService(db)
    service.ensure_default_catalog()
    return service.list_permissions()


@router.get('/assignments', response_model=RoleAssignmentListOut)
def list_assignments(
    user_id: str | None = None,
    role_code: str | None = None,
    scope_type: str | None = None,
    scope_id: str | None = None,
    include_revoked: bool = False,
    user: UserContext = Depends(require_permission('view_rbac')),
    db: Session = Depends(get_db),
):
    service = BusinessRBACService(db)
    items = service.list_assignments(
        actor=user,
        user_id=user_id,
        role_code=role_code,
        scope_type=scope_type,
        scope_id=scope_id,
        include_revoked=include_revoked,
    )
    return {'items': [service.serialize_assignment(item) for item in items], 'total': len(items)}


@router.post('/assignments', response_model=RoleAssignmentOut)
def create_assignment(payload: RoleAssignmentCreate, user: UserContext = Depends(require_permission('view_rbac')), db: Session = Depends(get_db)):
    service = BusinessRBACService(db)
    try:
        service.ensure_default_catalog()
        item = service.create_assignment(actor=user, **payload.model_dump())
        log_audit(
            db,
            action='rbac.assignment.create',
            status='success',
            message='Gán quyền nghiệp vụ thành công',
            user=user,
            target_type='rbac_assignment',
            target_id=item.id,
            metadata={'assignee': item.user_id, 'role_code': item.role_code, 'scope_type': item.scope_type, 'scope_id': item.scope_id, 'sync_openedx_requested': payload.sync_openedx},
        )
        return service.serialize_assignment(item)
    except HTTPException:
        raise
    except Exception as exc:
        log_audit(db, action='rbac.assignment.create', status='failed', error_type=AuditErrorType.VALIDATION_ERROR, message='Không thể hoàn tất thao tác phân quyền.', user=user, target_type='rbac_assignment')
        raise public_http_exception(status_code=400, code='RBAC_OPERATION_FAILED', message='Không thể hoàn tất thao tác phân quyền.', logger_name=__name__) from exc


@router.post('/assignments/batch', response_model=RoleAssignmentBatchOut)
def create_assignments_batch(payload: RoleAssignmentBatchCreate, user: UserContext = Depends(require_permission('view_rbac')), db: Session = Depends(get_db)):
    service = BusinessRBACService(db)
    try:
        service.ensure_default_catalog()
        items, created_count, reused_count = service.create_assignments_batch(actor=user, **payload.model_dump())
        log_audit(
            db,
            action='rbac.assignment.batch_create',
            status='success',
            message='Gán nhiều phạm vi nghiệp vụ thành công',
            user=user,
            target_type='rbac_assignment_batch',
            target_id=payload.user_id,
            metadata={
                'assignee': payload.user_id,
                'role_code': payload.role_code,
                'scope_type': payload.scope_type,
                'scope_ids': payload.scope_ids,
                'created_count': created_count,
                'reused_count': reused_count,
            },
        )
        return {
            'items': [service.serialize_assignment(item) for item in items],
            'created_count': created_count,
            'reused_count': reused_count,
            'total': len(items),
        }
    except HTTPException:
        db.rollback()
        raise
    except Exception as exc:
        db.rollback()
        log_audit(db, action='rbac.assignment.batch_create', status='failed', error_type=AuditErrorType.VALIDATION_ERROR, message='Không thể hoàn tất thao tác phân quyền hàng loạt.', user=user, target_type='rbac_assignment_batch', target_id=payload.user_id)
        raise public_http_exception(status_code=400, code='RBAC_BATCH_OPERATION_FAILED', message='Không thể hoàn tất thao tác phân quyền hàng loạt.', logger_name=__name__) from exc


@router.delete('/assignments/{assignment_id}', response_model=RoleAssignmentOut)
def revoke_assignment(assignment_id: str, payload: RoleAssignmentRevoke | None = None, user: UserContext = Depends(require_permission('view_rbac')), db: Session = Depends(get_db)):
    service = BusinessRBACService(db)
    try:
        item = service.revoke_assignment(assignment_id, actor=user, revoke_reason=(payload.revoke_reason if payload else ''))
        log_audit(
            db,
            action='rbac.assignment.revoke',
            status='success',
            message='Thu hồi quyền nghiệp vụ thành công',
            user=user,
            target_type='rbac_assignment',
            target_id=item.id,
            metadata={'assignee': item.user_id, 'role_code': item.role_code, 'scope_type': item.scope_type, 'scope_id': item.scope_id},
        )
        return service.serialize_assignment(item)
    except HTTPException:
        raise
    except Exception as exc:
        log_audit(db, action='rbac.assignment.revoke', status='failed', error_type=AuditErrorType.VALIDATION_ERROR, message='Không thể hoàn tất thao tác phân quyền.', user=user, target_type='rbac_assignment', target_id=assignment_id)
        raise public_http_exception(status_code=400, code='RBAC_OPERATION_FAILED', message='Không thể hoàn tất thao tác phân quyền.', logger_name=__name__) from exc


@router.get('/assignments/import-template')
def download_import_template(user: UserContext = Depends(require_permission('view_rbac')), db: Session = Depends(get_db)):
    service = BusinessRBACService(db)
    if not service.has_any_business_permission(user, 'reviewer.assign') and not service.has_any_business_permission(user, 'user.manage_all'):
        raise HTTPException(status_code=403, detail='Bạn không có quyền tải mẫu import phân quyền')
    content = _build_import_template()
    return StreamingResponse(
        BytesIO(content),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="ai-question-bank-rbac-import-template.xlsx"'},
    )


@router.post('/assignments/import', response_model=RoleAssignmentImportOut)
async def import_assignments(
    file: UploadFile = File(...),
    dry_run: bool = Query(default=False),
    user: UserContext = Depends(require_permission('view_rbac')),
    db: Session = Depends(get_db),
):
    service = BusinessRBACService(db)
    if not service.has_any_business_permission(user, 'reviewer.assign') and not service.has_any_business_permission(user, 'user.manage_all'):
        raise HTTPException(status_code=403, detail='Bạn không có quyền import phân quyền')
    name = (file.filename or '').lower()
    if not name.endswith(('.xlsx', '.xlsm')):
        raise HTTPException(status_code=400, detail='Chỉ hỗ trợ file Excel .xlsx/.xlsm')
    raw = await file.read()
    if len(raw) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail='File import quá lớn. Tối đa 5MB.')
    source_rows = _read_import_rows(raw)
    result_rows = []
    valid_rows = created_count = skipped_count = failed_count = 0
    for row in source_rows:
        base = {k: row.get(k) for k in ['row_index', 'user_id', 'email', 'role_code', 'scope_type', 'scope_id']}
        try:
            payload = RoleAssignmentCreate(
                user_id=row['user_id'], email=row.get('email'), role_code=row['role_code'],
                scope_type=row['scope_type'], scope_id=row['scope_id'],
                grant_reason=row.get('grant_reason') or 'Import Excel', sync_openedx=bool(row.get('sync_openedx')),
            )
            service._validate_assignment_scope(payload.role_code, payload.scope_type, payload.scope_id)
            if not service.can_grant(user, payload.role_code, payload.scope_type, payload.scope_id):
                raise HTTPException(status_code=403, detail='Bạn không được gán role này trong scope này')
            existing = service.active_assignments_query().filter(
                UserRoleAssignment.user_id == payload.user_id,
                UserRoleAssignment.role_code == payload.role_code,
                UserRoleAssignment.scope_type == payload.scope_type,
                UserRoleAssignment.scope_id == payload.scope_id,
            ).first()
            valid_rows += 1
            if dry_run:
                result_rows.append({**base, 'scope_label': service.scope_label(payload.scope_type, payload.scope_id), 'status': 'valid', 'message': 'Hợp lệ, chưa ghi DB vì đang dry-run', 'assignment': None})
            elif existing:
                skipped_count += 1
                result_rows.append({**base, 'scope_label': service.scope_label(payload.scope_type, payload.scope_id), 'status': 'skipped', 'message': 'Assignment đã tồn tại, bỏ qua', 'assignment': service.serialize_assignment(existing)})
            else:
                item = service.create_assignment(actor=user, **payload.model_dump())
                created_count += 1
                result_rows.append({**base, 'scope_label': service.scope_label(payload.scope_type, payload.scope_id), 'status': 'created', 'message': 'Đã tạo assignment', 'assignment': service.serialize_assignment(item)})
        except HTTPException as exc:
            failed_count += 1
            result_rows.append({**base, 'scope_label': None, 'status': 'failed', 'message': str(exc.detail), 'assignment': None})
        except Exception as exc:
            failed_count += 1
            result_rows.append({**base, 'scope_label': None, 'status': 'failed', 'message': str(exc), 'assignment': None})
    if not dry_run and (created_count or failed_count or skipped_count):
        log_audit(
            db,
            action='rbac.assignment.import',
            status='success' if failed_count == 0 else 'failed',
            error_type=None if failed_count == 0 else AuditErrorType.VALIDATION_ERROR,
            message=f'Import phân quyền Excel: tạo {created_count}, bỏ qua {skipped_count}, lỗi {failed_count}',
            user=user,
            target_type='rbac_assignment_import',
            metadata={'total_rows': len(source_rows), 'created_count': created_count, 'skipped_count': skipped_count, 'failed_count': failed_count},
        )
    return {
        'ok': failed_count == 0,
        'dry_run': dry_run,
        'total_rows': len(source_rows),
        'valid_rows': valid_rows,
        'created_count': created_count,
        'skipped_count': skipped_count,
        'failed_count': failed_count,
        'rows': result_rows[:500],
    }


@router.post('/bootstrap/system-admin', response_model=RBACBootstrapOut)
def bootstrap_system_admin(
    payload: RoleAssignmentCreate,
    db: Session = Depends(get_db),
    x_rbac_bootstrap_token: str | None = Header(default=None, alias='X-RBAC-Bootstrap-Token'),
):
    """Guarded one-time bootstrap for the first SYSTEM_ADMIN.

    Production is disabled unless RBAC_BOOTSTRAP_TOKEN is configured and the
    exact value is supplied in X-RBAC-Bootstrap-Token. After the first active
    SYSTEM_ADMIN exists, this endpoint refuses to create more admins.
    """
    if is_production():
        if not settings.rbac_bootstrap_token:
            raise HTTPException(status_code=403, detail='RBAC bootstrap bị tắt trong production. Hãy đăng nhập bằng Open edX superuser/AI_ADMIN rồi gán quyền.')
        if x_rbac_bootstrap_token != settings.rbac_bootstrap_token:
            raise HTTPException(status_code=403, detail='Sai RBAC bootstrap token')
    if settings.rbac_bootstrap_token and x_rbac_bootstrap_token != settings.rbac_bootstrap_token:
        raise HTTPException(status_code=403, detail='Sai RBAC bootstrap token')
    if payload.role_code != 'SYSTEM_ADMIN' or payload.scope_type != 'SYSTEM':
        raise HTTPException(status_code=400, detail='Bootstrap chỉ nhận role SYSTEM_ADMIN scope SYSTEM')
    service = BusinessRBACService(db)
    item, created = service.bootstrap_system_admin(user_id=payload.user_id, email=payload.email, reason=payload.grant_reason)
    if not created:
        return {'ok': False, 'created': False, 'message': 'Đã có SYSTEM_ADMIN. Hãy dùng API gán quyền bình thường.', 'assignment': None}
    return {'ok': True, 'created': True, 'message': 'Đã tạo SYSTEM_ADMIN đầu tiên.', 'assignment': service.serialize_assignment(item)}
