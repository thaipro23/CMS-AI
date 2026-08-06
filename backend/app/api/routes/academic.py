from __future__ import annotations

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import FileResponse, Response, StreamingResponse
from io import BytesIO
from pathlib import Path
import hashlib
import json
from typing import Any
from datetime import datetime
from sqlalchemy.orm import Session
from sqlalchemy import and_, func, or_, text
from sqlalchemy.exc import IntegrityError
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.rbac import UserContext, get_user_context, require_permission
from app.db.session import get_db
from app.models.academic import (
    AcademicCampus,
    AcademicBlock,
    AcademicClass,
    AcademicClassStudent,
    AcademicClassSyncJob,
    AcademicBulkOperationJob,
    AcademicTeacherReportJob,
    AcademicQuizDeadlineOverride,
    AcademicAssignmentDefenseScore,
    AcademicStudent,
    AcademicSubject,
    AcademicSyncRun,
    AcademicTeacherAssignment,
    AcademicTerm,
    AcademicSubjectDelivery,
    UdemyProgressImportBatch,
)
from app.models.rbac import UserRoleAssignment
from app.schemas.academic import (
    AcademicAPSyncIn,
    AcademicAPSyncOptionsOut,
    AcademicCampusOut,
    AcademicCampusUpsertIn,
    AcademicCampusUpdateIn,
    AcademicBlockOut,
    AcademicClassListOut,
    AcademicClassSyncJobOut,
    AcademicBulkOperationJobOut,
    AcademicTeacherReportJobOut,
    AcademicClassOut,
    AcademicClassCourseMappingCreateIn,
    AcademicClassCourseMappingOut,
    AcademicClassCourseMappingProposalOut,
    AcademicClassCourseMappingValidateIn,
    AcademicCourseMappingCreateIn,
    AcademicCourseMappingListOut,
    AcademicCourseMappingOut,
    AcademicCourseMappingValidateIn,
    AcademicCourseMappingValidationOut,
    AcademicHealthOut,
    AcademicImportFromJsonIn,
    AcademicEnrollmentSyncIn,
    AcademicEnrollmentSyncOut,
    AcademicFullCmsSyncIn,
    AcademicFullCmsSyncOut,
    AcademicLearningSyncIn,
    AcademicLearningSyncOut,
    AcademicLearningSummaryOut,
    AcademicImportResultOut,
    AcademicIdentityCleanupIn,
    AcademicIdentityCleanupOut,
    AcademicIdentityReconciliationOut,
    AcademicIdentityMigrationOut,
    AcademicMappingResolveOut,
    AcademicMappingSummaryOut,
    AcademicManualMappingImportIn,
    AcademicManualMappingImportOut,
    AcademicResolveClassUsersIn,
    AcademicStudentListOut,
    AcademicSubjectOut,
    AcademicSubjectDeliveryListOut,
    AcademicSubjectCatalogRefreshIn,
    AcademicSubjectCatalogRefreshOut,
    AcademicSubjectPlatformUpdateIn,
    AcademicSubjectPlatformBulkIn,
    AcademicSubjectPlatformMutationOut,
    UdemyPlanDetailOut,
    UdemyPlanVersionCreateIn,
    UdemyPlanImportPreviewOut,
    UdemyPlanImportCommitIn,
    UdemyPlanMutationOut,
    UdemySubjectPlanOut,
    UdemyProgressImportBatchOut,
    UdemyProgressImportJobOut,
    UdemyProgressSummaryOut,
    UdemyProgressDashboardOut,
    UdemyProgressStudentListOut,
    AcademicSubjectManagementListOut,
    AcademicSubjectAutoMapAllSyncIn,
    AcademicSubjectAutoMapAllSyncOut,
    AcademicSubjectCourseAutoMapOut,
    AcademicSyncCounters,
    AcademicSyncRunOut,
    AcademicTermOut,
    AcademicTermUpsertIn,
    AcademicTermWithBlocksOut,
    AcademicQuizDeadlineOverrideBulkIn,
    AcademicQuizDeadlineOverrideOut,
    AcademicAssignmentDefenseScoreBulkIn,
    AcademicAssignmentDefenseScoreOut,
)
from app.services.academic_service import AcademicService
from app.services.academic.subject_delivery import AcademicSubjectDeliveryService
from app.services.academic.udemy_plan import UdemyPlanService
from app.services.academic.udemy_progress import UdemyProgressService
from app.services.academic.ap_sync import AcademicAPSyncWorkflowService
from app.services.academic.assignment_external import AcademicAssignmentExternalWorkflowService
from app.services.ap_academic_sync import AcademicImportService
from app.services.audit_log import AuditErrorType, log_audit
from app.services.business_rbac import BusinessRBACService
from app.core.json_safe import json_safe_value
from app.core.operation_rate_limit import enforce_operation_rate_limit
from app.core.config import settings
from app.core.errors import public_http_exception
from app.services.question_bank.helpers import safe_upload_filename


def _safe_error_message(message: str = 'academic_operation_failed') -> dict[str, str]:
    return {
        'code': message,
        'message': 'Không thể hoàn tất thao tác học vụ. Vui lòng thử lại hoặc liên hệ quản trị.',
    }

router = APIRouter()


def _excel_value(value: Any) -> Any:
    if value is None:
        return ''
    if isinstance(value, (list, tuple, set)):
        return ', '.join(str(item) for item in value if item is not None)
    return value


def _grade10(value: Any) -> float | None:
    if value is None or value == '':
        return None
    try:
        number = float(value)
    except Exception:
        return None
    if 0 <= number <= 1:
        number *= 100.0
    return round(max(0.0, min(100.0, number)) / 10.0, 2)


def _component_key(item: dict[str, Any]) -> str:
    return str(item.get('key') or item.get('name') or '').strip()


def _component_name(item: dict[str, Any]) -> str:
    return str(item.get('name') or item.get('key') or 'Đầu điểm').strip()


def _component_score_text(item: dict[str, Any] | None) -> Any:
    if not item:
        return ''
    if item.get('percent') is not None:
        return _grade10(item.get('percent'))
    try:
        earned = float(item.get('earned'))
        possible = float(item.get('possible'))
        if possible > 0:
            return round(max(0.0, min(10.0, earned / possible * 10.0)), 2)
    except Exception:
        pass
    return ''


def _training_component_columns(report: dict[str, Any]) -> list[dict[str, str]]:
    columns: list[dict[str, str]] = []
    seen: set[str] = set()
    for item in report.get('items') or []:
        for cls in item.get('classes') or []:
            for score in cls.get('learning_component_summaries') or []:
                if not isinstance(score, dict):
                    continue
                key = _component_key(score)
                name = _component_name(score)
                dedupe = (key or name).lower()
                if not dedupe or dedupe in seen:
                    continue
                seen.add(dedupe)
                columns.append({'key': key or name, 'name': name})
    columns.sort(key=lambda item: str(item.get('name') or item.get('key') or '').lower())
    return columns


def _setup_sheet(ws, headers: list[str], widths: list[int] | None = None) -> None:
    header_fill = PatternFill('solid', fgColor='111827')
    for col, name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = Font(color='FFFFFF', bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.freeze_panes = 'A2'
    for idx, width in enumerate(widths or [], 1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _append_row(ws, values: list[Any]) -> None:
    ws.append([_excel_value(value) for value in values])


def _create_training_teacher_report_workbook(report: dict[str, Any]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = 'TongQuanGV'
    overview_headers = [
        'Hệ', 'Cơ sở', 'Giảng viên', 'Username', 'Email', 'Số môn', 'Môn', 'Tổng lớp',
        'Lớp CMS', 'Lớp Udemy', 'SV lượt lớp', 'SV CMS', 'SV Udemy', 'SV riêng biệt',
        'SV học lại', 'Lượt học lại', 'Đã đồng bộ CMS', 'Đã enroll', 'Có hoạt động CMS',
        'Course completion TB (%)', 'Điểm tổng TB (hệ 10)', 'SV Udemy đã import',
        'Tiến độ Udemy TB (%)', 'SV Udemy chậm tiến độ', 'Lần import Udemy gần nhất',
        'Lớp CMS chưa map Course', 'SV rủi ro', 'SV trễ deadline', 'Lượt quiz trễ',
        'SV không được thi', 'SV thiếu dữ liệu xét thi', 'Quiz chưa đạt', 'Assignment chưa chấm',
        'Chưa đồng bộ CMS', 'Chưa enroll', 'Chưa học', 'Tiến độ thấp', 'Điểm thấp',
        'Lỗi đồng bộ', 'Cập nhật CMS gần nhất', 'Cảnh báo'
    ]
    overview_widths = [
        12, 12, 28, 24, 30, 10, 28, 10, 10, 12, 12, 12, 12, 14, 16, 12,
        16, 12, 16, 20, 18, 18, 20, 22, 22, 22, 14, 18, 12, 18, 20, 14, 20,
        20, 14, 12, 14, 12, 14, 22, 52,
    ]
    _setup_sheet(ws, overview_headers, overview_widths)
    for item in report.get('items') or []:
        statuses = item.get('status_counts') or {}
        _append_row(ws, [
            item.get('branch'), item.get('campus'), item.get('teacher_name'), item.get('teacher_username'), item.get('teacher_email'),
            item.get('subject_count'), item.get('subject_codes'), item.get('class_count'), item.get('cms_class_count'), item.get('udemy_class_count'),
            item.get('student_count'), item.get('cms_student_count'), item.get('udemy_student_count'), item.get('unique_student_count'),
            item.get('relearn_student_count'), item.get('total_relearn_count'), item.get('cms_synced_count'), item.get('learning_enrolled_count'), item.get('learning_active_count'),
            item.get('learning_avg_progress_percent'), item.get('learning_avg_grade_10'), item.get('udemy_progress_student_count'),
            item.get('udemy_progress_average_percent'), item.get('udemy_progress_late_count'), item.get('udemy_progress_last_imported_at'),
            item.get('classes_without_course_count'), item.get('risk_student_count'), item.get('deadline_late_student_count'), item.get('deadline_late_quiz_count'),
            item.get('exam_not_eligible_student_count'), item.get('exam_insufficient_data_student_count'), item.get('quiz_failed_count'), item.get('assignment_not_graded_count'),
            statuses.get('cms_not_synced'), statuses.get('not_enrolled'), statuses.get('no_activity'), statuses.get('low_progress'), statuses.get('low_grade'), statuses.get('sync_error'),
            item.get('last_synced_at'), item.get('learning_alerts'),
        ])

    class_ws = wb.create_sheet('ChiTietLop')
    component_columns = _training_component_columns(report)
    class_headers = [
        'Giảng viên', 'Username GV', 'Hệ', 'Cơ sở', 'Học kỳ', 'Block', 'Môn', 'Tên môn',
        'Lớp', 'Tên lớp', 'Nền tảng', 'Subject delivery ID', 'Course CMS', 'Nguồn mapping',
        'SV', 'SV học lại', 'Lượt học lại', 'Đã đồng bộ CMS', 'Đã enroll', 'Có hoạt động CMS',
        'Course completion TB (%)', 'Điểm tổng TB (hệ 10)',
        'SV Udemy đã import', 'Tiến độ Udemy TB (%)', 'SV Udemy chậm', 'Mốc Udemy hiện tại',
        'Tuần Udemy hiện tại', 'Deadline Udemy hiện tại', 'Lần import Udemy gần nhất',
        *[column['name'] for column in component_columns],
        'Số Quiz', 'Quiz đã đến hạn', 'SV trễ deadline', 'Lượt quiz trễ', 'SV được thi',
        'SV không được thi', 'SV thiếu dữ liệu xét thi', 'Quiz chưa đạt', 'Assignment chưa chấm',
        'Đợt quiz kế tiếp', 'Ngày làm quiz kế tiếp', 'Deadline kế tiếp', 'Chưa đồng bộ CMS',
        'Chưa enroll', 'Chưa học', 'Tiến độ thấp', 'Điểm thấp', 'Lỗi đồng bộ',
        'Cập nhật CMS gần nhất', 'Cảnh báo'
    ]
    class_widths = [
        28, 22, 10, 12, 18, 16, 12, 30, 16, 26, 12, 38, 38, 20, 8, 16, 12,
        16, 12, 18, 20, 18, 18, 20, 16, 18, 16, 22, 22,
        *([12] * len(component_columns)),
        10, 14, 16, 14, 18, 18, 20, 16, 20, 18, 22, 22, 20, 14, 12, 14, 12,
        14, 22, 52,
    ]
    _setup_sheet(class_ws, class_headers, class_widths)
    for item in report.get('items') or []:
        for cls in item.get('classes') or []:
            statuses = cls.get('status_counts') or {}
            platform = str(cls.get('learning_platform') or 'cms').strip().lower()
            _append_row(class_ws, [
                item.get('teacher_name'), item.get('teacher_username'), cls.get('branch'), cls.get('campus'), cls.get('term_name'), cls.get('block_name'),
                cls.get('subject_code'), cls.get('subject_name'), cls.get('class_code'), cls.get('class_name'), platform.upper(), cls.get('subject_delivery_id'),
                cls.get('openedx_course_id') if platform != 'udemy' else None, cls.get('openedx_mapping_source') if platform != 'udemy' else None,
                cls.get('student_count'), cls.get('relearn_student_count'), cls.get('total_relearn_count'), cls.get('cms_synced_count'), cls.get('learning_enrolled_count'), cls.get('learning_active_count'),
                cls.get('learning_avg_progress_percent'), cls.get('learning_avg_grade_10'), cls.get('udemy_progress_student_count'), cls.get('udemy_progress_average_percent'),
                cls.get('udemy_progress_late_count'), cls.get('udemy_progress_required_percent'), cls.get('udemy_progress_current_week'), cls.get('udemy_progress_deadline_date'), cls.get('udemy_progress_last_imported_at'),
                *[
                    _component_score_text(next((score for score in (cls.get('learning_component_summaries') or []) if isinstance(score, dict) and (_component_key(score) == column['key'] or _component_name(score) == column['name'])), None))
                    for column in component_columns
                ],
                cls.get('deadline_quiz_count'), cls.get('deadline_due_quiz_count'), cls.get('deadline_late_student_count'), cls.get('deadline_late_quiz_count'),
                cls.get('exam_eligible_student_count'), cls.get('exam_not_eligible_student_count'), cls.get('exam_insufficient_data_student_count'), cls.get('quiz_failed_count'), cls.get('assignment_not_graded_count'),
                cls.get('deadline_next_quiz_label'), cls.get('deadline_next_quiz_from_date'), cls.get('deadline_next_quiz_due_date'),
                statuses.get('cms_not_synced'), statuses.get('not_enrolled'), statuses.get('no_activity'), statuses.get('low_progress'), statuses.get('low_grade'), statuses.get('sync_error'),
                cls.get('learning_last_synced_at'), cls.get('learning_alerts'),
            ])

    udemy_alert_ws = wb.create_sheet('UdemyChamTienDo')
    udemy_alert_headers = [
        'Giảng viên', 'Username GV', 'Hệ', 'Cơ sở', 'Học kỳ', 'Block', 'Môn', 'Tên môn',
        'Lớp', 'SV lớp', 'SV đã import', 'Tiến độ TB (%)', 'SV chậm tiến độ',
        'Mốc yêu cầu (%)', 'Tuần kế hoạch', 'Deadline', 'Lần import gần nhất', 'Mở dashboard'
    ]
    _setup_sheet(udemy_alert_ws, udemy_alert_headers, [28, 22, 10, 12, 18, 16, 12, 30, 16, 10, 14, 18, 18, 18, 16, 18, 22, 58])
    for item in report.get('items') or []:
        for cls in item.get('classes') or []:
            if str(cls.get('learning_platform') or '').strip().lower() != 'udemy':
                continue
            if int(cls.get('udemy_progress_late_count') or 0) <= 0:
                continue
            delivery_id = cls.get('subject_delivery_id')
            _append_row(udemy_alert_ws, [
                item.get('teacher_name'), item.get('teacher_username'), cls.get('branch'), cls.get('campus'), cls.get('term_name'), cls.get('block_name'),
                cls.get('subject_code'), cls.get('subject_name'), cls.get('class_code'), cls.get('student_count'), cls.get('udemy_progress_student_count'),
                cls.get('udemy_progress_average_percent'), cls.get('udemy_progress_late_count'), cls.get('udemy_progress_required_percent'), cls.get('udemy_progress_current_week'),
                cls.get('udemy_progress_deadline_date'), cls.get('udemy_progress_last_imported_at'),
                f'/subject-management/{delivery_id}/udemy' if delivery_id else None,
            ])

    watch_ws = wb.create_sheet('SinhVienCanTheoDoi')
    watch_headers = [
        'Giảng viên', 'Username GV', 'Học kỳ', 'Block', 'Môn', 'Tên môn', 'Lớp', 'Mã SV',
        'Username', 'Họ tên', 'Email', 'Học lại', 'Username CMS', 'Trạng thái', 'Enrollment',
        'Course completion (%)', 'Điểm tổng (hệ 10)', 'Quiz đã đến hạn', 'Quiz đã hoàn thành đúng hạn',
        'Quiz trễ', 'Danh sách quiz trễ', 'Đợt quiz kế tiếp', 'Deadline kế tiếp', 'Hoạt động cuối', 'Cập nhật cuối'
    ]
    _setup_sheet(watch_ws, watch_headers, [28, 22, 18, 16, 12, 30, 16, 14, 22, 28, 32, 22, 20, 16, 22, 18, 14, 20, 12, 34, 18, 18, 22, 22])
    for row in report.get('student_watch_rows') or []:
        _append_row(watch_ws, [
            row.get('teacher_name'), row.get('teacher_username'), row.get('term_name'), row.get('block_name'), row.get('subject_code'), row.get('subject_name'), row.get('class_code'),
            row.get('student_code'), row.get('student_username'), row.get('student_name'), row.get('student_email'), row.get('total_relearn'), row.get('openedx_username'), row.get('status_label'),
            row.get('enrollment_status'), row.get('progress_percent'), row.get('grade_10'), row.get('deadline_due_quiz_count'), row.get('deadline_completed_due_quiz_count'),
            row.get('deadline_late_quiz_count'), row.get('deadline_late_quizzes'), row.get('deadline_next_quiz_label'), row.get('deadline_next_quiz_due_date'), row.get('last_activity_at'), row.get('last_synced_at'),
        ])


    student_ws = wb.create_sheet('ChiTietSinhVien')
    student_headers = [
        'Giảng viên', 'Username GV', 'Học kỳ', 'Block', 'Môn', 'Lớp', 'Mã SV', 'Username', 'Họ tên',
        'Course completion (%)', 'Điểm tổng (hệ 10)', 'Quiz đạt', 'Quiz chưa đạt 100%', 'Quiz chưa làm',
        'Quiz trễ', 'Assignment', 'Điểm Assignment', 'Điều kiện thi', 'Lý do'
    ]
    _setup_sheet(student_ws, student_headers, [26, 20, 18, 14, 12, 16, 14, 20, 26, 18, 18, 12, 16, 12, 12, 16, 16, 18, 60])
    for row in report.get('student_watch_rows') or []:
        _append_row(student_ws, [
            row.get('teacher_name'), row.get('teacher_username'), row.get('term_name'), row.get('block_name'), row.get('subject_code'), row.get('class_code'),
            row.get('student_code'), row.get('student_username'), row.get('student_name'), row.get('progress_percent'), row.get('grade_10'),
            row.get('quiz_passed_count'), row.get('quiz_failed_count'), row.get('quiz_not_attempted_count'), row.get('deadline_late_quiz_count'),
            row.get('assignment_status'), row.get('assignment_score_10'), row.get('exam_status_label') or row.get('exam_status'), row.get('exam_reasons'),
        ])

    late_ws = wb.create_sheet('TreDeadlineQuiz')
    late_headers = ['Giảng viên', 'Môn', 'Lớp', 'Mã SV', 'Username', 'Họ tên', 'Số quiz trễ', 'Danh sách quiz trễ', 'Điều kiện thi', 'Lý do']
    _setup_sheet(late_ws, late_headers, [26, 12, 16, 14, 20, 26, 12, 38, 18, 60])
    for row in report.get('student_watch_rows') or []:
        if int(row.get('deadline_late_quiz_count') or 0) <= 0:
            continue
        _append_row(late_ws, [row.get('teacher_name'), row.get('subject_code'), row.get('class_code'), row.get('student_code'), row.get('student_username'), row.get('student_name'), row.get('deadline_late_quiz_count'), row.get('deadline_late_quizzes'), row.get('exam_status_label') or row.get('exam_status'), row.get('exam_reasons')])

    not100_ws = wb.create_sheet('QuizChuaDat100')
    not100_headers = ['Giảng viên', 'Môn', 'Lớp', 'Mã SV', 'Username', 'Họ tên', 'Quiz chưa đạt 100%', 'Quiz chưa làm', 'Lý do']
    _setup_sheet(not100_ws, not100_headers, [26, 12, 16, 14, 20, 26, 16, 14, 70])
    for row in report.get('student_watch_rows') or []:
        if int(row.get('quiz_failed_count') or 0) <= 0 and int(row.get('quiz_not_attempted_count') or 0) <= 0:
            continue
        _append_row(not100_ws, [row.get('teacher_name'), row.get('subject_code'), row.get('class_code'), row.get('student_code'), row.get('student_username'), row.get('student_name'), row.get('quiz_failed_count'), row.get('quiz_not_attempted_count'), row.get('exam_reasons')])

    assignment_ws = wb.create_sheet('AssignmentBaoVe')
    assignment_headers = ['Giảng viên', 'Môn', 'Lớp', 'Mã SV', 'Username', 'Họ tên', 'Trạng thái Assignment', 'Điểm Assignment', 'Điều kiện thi', 'Lý do']
    _setup_sheet(assignment_ws, assignment_headers, [26, 12, 16, 14, 20, 26, 20, 16, 18, 70])
    for row in report.get('student_watch_rows') or []:
        if not row.get('assignment_status'):
            continue
        _append_row(assignment_ws, [row.get('teacher_name'), row.get('subject_code'), row.get('class_code'), row.get('student_code'), row.get('student_username'), row.get('student_name'), row.get('assignment_status'), row.get('assignment_score_10'), row.get('exam_status_label') or row.get('exam_status'), row.get('exam_reasons')])

    exam_ws = wb.create_sheet('KhongDuocThi')
    exam_headers = ['Giảng viên', 'Môn', 'Lớp', 'Mã SV', 'Username', 'Họ tên', 'Điều kiện thi', 'Lý do']
    _setup_sheet(exam_ws, exam_headers, [26, 12, 16, 14, 20, 26, 18, 90])
    for row in report.get('student_watch_rows') or []:
        if row.get('exam_status') != 'not_eligible':
            continue
        _append_row(exam_ws, [row.get('teacher_name'), row.get('subject_code'), row.get('class_code'), row.get('student_code'), row.get('student_username'), row.get('student_name'), row.get('exam_status_label') or 'Không được thi', row.get('exam_reasons')])

    guide = wb.create_sheet('HuongDan')
    guide['A1'] = 'Báo cáo quản lý đào tạo theo giảng viên'
    guide['A1'].font = Font(size=15, bold=True)
    notes = [
        'Mỗi dòng TongQuanGV là một giảng viên theo bộ lọc trên AI Server.',
        'SV lượt lớp tính theo class-student enrollment; một sinh viên học nhiều lớp có thể được tính nhiều lần.',
        'SV riêng biệt là số sinh viên không trùng trong phạm vi các lớp của giảng viên đó.',
        'Course completion lấy từ progress CMS/Open edX; Điểm tổng được quy đổi từ phần trăm sang hệ 10.',
        'SinhVienCanTheoDoi chỉ liệt kê sinh viên có trạng thái cần xử lý: chưa đồng bộ CMS, chưa enroll, chưa học, tiến độ thấp, điểm thấp, lỗi đồng bộ hoặc trễ deadline quiz.',
        'Deadline quiz được tính theo quy tắc: mỗi block 7 tuần, 6 tuần đầu dành deadline quiz từ Thứ 2 đến Thứ 7, tuần 7 Ôn+Thi; số quiz được chia đều vào 6 tuần và phần dư dồn vào các tuần đầu.',
        'Rule mới: tất cả Quiz phải đạt 100% và hoàn thành trước hoặc đúng deadline. Quiz chưa làm, dưới 100%, làm sau deadline hoặc làm trước thời gian học đều không đạt điều kiện.',
        'Assignment là điểm bảo vệ do giáo viên nhập thủ công; điểm Assignment từ CMS nếu có chỉ là tham khảo.',
        'Lớp Udemy không bị tính là thiếu Course CMS, chưa enroll hoặc chưa đồng bộ CMS.',
        'Tiến độ Udemy được tính từ snapshot import mới nhất và đối chiếu lại với mốc kế hoạch đã đến hạn tại thời điểm mở báo cáo.',
        'Sheet UdemyChamTienDo chỉ liệt kê lớp Udemy còn sinh viên chậm theo mốc hiện hành; xem dashboard môn để tải danh sách sinh viên chi tiết.',
        'Final test chưa áp dụng rule chính thức trong bản này.',
    ]
    for idx, note in enumerate(notes, 3):
        guide.cell(row=idx, column=1, value=f'- {note}')
    guide.column_dimensions['A'].width = 110

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        if sheet.max_row > 1:
            sheet.auto_filter.ref = sheet.dimensions

    return wb


def _build_training_teacher_report_xlsx(report: dict[str, Any]) -> bytes:
    raw = BytesIO()
    _create_training_teacher_report_workbook(report).save(raw)
    return raw.getvalue()


def _write_training_teacher_report_xlsx(report: dict[str, Any], path: Path) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    _create_training_teacher_report_workbook(report).save(path)
    return path.stat().st_size


def _requester_context_json(user: UserContext) -> dict[str, Any]:
    """Store enough requester identity for Celery to re-check real RBAC.

    Do not store bearer tokens or cookies. The worker can reconstruct a
    UserContext from this identity and the current DB business-RBAC assignments.
    """
    return json_safe_value({
        'user_id': user.user_id,
        'username': user.username,
        'email': user.email,
        'role': user.role,
        'permissions': sorted(list(user.permissions or set())),
        'course_ids': list(user.course_ids or []) if user.course_ids is not None else None,
    })


def _advisory_xact_lock_for_key(db: Session, key: str) -> None:
    """Best-effort PostgreSQL transaction lock; no-op on SQLite tests."""
    try:
        bind = db.get_bind()
        if bind and bind.dialect.name == 'postgresql':
            db.execute(text('SELECT pg_advisory_xact_lock(hashtext(:key))'), {'key': key})
    except Exception:
        # Never fail an operator action only because the defensive lock is unavailable.
        pass

def _enqueue_class_sync_job(
    *,
    db: Session,
    user: UserContext,
    class_id: str,
    job_type: str,
    force: bool,
    limit: int,
    mode: str | None = None,
    auto_map_course: bool | None = None,
    sync_learning: bool | None = None,
) -> AcademicClassSyncJob:
    service = AcademicService(db)
    service.assert_can_access_class(user, class_id)
    class_row = db.get(AcademicClass, class_id)
    if not class_row:
        raise HTTPException(status_code=404, detail='Không tìm thấy lớp')
    AcademicSubjectDeliveryService(db).assert_cms_workflow_allowed_for_class(class_id, job_type=job_type)
    clean_limit = max(1, min(500, int(limit or 500)))
    request_key_payload = {
        'class_id': class_id,
        'job_type': job_type,
        'force': bool(force),
        'limit': clean_limit,
        'mode': mode,
        'auto_map_course': auto_map_course,
        'sync_learning': sync_learning,
        'requested_by': user.user_id,
    }
    request_key = hashlib.sha256(
        json.dumps(request_key_payload, sort_keys=True, ensure_ascii=True).encode('utf-8')
    ).hexdigest()

    _advisory_xact_lock_for_key(db, f'academic-class-sync:{class_id}')

    # Class sync jobs mutate the same CMS/Open edX and snapshot rows. Returning
    # the existing active job makes the operation idempotent across refresh/F5
    # and prevents users from accidentally enqueueing duplicate jobs.
    existing_job = (
        db.query(AcademicClassSyncJob)
        .filter(
            AcademicClassSyncJob.class_id == class_id,
            AcademicClassSyncJob.status.in_(['queued', 'running']),
        )
        .order_by(AcademicClassSyncJob.created_at.desc())
        .first()
    )
    if existing_job:
        return existing_job

    job = AcademicClassSyncJob(
        job_type=job_type,
        status='queued',
        class_id=class_id,
        requested_by=user.user_id,
        force=bool(force),
        limit=clean_limit,
        mode=mode,
        progress_current=0,
        progress_total=100,
        progress_label='Đang chờ xử lý',
        request_json=json_safe_value({
            'force': bool(force),
            'limit': clean_limit,
            'mode': mode,
            'auto_map_course': auto_map_course,
            'sync_learning': sync_learning,
            'requester_context': _requester_context_json(user),
            'approved_class_id': class_id,
            'approved_campus_codes': [str(class_row.campus or '').strip().lower()] if class_row.campus else [],
            'approved_branch': str(class_row.branch or '').strip().lower() or None,
            'scope_enforced_by_backend': True,
            'request_key': request_key,
        }),
        result_json={},
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    from app.worker import academic_class_sync_task
    academic_class_sync_task.delay(job.id)
    return job




def _require_academic_sync_permission(
    user: UserContext = Depends(get_user_context),
    db: Session = Depends(get_db),
) -> UserContext:
    """Allow Student Ops CMS/Open edX mutations only for campus/system admins.

    v25.9.16.7.2.64.13 splits Quiz Bank roles from Student Ops. Bank permissions
    such as course.sync no longer authorize AP/CMS class/enrollment mutations.
    """
    if 'manage_settings' in user.permissions:
        return user
    try:
        service = BusinessRBACService(db)
        if service.has_any_business_permission(user, 'manage_settings') or service.has_any_business_permission(user, 'academic.manage_campus'):
            return user
    except HTTPException:
        raise
    except Exception:
        pass
    raise HTTPException(
        status_code=status.HTTP_403_FORBIDDEN,
        detail='Bạn không có quyền đồng bộ/thao tác học vụ CMS/Open edX.',
    )



def _require_academic_catalog_admin(
    user: UserContext = Depends(get_user_context),
    db: Session = Depends(get_db),
) -> UserContext:
    """Restrict global AP imports and campus catalog mutations to system administrators.

    CAMPUS_OWNER is intentionally excluded: that role is scoped to class operations
    inside assigned campuses and must not mutate the global campus catalog or enqueue
    cross-campus AP import jobs.
    """
    if 'manage_settings' in set(user.permissions or []):
        return user
    BusinessRBACService(db).require_system_admin(user)
    return user


def _require_academic_view_permission(
    user: UserContext = Depends(get_user_context),
    db: Session = Depends(get_db),
) -> UserContext:
    """Academic read access for teacher/class operations.

    This keeps campus-management users out of Question Bank read routes while
    still letting them open terms/campuses/teacher-management.
    """
    if 'manage_settings' in set(user.permissions or []):
        return user
    try:
        service = BusinessRBACService(db)
        if service.has_any_business_permission(user, 'view_training_reports') or service.has_any_business_permission(user, 'academic.view'):
            return user
        decision = AcademicService(db).access_decision(user)
        if decision.unrestricted or decision.teacher_ids or decision.campus_codes:
            return user
    except HTTPException:
        raise
    except Exception:
        pass
    raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail='Bạn không có quyền xem dữ liệu đào tạo')


def _require_training_write_permission(
    user: UserContext = Depends(get_user_context),
    db: Session = Depends(get_db),
) -> UserContext:
    """Allow deadline/class training mutations for training-capable users."""
    allowed = {'manage_training_deadlines', 'view_training_reports', 'manage_settings'}
    if allowed.intersection(set(user.permissions or [])):
        return user
    try:
        service = BusinessRBACService(db)
        for permission in allowed:
            if service.has_any_business_permission(user, permission):
                return user
    except HTTPException:
        raise
    except Exception:
        pass
    raise HTTPException(
        status_code=status.HTTP_403_FORBIDDEN,
        detail='Bạn không có quyền cấu hình đào tạo.',
    )

def _enqueue_teacher_report_job(
    *,
    db: Session,
    user: UserContext,
    job_type: str,
    term_id: str,
    branch: str | None = None,
    campus: str | None = None,
    search: str | None = None,
    learning_status: str | None = None,
    teacher_id: str | None = None,
) -> AcademicTeacherReportJob:
    if not term_id:
        raise HTTPException(status_code=422, detail='Thiếu học kỳ để chạy báo cáo giáo viên')
    rbac = BusinessRBACService(db)
    # A limited CAMPUS_MANAGER must choose a concrete campus before materializing
    # cache/export files. Otherwise a durable worker job with campus=None could
    # build all-campus data outside the actor scope. Subject/teacher scoped users
    # remain constrained by AcademicService in the report query itself.
    rbac.ensure_requested_campus_filter_allowed(
        user,
        campus,
        require_filter_when_scoped=True,
        action='tạo báo cáo giáo viên',
    )
    campus_scope = rbac.campus_scope_for_user(user)
    normalized_branch = branch.strip().lower() if branch else None
    normalized_campus = campus.strip().lower() if campus else None
    request_key_payload = {
        'job_type': job_type,
        'term_id': term_id,
        'branch': normalized_branch,
        'campus': normalized_campus,
        'search': (search or '').strip(),
        'learning_status': (learning_status or '').strip(),
        'teacher_id': (teacher_id or '').strip(),
        'requested_by': user.user_id,
    }
    request_key = hashlib.sha256(json.dumps(request_key_payload, sort_keys=True, ensure_ascii=True).encode('utf-8')).hexdigest()
    active_query = db.query(AcademicTeacherReportJob).filter(
        AcademicTeacherReportJob.job_type == job_type,
        AcademicTeacherReportJob.term_id == term_id,
        AcademicTeacherReportJob.branch == normalized_branch,
        AcademicTeacherReportJob.campus == normalized_campus,
        AcademicTeacherReportJob.status.in_(['queued', 'running']),
    )
    if job_type == 'export_excel':
        active_query = active_query.filter(AcademicTeacherReportJob.requested_by == user.user_id)
    for active in active_query.order_by(AcademicTeacherReportJob.created_at.desc()).limit(20).all():
        active_request = active.request_json if isinstance(active.request_json, dict) else {}
        if job_type == 'rebuild_cache' or active_request.get('request_key') == request_key:
            return active
    job = AcademicTeacherReportJob(
        job_type=job_type,
        status='queued',
        term_id=term_id,
        branch=normalized_branch,
        campus=normalized_campus,
        requested_by=user.user_id,
        progress_current=0,
        progress_total=100,
        progress_label='Đang chờ xử lý',
        request_json=json_safe_value({
            'term_id': term_id,
            'branch': branch,
            'campus': campus,
            'search': search,
            'learning_status': learning_status,
            'teacher_id': teacher_id,
            'requester_context': _requester_context_json(user),
            'approved_campus_codes': campus_scope.get('campus_codes') or ([] if campus_scope.get('unrestricted') else []),
            'campus_scope_unrestricted': bool(campus_scope.get('unrestricted')),
            'scope_enforced_by_backend': True,
            'request_key': request_key,
        }),
        result_json={},
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    from app.worker import academic_teacher_report_job_task
    academic_teacher_report_job_task.delay(job.id)
    return job


@router.post('/training/teachers/report-cache/jobs', response_model=AcademicTeacherReportJobOut)
def enqueue_training_teacher_cache_job(
    term_id: str = Query(...),
    branch: str | None = Query(None),
    campus: str | None = Query(None),
    user: UserContext = Depends(_require_academic_view_permission),
    db: Session = Depends(get_db),
):
    return _enqueue_teacher_report_job(db=db, user=user, job_type='rebuild_cache', term_id=term_id, branch=branch, campus=campus)


@router.post('/training/teachers/export/jobs', response_model=AcademicTeacherReportJobOut)
def enqueue_training_teacher_export_job(
    term_id: str = Query(...),
    branch: str | None = Query(None),
    campus: str | None = Query(None),
    search: str | None = Query(None),
    learning_status: str | None = Query(None),
    teacher_id: str | None = Query(None),
    user: UserContext = Depends(_require_academic_view_permission),
    db: Session = Depends(get_db),
):
    return _enqueue_teacher_report_job(db=db, user=user, job_type='export_excel', term_id=term_id, branch=branch, campus=campus, search=search, learning_status=learning_status, teacher_id=teacher_id)



@router.get('/training/teachers/report-jobs', response_model=list[AcademicTeacherReportJobOut])
def list_training_teacher_report_jobs(
    status_filter: str = Query('all', alias='status'),
    limit: int = Query(20, ge=1, le=80),
    user: UserContext = Depends(_require_academic_view_permission),
    db: Session = Depends(get_db),
):
    query = db.query(AcademicTeacherReportJob)
    if status_filter == 'active':
        query = query.filter(AcademicTeacherReportJob.status.in_(['queued', 'running']))
    elif status_filter and status_filter != 'all':
        query = query.filter(AcademicTeacherReportJob.status == status_filter)
    service = BusinessRBACService(db)
    candidates = query.order_by(AcademicTeacherReportJob.created_at.desc()).limit(max(limit * 5, 50)).all()
    visible: list[AcademicTeacherReportJob] = []
    for job in candidates:
        if service.can_access_academic_scope(user, campus=job.campus, requested_by=job.requested_by, request_json=job.request_json if isinstance(job.request_json, dict) else {}):
            visible.append(job)
        if len(visible) >= limit:
            break
    return visible


@router.get('/training/teachers/report-jobs/{job_id}', response_model=AcademicTeacherReportJobOut)
def get_training_teacher_report_job(
    job_id: str,
    user: UserContext = Depends(_require_academic_view_permission),
    db: Session = Depends(get_db),
):
    job = db.get(AcademicTeacherReportJob, job_id)
    if not job:
        raise HTTPException(status_code=404, detail='Không tìm thấy job báo cáo giáo viên')
    BusinessRBACService(db).require_academic_scope(user, campus=job.campus, requested_by=job.requested_by, request_json=job.request_json if isinstance(job.request_json, dict) else {}, action='xem job báo cáo giáo viên')
    return job


@router.get('/training/teachers/report-jobs/{job_id}/download')
def download_training_teacher_report_job_file(
    job_id: str,
    user: UserContext = Depends(_require_academic_view_permission),
    db: Session = Depends(get_db),
):
    job = db.get(AcademicTeacherReportJob, job_id)
    if not job or job.status != 'completed' or not job.file_path:
        raise HTTPException(status_code=404, detail='File báo cáo chưa sẵn sàng')
    BusinessRBACService(db).require_academic_scope(user, campus=job.campus, requested_by=job.requested_by, request_json=job.request_json if isinstance(job.request_json, dict) else {}, action='tải file báo cáo giáo viên')
    root = Path(settings.local_storage_path or '/app/.runtime').expanduser().resolve()
    path = Path(job.file_path).expanduser().resolve()
    if root not in path.parents and path != root:
        raise HTTPException(status_code=403, detail='Đường dẫn file báo cáo không hợp lệ')
    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail='File báo cáo không còn tồn tại')
    filename = job.file_name or path.name
    return StreamingResponse(
        path.open('rb'),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )



def _require_assignment_score_permission_for_class(db: Session, user: UserContext, class_id: str) -> None:
    """Assignment score input is limited to system admin or campus manager."""
    service = BusinessRBACService(db)
    if service.is_system_admin(user):
        return
    cls = db.get(AcademicClass, class_id)
    if not cls:
        raise HTTPException(status_code=404, detail='Không tìm thấy lớp')
    if service.can_manage_assignment_scores_for_campus(user, cls.campus):
        return
    raise HTTPException(status_code=403, detail='Chỉ Quản trị viên hoặc quyền cơ sở được nhập/sửa điểm Assignment')


def _require_academic_admin(db: Session, user: UserContext) -> None:
    service = BusinessRBACService(db)
    if not service.is_system_admin(user):
        # manage_settings also lets current legacy admins pass through the normal require_permission bridge.
        service.require_system_admin(user)


@router.get('/health', response_model=AcademicHealthOut)
def academic_health(user: UserContext = Depends(require_permission('view_questions')), db: Session = Depends(get_db)):
    last_sync = db.query(AcademicSyncRun).order_by(AcademicSyncRun.created_at.desc()).first()
    return {
        'ok': True,
        'terms': db.query(func.count(AcademicTerm.id)).scalar() or 0,
        'classes': db.query(func.count(AcademicClass.id)).scalar() or 0,
        'students': db.query(func.count(AcademicStudent.id)).scalar() or 0,
        'assignments': db.query(func.count(AcademicTeacherAssignment.id)).scalar() or 0,
        'last_sync': last_sync,
    }



@router.get('/training/teachers')
def list_training_teacher_report(
    term_id: str | None = None,
    branch: str | None = None,
    campus: str | None = None,
    search: str | None = None,
    learning_status: str | None = Query(None, description='Lọc giáo viên theo cảnh báo học tập'),
    teacher_id: str | None = Query(None, description='Lọc đúng một giáo viên để mở trang lớp'),
    include_classes: bool = Query(False, description='Chỉ bật khi mở chi tiết giảng viên để tránh payload lớn ở danh sách'),
    fresh: bool = Query(False, description='Bỏ cache khi cần đối soát số liệu mới nhất'),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    user: UserContext = Depends(_require_academic_view_permission),
    db: Session = Depends(get_db),
):
    return AcademicService(db).training_teacher_report(
        user,
        term_id=term_id,
        branch=branch,
        campus=campus,
        search=search,
        learning_status=learning_status,
        teacher_id=teacher_id,
        page=page,
        page_size=page_size,
        include_classes=include_classes or bool(teacher_id),
        use_cache=not fresh,
    )


@router.get('/training/teachers/export')
def export_training_teacher_report(
    term_id: str | None = None,
    branch: str | None = None,
    campus: str | None = None,
    search: str | None = None,
    learning_status: str | None = Query(None, description='Lọc giáo viên theo cảnh báo học tập'),
    teacher_id: str | None = Query(None, description='Lọc đúng một giáo viên để xuất lớp'),
    user: UserContext = Depends(_require_academic_view_permission),
    db: Session = Depends(get_db),
):
    service = AcademicService(db)
    preview = service.training_teacher_report(
        user,
        term_id=term_id,
        branch=branch,
        campus=campus,
        search=search,
        learning_status=learning_status,
        teacher_id=teacher_id,
        page=1,
        page_size=1,
        include_classes=False,
        include_students=False,
        use_cache=True,
    )
    summary = preview.get('summary') or {}
    teacher_total = int(preview.get('total') or 0)
    student_total = int(summary.get('unique_student_count') or summary.get('student_count') or 0)
    if (
        teacher_total > int(settings.academic_teacher_report_sync_export_max_teachers)
        or student_total > int(settings.academic_teacher_report_sync_export_max_students)
    ):
        raise public_http_exception(
            status_code=409,
            code='TEACHER_EXPORT_REQUIRES_BACKGROUND_JOB',
            message='Báo cáo vượt giới hạn xuất trực tiếp. Hãy dùng tác vụ Xuất Excel nền.',
            logger_name=__name__,
        )
    report = service.training_teacher_report(
        user,
        term_id=term_id,
        branch=branch,
        campus=campus,
        search=search,
        learning_status=learning_status,
        teacher_id=teacher_id,
        page=1,
        page_size=max(1, min(200, int(settings.academic_teacher_report_sync_export_max_teachers))),
        include_all=True,
        include_students=True,
        use_cache=False,
    )
    content = _build_training_teacher_report_xlsx(report)
    return StreamingResponse(
        BytesIO(content),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="teacher-management-report.xlsx"'},
    )


@router.get('/terms', response_model=list[AcademicTermOut])
def list_terms(
    branch: str | None = None,
    active: bool | None = True,
    user: UserContext = Depends(_require_academic_view_permission),
    db: Session = Depends(get_db),
):
    return AcademicService(db).list_terms(branch=branch, active=active)




@router.get('/terms/{term_id}/with-blocks', response_model=AcademicTermWithBlocksOut)
def get_term_with_blocks(
    term_id: str,
    active_blocks: bool | None = None,
    user: UserContext = Depends(require_permission('manage_settings')),
    db: Session = Depends(get_db),
):
    _require_academic_admin(db, user)
    service = AcademicService(db)
    term = db.query(AcademicTerm).filter(AcademicTerm.id == term_id).first()
    if not term:
        raise HTTPException(status_code=404, detail='Không tìm thấy học kỳ')
    blocks = service.list_blocks(term_id=term_id, active=active_blocks)
    data = AcademicTermOut.model_validate(term).model_dump()
    data['blocks'] = [AcademicBlockOut.model_validate(item).model_dump() for item in blocks]
    return data


@router.post('/terms', response_model=AcademicTermWithBlocksOut)
def save_academic_term(
    payload: AcademicTermUpsertIn,
    user: UserContext = Depends(require_permission('manage_settings')),
    db: Session = Depends(get_db),
):
    _require_academic_admin(db, user)
    term = AcademicService(db).save_term_with_blocks(payload.model_dump())
    blocks = AcademicService(db).list_blocks(term_id=term.id, active=None)
    log_audit(db, action='academic.term.upsert', status='success', message='Lưu học kỳ/block thành công', user=user, target_type='academic_term', target_id=term.id, metadata={'term_code': term.term_code, 'branch': term.branch, 'block_count': len(blocks)})
    data = AcademicTermOut.model_validate(term).model_dump()
    data['blocks'] = [AcademicBlockOut.model_validate(item).model_dump() for item in blocks]
    return data


@router.delete('/terms/{term_id}', response_model=AcademicTermWithBlocksOut)
def delete_academic_term(
    term_id: str,
    user: UserContext = Depends(require_permission('manage_settings')),
    db: Session = Depends(get_db),
):
    _require_academic_admin(db, user)
    service = AcademicService(db)
    term = db.query(AcademicTerm).filter(AcademicTerm.id == term_id).first()
    if not term:
        raise HTTPException(status_code=404, detail='Không tìm thấy học kỳ')
    term.active = False
    meta = dict(term.metadata_json or {})
    meta.update({'deleted_from_ui': True})
    term.metadata_json = meta
    blocks = service.list_blocks(term_id=term.id, active=None)
    for block in blocks:
        block.active = False
    db.commit()
    db.refresh(term)
    log_audit(db, action='academic.term.delete', status='success', message='Đã xóa/ẩn học kỳ và block', user=user, target_type='academic_term', target_id=term.id, metadata={'term_code': term.term_code, 'branch': term.branch})
    data = AcademicTermOut.model_validate(term).model_dump()
    data['blocks'] = [AcademicBlockOut.model_validate(item).model_dump() for item in blocks]
    return data


@router.get('/blocks', response_model=list[AcademicBlockOut])
def list_blocks(
    term_id: str,
    active: bool | None = True,
    user: UserContext = Depends(_require_academic_view_permission),
    db: Session = Depends(get_db),
):
    return AcademicService(db).list_blocks(term_id=term_id, active=active)


@router.get('/subject-deliveries', response_model=AcademicSubjectDeliveryListOut)
def list_academic_subject_deliveries(
    term_id: str | None = None,
    block_id: str | None = None,
    branch: str | None = None,
    learning_platform: str | None = Query(None, alias='platform'),
    search: str | None = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    user: UserContext = Depends(_require_academic_catalog_admin),
    db: Session = Depends(get_db),
):
    del user
    return AcademicSubjectDeliveryService(db).list_deliveries(
        term_id=term_id,
        block_id=block_id,
        branch=branch,
        learning_platform=learning_platform,
        search=search,
        page=page,
        page_size=page_size,
    )


@router.post('/subject-deliveries/catalog-refresh/jobs', response_model=AcademicSubjectCatalogRefreshOut)
def enqueue_academic_subject_catalog_refresh(
    payload: AcademicSubjectCatalogRefreshIn,
    user: UserContext = Depends(_require_academic_catalog_admin),
    db: Session = Depends(get_db),
):
    term = db.get(AcademicTerm, payload.term_id)
    if not term:
        raise HTTPException(status_code=404, detail='Không tìm thấy học kỳ.')
    if payload.block_id:
        block = db.get(AcademicBlock, payload.block_id)
        if not block or block.term_id != term.id:
            raise HTTPException(status_code=422, detail='Block không thuộc học kỳ đã chọn.')

    branch_value = AcademicSubjectDeliveryService.normalize_branch(payload.branch or term.branch)
    active_candidates = (
        db.query(AcademicBulkOperationJob)
        .filter(
            AcademicBulkOperationJob.job_type == 'subject_catalog_refresh',
            AcademicBulkOperationJob.term_id == payload.term_id,
            AcademicBulkOperationJob.branch == branch_value,
            AcademicBulkOperationJob.status.in_(['queued', 'running']),
        )
        .order_by(AcademicBulkOperationJob.created_at.desc())
        .limit(20)
        .all()
    )
    for active in active_candidates:
        request_json = active.request_json if isinstance(active.request_json, dict) else {}
        if (request_json.get('block_id') or None) == (payload.block_id or None):
            return {
                'ok': True,
                'message': 'Đang có tác vụ lấy danh sách môn cho phạm vi này. Hệ thống dùng lại tác vụ hiện có; F5 không làm mất tiến trình.',
                'job_id': active.id,
                'status': active.status,
                'term_id': payload.term_id,
                'block_id': payload.block_id,
                'branch': branch_value,
            }

    request_json = {
        'term_id': payload.term_id,
        'block_id': payload.block_id,
        'branch': branch_value,
        'requester_context': _requester_context_json(user),
        'policy_version': 'udemy-subject-management/batch31',
    }
    job = AcademicBulkOperationJob(
        job_type='subject_catalog_refresh',
        status='queued',
        term_id=payload.term_id,
        branch=branch_value,
        campus=None,
        requested_by=user.user_id,
        progress_current=0,
        progress_total=100,
        progress_label='Đã đưa yêu cầu lấy danh sách môn từ AP vào hàng đợi',
        request_json=json_safe_value(request_json),
        result_json={},
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    from app.worker import academic_subject_catalog_refresh_task
    academic_subject_catalog_refresh_task.delay(job.id)
    message = 'Đã tạo tác vụ lấy danh sách môn từ AP. Bạn có thể F5 hoặc chuyển màn hình; tác vụ vẫn tiếp tục chạy.'
    log_audit(
        db,
        action='academic.subject_delivery.catalog_refresh.enqueue',
        status='queued',
        message=message,
        user=user,
        target_type='academic_bulk_operation_job',
        target_id=job.id,
        metadata=json_safe_value(request_json),
    )
    return {
        'ok': True,
        'message': message,
        'job_id': job.id,
        'status': job.status,
        'term_id': payload.term_id,
        'block_id': payload.block_id,
        'branch': branch_value,
    }


@router.patch('/subject-deliveries/{delivery_id}/platform', response_model=AcademicSubjectPlatformMutationOut)
def update_academic_subject_delivery_platform(
    delivery_id: str,
    payload: AcademicSubjectPlatformUpdateIn,
    user: UserContext = Depends(_require_academic_catalog_admin),
    db: Session = Depends(get_db),
):
    service = AcademicSubjectDeliveryService(db)
    delivery = service.set_platform(delivery_id, payload.learning_platform, actor=user.user_id or user.username)
    platform_label = {'cms': 'CMS', 'udemy': 'Udemy'}.get(delivery.learning_platform, 'Chưa chọn')
    message = f'Đã đặt nền tảng môn thành {platform_label}. Dữ liệu lịch sử không bị xóa.'
    log_audit(
        db,
        action='academic.subject_delivery.platform.update',
        status='success',
        message=message,
        user=user,
        target_type='academic_subject_delivery',
        target_id=delivery.id,
        metadata={'learning_platform': delivery.learning_platform, 'term_id': delivery.term_id, 'block_id': delivery.block_id, 'subject_id': delivery.subject_id, 'branch': delivery.branch},
    )
    return {'ok': True, 'message': message, 'updated': 1, 'items': []}


@router.post('/subject-deliveries/platform/bulk', response_model=AcademicSubjectPlatformMutationOut)
def bulk_update_academic_subject_delivery_platform(
    payload: AcademicSubjectPlatformBulkIn,
    user: UserContext = Depends(_require_academic_catalog_admin),
    db: Session = Depends(get_db),
):
    service = AcademicSubjectDeliveryService(db)
    rows = service.bulk_set_platform(payload.delivery_ids, payload.learning_platform, actor=user.user_id or user.username)
    platform_label = {'cms': 'CMS', 'udemy': 'Udemy'}.get(payload.learning_platform, 'Chưa chọn')
    message = f'Đã cập nhật {len(rows)} môn sang {platform_label}. Dữ liệu lịch sử không bị xóa.'
    log_audit(
        db,
        action='academic.subject_delivery.platform.bulk_update',
        status='success',
        message=message,
        user=user,
        target_type='academic_subject_delivery',
        target_id=rows[0].id if rows else None,
        metadata={'learning_platform': payload.learning_platform, 'delivery_ids': [row.id for row in rows], 'updated': len(rows)},
    )
    return {'ok': True, 'message': message, 'updated': len(rows), 'items': []}


@router.get('/udemy/plans/import-template.xlsx')
def download_udemy_plan_import_template(
    user: UserContext = Depends(_require_academic_catalog_admin),
):
    del user
    return Response(
        content=UdemyPlanService.build_template(),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="udemy-plan-import-template.xlsx"'},
    )


@router.post('/udemy/plans/import/preview', response_model=UdemyPlanImportPreviewOut)
async def preview_udemy_plan_import(
    file: UploadFile = File(...),
    branch: str = Form('poly'),
    user: UserContext = Depends(_require_academic_catalog_admin),
    db: Session = Depends(get_db),
):
    filename = safe_upload_filename(file.filename or 'udemy-plan.xlsx')
    if not filename.lower().endswith('.xlsx'):
        raise HTTPException(status_code=400, detail='Chỉ hỗ trợ file .xlsx cho kế hoạch Udemy.')
    raw = await file.read()
    service = UdemyPlanService(db)
    try:
        parsed = service.parse_workbook(raw, filename=filename, branch=branch, requested_by=user.user_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    token = service.persist_preview(parsed)
    message = 'File hợp lệ, có thể xác nhận import.' if parsed['can_commit'] else 'File còn lỗi. Hãy sửa các dòng lỗi rồi upload lại.'
    log_audit(
        db,
        action='academic.udemy_plan.import.preview',
        status='success' if parsed['can_commit'] else 'warning',
        message='Đã kiểm tra file kế hoạch Udemy',
        user=user,
        target_type='udemy_plan_import_preview',
        target_id=token,
        metadata={
            'filename': filename,
            'file_sha256': parsed['file_sha256'],
            'branch': parsed['branch'],
            'total_rows': parsed['total_rows'],
            'valid_count': parsed['valid_count'],
            'error_count': parsed['error_count'],
            'warning_count': parsed['warning_count'],
        },
    )
    return {
        'ok': True,
        'preview_token': token,
        'filename': filename,
        'file_sha256': parsed['file_sha256'],
        'branch': parsed['branch'],
        'total_rows': parsed['total_rows'],
        'valid_count': parsed['valid_count'],
        'error_count': parsed['error_count'],
        'warning_count': parsed['warning_count'],
        'can_commit': parsed['can_commit'],
        'rows': parsed['rows'][:200],
        'errors': parsed['errors'][:200],
        'warnings': parsed['warnings'][:200],
        'message': message,
    }


@router.get('/udemy/plans/import/errors/{preview_token}.xlsx')
def download_udemy_plan_import_errors(
    preview_token: str,
    user: UserContext = Depends(_require_academic_catalog_admin),
    db: Session = Depends(get_db),
):
    _path, preview = UdemyPlanService(db).load_preview(preview_token, requested_by=user.user_id)
    return Response(
        content=UdemyPlanService.build_error_workbook(preview),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="udemy-plan-import-errors.xlsx"'},
    )


@router.post('/udemy/plans/import/commit', response_model=UdemyPlanMutationOut)
def commit_udemy_plan_import(
    payload: UdemyPlanImportCommitIn,
    user: UserContext = Depends(_require_academic_catalog_admin),
    db: Session = Depends(get_db),
):
    service = UdemyPlanService(db)
    preview_path, preview = service.load_preview(payload.preview_token, requested_by=user.user_id)
    plans = service.commit_preview(preview, actor=user.user_id or user.username)
    preview_path.unlink(missing_ok=True)
    message = f'Đã import {len(plans)} kế hoạch Udemy. Mỗi môn được tạo phiên bản mới và giữ nguyên lịch sử cũ.'
    log_audit(
        db,
        action='academic.udemy_plan.import.commit',
        status='success',
        message=message,
        user=user,
        target_type='udemy_subject_plan',
        target_id=plans[0].id if plans else None,
        metadata={
            'filename': preview.get('filename'),
            'file_sha256': preview.get('file_sha256'),
            'created_count': len(plans),
            'plan_ids': [item.id for item in plans],
        },
    )
    return {'ok': True, 'message': message, 'created_count': len(plans), 'plans': [service.serialize_plan(item) for item in plans]}


@router.post('/udemy/progress/import/jobs', response_model=UdemyProgressImportJobOut)
async def create_udemy_progress_import_job(
    files: list[UploadFile] = File(...),
    term_id: str = Form(...),
    block_id: str = Form(...),
    branch: str = Form('poly'),
    delivery_id: str | None = Form(None),
    force_reimport: bool = Form(False),
    user: UserContext = Depends(_require_academic_catalog_admin),
    db: Session = Depends(get_db),
):
    enforce_operation_rate_limit(
        namespace='udemy-progress-upload',
        actor_id=str(user.user_id or user.username or 'unknown'),
        limit=int(settings.academic_udemy_upload_rate_limit_per_minute),
        code='UDEMY_UPLOAD_RATE_LIMITED',
        message='Bạn thao tác import Udemy quá nhanh. Vui lòng đợi rồi thử lại.',
    )
    try:
        active_import_job_ids = {
            str(row[0])
            for row in db.query(AcademicBulkOperationJob.id).filter(
                AcademicBulkOperationJob.job_type == 'udemy_progress_import',
                AcademicBulkOperationJob.status.in_(['queued', 'running']),
            ).all()
        }
        UdemyProgressService.cleanup_expired_artifacts(protected_import_job_ids=active_import_job_ids)
    except Exception:
        # Periodic cleanup task is authoritative; enqueue must remain available
        # when a stale artifact cannot be removed because of filesystem races.
        pass
    if not files:
        raise HTTPException(status_code=400, detail='Hãy chọn ít nhất một file Udemy .xlsx.')
    if len(files) > UdemyProgressService.MAX_FILES:
        raise HTTPException(status_code=400, detail=f'Mỗi lần tối đa {UdemyProgressService.MAX_FILES} file.')
    if delivery_id and len(files) != 1:
        raise HTTPException(status_code=400, detail='Import tại một môn chỉ nhận một file. Hãy bỏ chọn môn để import nhiều file theo tên mã môn.')

    branch_value = str(branch or 'poly').strip().lower()
    service = UdemyProgressService(db)
    prepared: list[dict[str, Any]] = []
    total_bytes = 0
    for upload in files:
        filename = safe_upload_filename(upload.filename or 'udemy-progress.xlsx')
        try:
            UdemyProgressService.validate_upload_metadata(filename=filename, content_type=upload.content_type)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=f'{filename}: {exc}') from exc
        raw = await upload.read(UdemyProgressService.MAX_FILE_BYTES + 1)
        total_bytes += len(raw)
        if len(raw) > UdemyProgressService.MAX_FILE_BYTES:
            raise HTTPException(status_code=400, detail=f'File {filename} vượt giới hạn {max(1, UdemyProgressService.MAX_FILE_BYTES // (1024 * 1024))} MB.')
        if total_bytes > UdemyProgressService.MAX_TOTAL_UPLOAD_BYTES:
            raise HTTPException(status_code=400, detail=f'Tổng dung lượng upload vượt giới hạn {max(1, UdemyProgressService.MAX_TOTAL_UPLOAD_BYTES // (1024 * 1024))} MB.')
        try:
            service._validate_xlsx(raw)
            delivery, subject = service.resolve_delivery(
                term_id=term_id,
                block_id=block_id,
                branch=branch_value,
                filename=filename,
                delivery_id=delivery_id,
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=f'{filename}: {exc}') from exc
        prepared.append({
            'filename': filename,
            'raw': raw,
            'hash': hashlib.sha256(raw).hexdigest(),
            'delivery': delivery,
            'subject': subject,
        })

    parent = AcademicBulkOperationJob(
        job_type='udemy_progress_import',
        status='queued',
        term_id=term_id,
        branch=branch_value,
        campus=None,
        requested_by=user.user_id,
        progress_current=0,
        progress_total=100,
        progress_label='Đã đưa file tiến độ Udemy vào hàng đợi',
        request_json={},
        result_json={},
    )
    db.add(parent)
    db.flush()
    root = Path(settings.local_storage_path or '/app/.runtime').expanduser().resolve()
    upload_dir = root / 'udemy-progress-imports' / parent.id
    upload_dir.mkdir(parents=True, exist_ok=True)

    batches: list[UdemyProgressImportBatch] = []
    queued_ids: list[str] = []
    duplicate_count = 0
    try:
        for item in prepared:
            try:
                bind = db.get_bind()
                if bind is not None and bind.dialect.name == 'postgresql':
                    db.execute(
                        text('SELECT pg_advisory_xact_lock(hashtext(:key))'),
                        {'key': f'udemy-progress:{item["delivery"].id}:{item["hash"]}'},
                    )
            except Exception:
                # The unique idempotency key remains the final protection when the
                # database user cannot acquire advisory locks.
                pass
            prior = db.query(UdemyProgressImportBatch).filter(
                UdemyProgressImportBatch.subject_delivery_id == item['delivery'].id,
                UdemyProgressImportBatch.file_hash == item['hash'],
                UdemyProgressImportBatch.status.in_(['queued', 'running', 'completed']),
            ).order_by(UdemyProgressImportBatch.created_at.desc()).first()
            is_duplicate = bool(prior and not force_reimport)
            batch = UdemyProgressImportBatch(
                parent_job_id=parent.id,
                subject_delivery_id=item['delivery'].id,
                duplicate_of_batch_id=prior.id if is_duplicate else None,
                idempotency_key=(
                    f'duplicate:{item["delivery"].id}:{item["hash"]}:{parent.id}'
                    if is_duplicate else
                    (f'force:{item["delivery"].id}:{item["hash"]}:{parent.id}' if force_reimport else f'{item["delivery"].id}:{item["hash"]}')
                ),
                file_name=item['filename'],
                file_hash=item['hash'],
                file_size_bytes=len(item['raw']),
                status='skipped' if is_duplicate else 'queued',
                force_reimport=bool(force_reimport),
                requested_by=user.user_id,
                request_json=json_safe_value({
                    'term_id': term_id,
                    'block_id': block_id,
                    'branch': branch_value,
                    'subject_code': item['subject'].subject_code,
                    'requester_context': _requester_context_json(user),
                    'policy_version': 'udemy-progress-import/batch35',
                }),
                result_json=(
                    {'ok': True, 'skipped': True, 'message': 'File trùng đã được import trước đó.', 'duplicate_of_batch_id': prior.id}
                    if is_duplicate else {}
                ),
                finished_at=datetime.utcnow() if is_duplicate else None,
            )
            db.add(batch)
            db.flush()
            if is_duplicate:
                duplicate_count += 1
            else:
                path = upload_dir / f'{batch.id}-{item["filename"]}'
                path.write_bytes(item['raw'])
                batch.file_path = str(path)
                db.add(batch)
                queued_ids.append(batch.id)
            batches.append(batch)

        parent.request_json = json_safe_value({
            'term_id': term_id,
            'block_id': block_id,
            'branch': branch_value,
            'delivery_id': delivery_id,
            'force_reimport': bool(force_reimport),
            'batch_ids': [item.id for item in batches],
            'queued_batch_ids': queued_ids,
            'requester_context': _requester_context_json(user),
            'policy_version': 'udemy-progress-import/batch35',
        })
        if not queued_ids:
            parent.status = 'completed'
            parent.progress_current = 100
            parent.progress_label = 'Tất cả file đã được import trước đó; không tạo dữ liệu trùng'
            parent.result_json = {'ok': True, 'queued_count': 0, 'duplicate_count': duplicate_count}
            parent.finished_at = datetime.utcnow()
        db.add(parent)
        db.commit()
    except Exception:
        db.rollback()
        for item in upload_dir.glob('*'):
            if item.is_file():
                item.unlink(missing_ok=True)
        raise

    if queued_ids:
        from app.worker import academic_udemy_progress_import_task
        academic_udemy_progress_import_task.delay(parent.id)
    message = (
        f'Đã xếp hàng {len(queued_ids)} file Udemy; bỏ qua {duplicate_count} file trùng.'
        if queued_ids else
        f'Không xếp hàng file mới; {duplicate_count} file đã được import trước đó.'
    )
    log_audit(
        db,
        action='academic.udemy_progress.import.enqueue',
        status='queued' if queued_ids else 'success',
        message=message,
        user=user,
        target_type='academic_bulk_operation_job',
        target_id=parent.id,
        metadata=json_safe_value({'batch_ids': [item.id for item in batches], 'queued_count': len(queued_ids), 'duplicate_count': duplicate_count}),
    )
    return {
        'ok': True,
        'message': message,
        'job_id': parent.id,
        'status': parent.status,
        'queued_count': len(queued_ids),
        'duplicate_count': duplicate_count,
        'batches': [service.batch_to_dict(item) for item in batches],
    }


@router.post('/udemy/progress/import-batches/{batch_id}/retry', response_model=UdemyProgressImportJobOut)
def retry_udemy_progress_import_batch(
    batch_id: str,
    user: UserContext = Depends(_require_academic_catalog_admin),
    db: Session = Depends(get_db),
):
    enforce_operation_rate_limit(
        namespace='udemy-progress-retry',
        actor_id=str(user.user_id or user.username or 'unknown'),
        limit=int(settings.academic_udemy_upload_rate_limit_per_minute),
        code='UDEMY_RETRY_RATE_LIMITED',
        message='Bạn thao tác thử lại import quá nhanh. Vui lòng đợi rồi thử lại.',
    )
    source = db.get(UdemyProgressImportBatch, batch_id)
    if not source:
        raise HTTPException(status_code=404, detail='Không tìm thấy batch Udemy cần thử lại.')
    if source.status != 'failed':
        raise HTTPException(status_code=409, detail='Chỉ batch thất bại mới được thử lại. File trùng có thể dùng tùy chọn “Import lại có chủ đích”.')
    delivery = db.get(AcademicSubjectDelivery, source.subject_delivery_id)
    subject = db.get(AcademicSubject, delivery.subject_id) if delivery else None
    if not delivery or not delivery.active or not subject:
        raise HTTPException(status_code=409, detail='Môn Udemy của batch không còn tồn tại.')
    if delivery.learning_platform != 'udemy':
        raise HTTPException(status_code=409, detail='Môn đã được chuyển khỏi Udemy nên không thể thử lại file này.')

    root = Path(settings.local_storage_path or '/app/.runtime').expanduser().resolve()
    source_path = Path(str(source.file_path or '')).expanduser().resolve()
    try:
        source_path.relative_to(root)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail='Đường dẫn file import cũ không hợp lệ.') from exc
    if not source_path.is_file():
        raise HTTPException(status_code=410, detail='File import gốc đã hết thời gian lưu. Hãy tải file lên lại.')
    raw = source_path.read_bytes()
    try:
        UdemyProgressService._validate_xlsx(raw)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f'File import cũ không còn hợp lệ: {exc}') from exc
    if hashlib.sha256(raw).hexdigest() != source.file_hash:
        raise HTTPException(status_code=409, detail='Checksum file import cũ đã thay đổi. Hãy tải file gốc lên lại.')

    parent = AcademicBulkOperationJob(
        job_type='udemy_progress_import',
        status='queued',
        term_id=delivery.term_id,
        branch=str(delivery.branch or 'poly').lower(),
        campus=None,
        requested_by=user.user_id,
        progress_current=0,
        progress_total=100,
        progress_label=f'Đã xếp hàng thử lại file {source.file_name}'[:255],
        request_json={},
        result_json={},
    )
    db.add(parent)
    db.flush()
    upload_dir = root / 'udemy-progress-imports' / parent.id
    upload_dir.mkdir(parents=True, exist_ok=True)
    retry_batch = UdemyProgressImportBatch(
        parent_job_id=parent.id,
        subject_delivery_id=delivery.id,
        duplicate_of_batch_id=source.id,
        idempotency_key=f'retry:{source.id}:{parent.id}',
        file_name=source.file_name,
        file_hash=source.file_hash,
        file_size_bytes=len(raw),
        parser_format=source.parser_format,
        status='queued',
        force_reimport=True,
        requested_by=user.user_id,
        request_json=json_safe_value({
            'term_id': delivery.term_id,
            'block_id': delivery.block_id,
            'branch': str(delivery.branch or 'poly').lower(),
            'subject_code': subject.subject_code,
            'retry_of_batch_id': source.id,
            'requester_context': _requester_context_json(user),
            'policy_version': 'udemy-progress-import/batch35.1',
        }),
        result_json={},
    )
    db.add(retry_batch)
    db.flush()
    retry_path = upload_dir / f'{retry_batch.id}-{safe_upload_filename(source.file_name)}'
    try:
        retry_path.write_bytes(raw)
        retry_batch.file_path = str(retry_path)
        parent.request_json = json_safe_value({
            'term_id': delivery.term_id,
            'block_id': delivery.block_id,
            'branch': str(delivery.branch or 'poly').lower(),
            'delivery_id': delivery.id,
            'force_reimport': True,
            'retry_of_batch_id': source.id,
            'batch_ids': [retry_batch.id],
            'queued_batch_ids': [retry_batch.id],
            'requester_context': _requester_context_json(user),
            'policy_version': 'udemy-progress-import/batch35.1',
        })
        db.add(retry_batch)
        db.add(parent)
        db.commit()
    except Exception:
        db.rollback()
        retry_path.unlink(missing_ok=True)
        raise

    from app.worker import academic_udemy_progress_import_task
    academic_udemy_progress_import_task.delay(parent.id)
    message = f'Đã xếp hàng thử lại file {source.file_name}.'
    log_audit(
        db,
        action='academic.udemy_progress.import.retry',
        status='queued',
        message=message,
        user=user,
        target_type='udemy_progress_import_batch',
        target_id=retry_batch.id,
        metadata=json_safe_value({'retry_of_batch_id': source.id, 'parent_job_id': parent.id, 'file_hash': source.file_hash}),
    )
    return {
        'ok': True,
        'message': message,
        'job_id': parent.id,
        'status': parent.status,
        'queued_count': 1,
        'duplicate_count': 0,
        'batches': [UdemyProgressService(db).batch_to_dict(retry_batch)],
    }


@router.get('/udemy/progress/import-batches', response_model=list[UdemyProgressImportBatchOut])
def list_udemy_progress_import_batches(
    delivery_id: str | None = None,
    parent_job_id: str | None = None,
    status_filter: str | None = Query(None, alias='status'),
    limit: int = Query(50, ge=1, le=200),
    user: UserContext = Depends(_require_academic_catalog_admin),
    db: Session = Depends(get_db),
):
    del user
    query = db.query(UdemyProgressImportBatch)
    if delivery_id:
        query = query.filter(UdemyProgressImportBatch.subject_delivery_id == delivery_id)
    if parent_job_id:
        query = query.filter(UdemyProgressImportBatch.parent_job_id == parent_job_id)
    if status_filter:
        query = query.filter(UdemyProgressImportBatch.status == status_filter)
    rows = query.order_by(UdemyProgressImportBatch.created_at.desc()).limit(limit).all()
    service = UdemyProgressService(db)
    return [service.batch_to_dict(item) for item in rows]


@router.get('/udemy/progress/import-batches/{batch_id}/errors.xlsx')
def download_udemy_progress_import_errors(
    batch_id: str,
    user: UserContext = Depends(_require_academic_catalog_admin),
    db: Session = Depends(get_db),
):
    del user
    batch = db.get(UdemyProgressImportBatch, batch_id)
    if not batch or not batch.error_report_path:
        raise HTTPException(status_code=404, detail='Batch chưa có file lỗi để tải.')
    root = Path(settings.local_storage_path or '/app/.runtime').expanduser().resolve()
    path = Path(batch.error_report_path).expanduser().resolve()
    try:
        path.relative_to(root)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail='Đường dẫn file lỗi không hợp lệ.') from exc
    if not path.is_file():
        raise HTTPException(status_code=404, detail='File lỗi đã hết hạn hoặc không còn tồn tại.')
    return FileResponse(
        path,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        filename=f'udemy-progress-errors-{batch.id[:8]}.xlsx',
    )




def _udemy_delivery_access_scope(
    db: Session,
    user: UserContext,
    delivery_id: str,
) -> tuple[AcademicSubjectDelivery, set[str] | None, str]:
    delivery = db.get(AcademicSubjectDelivery, delivery_id)
    if not delivery or not delivery.active:
        raise HTTPException(status_code=404, detail='Không tìm thấy môn Udemy.')
    subject = db.get(AcademicSubject, delivery.subject_id)
    if not subject:
        raise HTTPException(status_code=404, detail='Không tìm thấy môn học.')
    if delivery.learning_platform != 'udemy':
        raise HTTPException(status_code=409, detail='Môn này không được cấu hình trên nền tảng Udemy.')
    decision = AcademicService(db).access_decision(user)
    if decision.unrestricted:
        return delivery, None, 'Toàn bộ môn'

    class_query = db.query(AcademicClass.id).filter(
        AcademicClass.subject_id == delivery.subject_id,
        AcademicClass.term_id == delivery.term_id,
        AcademicClass.block_id == delivery.block_id,
        func.lower(func.coalesce(AcademicClass.branch, delivery.branch)) == str(delivery.branch).lower(),
        AcademicClass.active.is_(True),
    )
    predicates = []
    if decision.campus_codes:
        predicates.append(func.lower(func.coalesce(AcademicClass.campus, '')).in_(sorted(decision.campus_codes)))
    if decision.teacher_ids:
        teacher_class_ids = db.query(AcademicTeacherAssignment.class_id).filter(
            AcademicTeacherAssignment.teacher_id.in_(sorted(decision.teacher_ids)),
        )
        predicates.append(AcademicClass.id.in_(teacher_class_ids))
    if subject.subject_code and subject.subject_code.strip().lower() in decision.subject_codes:
        predicates.append(AcademicClass.subject_id == subject.id)
    if not predicates:
        raise HTTPException(status_code=403, detail='Bạn chưa được phân quyền xem lớp Udemy nào trong môn này.')
    class_ids = {row[0] for row in class_query.filter(or_(*predicates)).all()}
    if not class_ids:
        raise HTTPException(status_code=403, detail='Bạn không được phân công hoặc phân quyền xem lớp Udemy trong môn này.')
    scope_bits = []
    if decision.teacher_ids:
        scope_bits.append('lớp được AP phân công')
    if decision.campus_codes:
        scope_bits.append('cơ sở được phân quyền')
    return delivery, class_ids, ' và '.join(scope_bits) or 'Phạm vi được phân quyền'


@router.get('/subject-deliveries/{delivery_id}/udemy-progress/summary', response_model=UdemyProgressSummaryOut)
def get_udemy_progress_summary(
    delivery_id: str,
    user: UserContext = Depends(_require_academic_view_permission),
    db: Session = Depends(get_db),
):
    _delivery, allowed_class_ids, scope_label = _udemy_delivery_access_scope(db, user, delivery_id)
    return UdemyProgressService(db).dashboard(
        delivery_id, allowed_class_ids=allowed_class_ids, scope_label=scope_label,
    )['summary']


@router.get('/subject-deliveries/{delivery_id}/udemy-progress/dashboard', response_model=UdemyProgressDashboardOut)
def get_udemy_progress_dashboard(
    delivery_id: str,
    user: UserContext = Depends(_require_academic_view_permission),
    db: Session = Depends(get_db),
):
    _delivery, allowed_class_ids, scope_label = _udemy_delivery_access_scope(db, user, delivery_id)
    return UdemyProgressService(db).dashboard(
        delivery_id, allowed_class_ids=allowed_class_ids, scope_label=scope_label,
    )


@router.get('/subject-deliveries/{delivery_id}/udemy-progress/students', response_model=UdemyProgressStudentListOut)
def list_udemy_progress_students(
    delivery_id: str,
    q: str | None = None,
    class_id: str | None = None,
    status_filter: str | None = Query(None, alias='status'),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    sort_by: str = Query('student'),
    sort_dir: str = Query('asc'),
    user: UserContext = Depends(_require_academic_view_permission),
    db: Session = Depends(get_db),
):
    _delivery, allowed_class_ids, _scope_label = _udemy_delivery_access_scope(db, user, delivery_id)
    return UdemyProgressService(db).list_students(
        delivery_id,
        allowed_class_ids=allowed_class_ids,
        q=q, class_id=class_id, status_filter=status_filter,
        page=page, page_size=page_size, sort_by=sort_by, sort_dir=sort_dir,
    )


@router.get('/subject-deliveries/{delivery_id}/udemy-progress/export.xlsx')
def export_udemy_progress(
    delivery_id: str,
    q: str | None = None,
    class_id: str | None = None,
    status_filter: str | None = Query(None, alias='status'),
    user: UserContext = Depends(_require_academic_view_permission),
    db: Session = Depends(get_db),
):
    delivery, allowed_class_ids, scope_label = _udemy_delivery_access_scope(db, user, delivery_id)
    service = UdemyProgressService(db)
    count = service.list_students(
        delivery_id, allowed_class_ids=allowed_class_ids, q=q, class_id=class_id,
        status_filter=status_filter, page=1, page_size=1,
    )['total']
    if count > int(settings.academic_udemy_sync_export_max_rows):
        raise HTTPException(
            status_code=409,
            detail={
                'code': 'UDEMY_EXPORT_REQUIRES_BACKGROUND_JOB',
                'message': 'Báo cáo lớn phải chạy bằng job export nền.',
                'row_count': count,
                'sync_limit': int(settings.academic_udemy_sync_export_max_rows),
            },
        )
    subject = db.get(AcademicSubject, delivery.subject_id)
    raw = service.export_workbook(
        delivery_id, allowed_class_ids=allowed_class_ids, scope_label=scope_label,
        q=q, class_id=class_id, status_filter=status_filter,
        max_rows=int(settings.academic_udemy_sync_export_max_rows),
    )
    code = safe_upload_filename((subject.subject_code if subject else 'udemy') or 'udemy')
    return StreamingResponse(
        BytesIO(raw),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="udemy-progress-{code}.xlsx"'},
    )


@router.post('/subject-deliveries/{delivery_id}/udemy-progress/export-jobs', response_model=AcademicBulkOperationJobOut)
def create_udemy_progress_export_job(
    delivery_id: str,
    q: str | None = None,
    class_id: str | None = None,
    status_filter: str | None = Query(None, alias='status'),
    user: UserContext = Depends(_require_academic_view_permission),
    db: Session = Depends(get_db),
):
    enforce_operation_rate_limit(
        namespace='udemy-progress-export',
        actor_id=str(user.user_id or user.username or 'unknown'),
        limit=max(2, int(settings.academic_udemy_upload_rate_limit_per_minute)),
        code='UDEMY_EXPORT_RATE_LIMITED',
        message='Bạn yêu cầu xuất báo cáo quá nhanh. Vui lòng đợi rồi thử lại.',
    )
    actor_id = str(user.user_id or user.username or '')
    delivery, allowed_class_ids, scope_label = _udemy_delivery_access_scope(db, user, delivery_id)
    service = UdemyProgressService(db)
    count = int(service.list_students(
        delivery_id, allowed_class_ids=allowed_class_ids, q=q, class_id=class_id,
        status_filter=status_filter, page=1, page_size=1,
    )['total'])
    filter_payload = json_safe_value({
        'delivery_id': delivery_id,
        'q': str(q or '').strip() or None,
        'class_id': class_id,
        'status': status_filter,
        'allowed_class_ids': sorted(allowed_class_ids) if allowed_class_ids is not None else None,
        'scope_label': scope_label,
        'row_count': count,
    })
    signature = hashlib.sha256(json.dumps(filter_payload, sort_keys=True, ensure_ascii=False).encode('utf-8')).hexdigest()
    active_candidates = db.query(AcademicBulkOperationJob).filter(
        AcademicBulkOperationJob.job_type == 'udemy_progress_export',
        AcademicBulkOperationJob.requested_by == actor_id,
        AcademicBulkOperationJob.status.in_(['queued', 'running']),
    ).order_by(AcademicBulkOperationJob.created_at.desc()).limit(20).all()
    for existing in active_candidates:
        request = existing.request_json if isinstance(existing.request_json, dict) else {}
        if request.get('export_signature') == signature:
            return existing
    job = AcademicBulkOperationJob(
        job_type='udemy_progress_export',
        status='queued',
        term_id=delivery.term_id,
        branch=str(delivery.branch or 'poly').lower(),
        campus=None,
        requested_by=actor_id,
        progress_current=0,
        progress_total=100,
        progress_label=f'Đã xếp hàng export Udemy ({count} dòng)',
        request_json=json_safe_value({
            **filter_payload,
            'export_signature': signature,
            'requester_context': _requester_context_json(user),
            'policy_version': 'udemy-progress-export/batch35',
        }),
        result_json={},
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    from app.worker import academic_udemy_progress_export_task
    academic_udemy_progress_export_task.delay(job.id)
    log_audit(
        db,
        action='academic.udemy_progress.export.enqueue',
        status='queued',
        message=f'Đã xếp hàng export Udemy {count} dòng.',
        user=user,
        target_type='academic_bulk_operation_job',
        target_id=job.id,
        metadata=json_safe_value({'delivery_id': delivery_id, 'row_count': count, 'export_signature': signature}),
    )
    return job


@router.get('/udemy/progress/export-jobs/{job_id}/download')
def download_udemy_progress_export_job(
    job_id: str,
    user: UserContext = Depends(_require_academic_view_permission),
    db: Session = Depends(get_db),
):
    job = db.get(AcademicBulkOperationJob, job_id)
    if not job or job.job_type != 'udemy_progress_export':
        raise HTTPException(status_code=404, detail='Không tìm thấy job export Udemy.')
    request = job.request_json if isinstance(job.request_json, dict) else {}
    delivery_id = str(request.get('delivery_id') or '')
    _delivery, current_allowed_ids, _scope_label = _udemy_delivery_access_scope(db, user, delivery_id)
    requested_ids = request.get('allowed_class_ids')
    if current_allowed_ids is not None and str(job.requested_by or '') != str(user.user_id or user.username or ''):
        raise HTTPException(status_code=403, detail='Báo cáo này thuộc job của người dùng khác trong cùng phạm vi.')
    if requested_ids is None:
        if current_allowed_ids is not None:
            raise HTTPException(status_code=403, detail='Quyền hiện tại không còn cho phép tải báo cáo toàn môn.')
    elif current_allowed_ids is not None and not set(str(item) for item in requested_ids).issubset(current_allowed_ids):
        raise HTTPException(status_code=403, detail='Phạm vi quyền hiện tại không còn bao phủ báo cáo này.')
    if job.status != 'completed':
        raise HTTPException(status_code=409, detail='Job export chưa hoàn tất.')
    result = job.result_json if isinstance(job.result_json, dict) else {}
    root = Path(settings.local_storage_path or '/app/.runtime').expanduser().resolve()
    filename = safe_upload_filename(str(result.get('file_name') or ''))
    if not filename or not filename.lower().endswith('.xlsx'):
        raise HTTPException(status_code=400, detail='Tên file export không hợp lệ.')
    path = (root / 'udemy-progress-exports' / filename).resolve()
    try:
        path.relative_to(root / 'udemy-progress-exports')
    except ValueError as exc:
        raise HTTPException(status_code=400, detail='Đường dẫn file export không hợp lệ.') from exc
    if not path.is_file():
        raise HTTPException(status_code=410, detail='File export đã hết thời gian lưu. Hãy tạo lại báo cáo.')
    return FileResponse(
        path,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        filename=filename,
    )


@router.get('/subject-deliveries/{delivery_id}/udemy-plan', response_model=UdemyPlanDetailOut)
def get_udemy_plan_detail(
    delivery_id: str,
    user: UserContext = Depends(_require_academic_catalog_admin),
    db: Session = Depends(get_db),
):
    del user
    return UdemyPlanService(db).get_plan_detail(delivery_id)


@router.get('/subject-deliveries/{delivery_id}/udemy-plan/history', response_model=list[UdemySubjectPlanOut])
def get_udemy_plan_history(
    delivery_id: str,
    user: UserContext = Depends(_require_academic_catalog_admin),
    db: Session = Depends(get_db),
):
    del user
    return UdemyPlanService(db).list_plan_history(delivery_id)


@router.post('/subject-deliveries/{delivery_id}/udemy-plan', response_model=UdemyPlanMutationOut)
def create_manual_udemy_plan_version(
    delivery_id: str,
    payload: UdemyPlanVersionCreateIn,
    user: UserContext = Depends(_require_academic_catalog_admin),
    db: Session = Depends(get_db),
):
    service = UdemyPlanService(db)
    plan = service.create_version(
        delivery_id=delivery_id,
        item_count=payload.item_count,
        milestones=[item.model_dump(mode='json') for item in payload.milestones],
        actor=user.user_id or user.username,
        source='manual',
        note=payload.note,
        metadata={'policy_version': 'udemy-plan-management/batch32'},
    )
    message = f'Đã lưu phiên bản kế hoạch Udemy v{plan.version}. Phiên bản cũ vẫn được giữ trong lịch sử.'
    log_audit(
        db,
        action='academic.udemy_plan.version.create',
        status='success',
        message=message,
        user=user,
        target_type='udemy_subject_plan',
        target_id=plan.id,
        metadata={'delivery_id': delivery_id, 'version': plan.version, 'item_count': plan.item_count},
    )
    return {'ok': True, 'message': message, 'created_count': 1, 'plans': [service.serialize_plan(plan)]}


@router.get('/subjects', response_model=list[AcademicSubjectOut])
def list_subjects(
    term_id: str | None = None,
    block_id: str | None = None,
    search: str | None = None,
    branch: str | None = None,
    user: UserContext = Depends(_require_academic_view_permission),
    db: Session = Depends(get_db),
):
    return AcademicService(db).list_subjects(term_id=term_id, block_id=block_id, search=search, branch=branch)


@router.get('/teacher/classes', response_model=AcademicClassListOut)
def list_teacher_classes(
    term_id: str | None = None,
    block_id: str | None = None,
    subject_id: str | None = None,
    campus: str | None = None,
    branch: str | None = None,
    search: str | None = None,
    learning_status: str | None = Query(None, description='Lọc trạng thái học tập/cảnh báo'),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    user: UserContext = Depends(_require_academic_view_permission),
    db: Session = Depends(get_db),
):
    return AcademicService(db).list_teacher_classes(
        user,
        term_id=term_id,
        block_id=block_id,
        subject_id=subject_id,
        campus=campus,
        branch=branch,
        search=search,
        learning_status=learning_status,
        page=page,
        page_size=page_size,
    )



@router.get('/teacher/subjects', response_model=AcademicSubjectManagementListOut)
def list_teacher_subjects(
    term_id: str | None = None,
    branch: str | None = None,
    campus: str | None = None,
    search: str | None = None,
    learning_status: str | None = Query(None, description='Lọc trạng thái học tập/cảnh báo'),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    user: UserContext = Depends(_require_academic_view_permission),
    db: Session = Depends(get_db),
):
    return AcademicService(db).list_teacher_subjects(
        user, term_id=term_id, branch=branch, campus=campus, search=search, learning_status=learning_status, page=page, page_size=page_size
    )



@router.post('/subjects/course-mapping/auto-all-sync/jobs', response_model=AcademicSubjectAutoMapAllSyncOut)
def auto_map_all_subject_courses_and_enqueue_sync_jobs(
    payload: AcademicSubjectAutoMapAllSyncIn,
    user: UserContext = Depends(_require_academic_sync_permission),
    db: Session = Depends(get_db),
):
    """Create a durable worker-backed bulk auto-map + full CMS sync job.

    Auto-map all can touch thousands of subjects/classes and can enqueue many
    child class-sync jobs. It must behave like question-generation jobs: return
    immediately, survive F5, and be visible to other operators in /jobs.
    """
    if not payload.term_id:
        raise HTTPException(status_code=422, detail='Thiếu học kỳ để auto map tất cả')
    branch_value = payload.branch.strip().lower() if payload.branch and payload.branch.strip() else None
    campus_value = payload.campus.strip().lower() if payload.campus and payload.campus.strip() else None
    rbac = BusinessRBACService(db)
    rbac.ensure_requested_campus_filter_allowed(
        user,
        campus_value,
        require_filter_when_scoped=True,
        action='chạy Auto map tất cả',
    )
    campus_scope = rbac.campus_scope_for_user(user)
    request_json = json_safe_value({
        'term_id': payload.term_id,
        'branch': branch_value,
        'campus': campus_value,
        'search': payload.search,
        'learning_status': payload.learning_status,
        'force': bool(payload.force),
        'limit': max(1, min(500, int(payload.limit or 500))),
        'mode': payload.mode,
        'sync_learning': bool(payload.sync_learning),
        'max_classes': max(1, min(5000, int(payload.max_classes or 3000))),
        'requester_context': _requester_context_json(user),
        'scope_enforced_in_worker': True,
        'approved_campus_codes': campus_scope.get('campus_codes') or ([] if campus_scope.get('unrestricted') else []),
        'campus_scope_unrestricted': bool(campus_scope.get('unrestricted')),
        'scope_enforced_by_backend': True,
    })

    active_candidates = (
        db.query(AcademicBulkOperationJob)
        .filter(
            AcademicBulkOperationJob.job_type == 'subject_auto_map_all_sync',
            AcademicBulkOperationJob.term_id == payload.term_id,
            AcademicBulkOperationJob.branch == branch_value,
            AcademicBulkOperationJob.campus == campus_value,
            AcademicBulkOperationJob.status.in_(['queued', 'running']),
        )
        .order_by(AcademicBulkOperationJob.created_at.desc())
        .limit(20)
        .all()
    )
    active = None
    for candidate in active_candidates:
        candidate_request = candidate.request_json if isinstance(candidate.request_json, dict) else {}
        if (
            (candidate_request.get('search') or None) == (payload.search or None)
            and (candidate_request.get('learning_status') or None) == (payload.learning_status or None)
            and bool(candidate_request.get('force', True)) == bool(payload.force)
            and bool(candidate_request.get('sync_learning', True)) == bool(payload.sync_learning)
        ):
            active = candidate
            break
    if active and not rbac.can_access_academic_scope(user, campus=active.campus, requested_by=active.requested_by, request_json=active.request_json if isinstance(active.request_json, dict) else {}):
        active = None
    if active:
        message = 'Đang có job Auto map tất cả chạy cho bộ lọc này. Hệ thống dùng lại job hiện có, F5 không làm mất tiến trình.'
        return {
            'ok': True,
            'message': message,
            'job_id': active.id,
            'status': active.status,
            'term_id': payload.term_id,
            'branch': branch_value,
            'campus': campus_value,
            'subject_total': int((active.result_json or {}).get('subject_total') or 0) if isinstance(active.result_json, dict) else 0,
            'subject_mapped': int((active.result_json or {}).get('subject_mapped') or 0) if isinstance(active.result_json, dict) else 0,
            'subject_already_mapped': int((active.result_json or {}).get('subject_already_mapped') or 0) if isinstance(active.result_json, dict) else 0,
            'subject_failed': int((active.result_json or {}).get('subject_failed') or 0) if isinstance(active.result_json, dict) else 0,
            'class_total': int((active.result_json or {}).get('class_total') or 0) if isinstance(active.result_json, dict) else 0,
            'jobs_queued': int((active.result_json or {}).get('jobs_queued') or 0) if isinstance(active.result_json, dict) else 0,
            'jobs_reused': int((active.result_json or {}).get('jobs_reused') or 0) if isinstance(active.result_json, dict) else 0,
            'jobs_skipped': int((active.result_json or {}).get('jobs_skipped') or 0) if isinstance(active.result_json, dict) else 0,
            'capped': bool((active.result_json or {}).get('capped')) if isinstance(active.result_json, dict) else False,
            'subject_results': [],
            'job_ids': [],
        }

    # Snapshot the exact class scope that was authorized at enqueue time.
    # The worker will only enqueue child jobs for these class IDs, so a replayed
    # or modified parent job cannot widen itself to other campuses/subjects.
    try:
        approved_preview = AcademicService(db).auto_map_subject_courses_for_filter(
            user,
            term_id=payload.term_id,
            branch=branch_value,
            campus=campus_value,
            search=payload.search,
            learning_status=payload.learning_status,
            max_classes=max(1, min(5000, int(payload.max_classes or 3000))),
            dry_run=True,
        )
        request_json['approved_class_ids'] = list(approved_preview.get('class_ids') or [])
        request_json['approved_class_total'] = int(approved_preview.get('class_total') or 0)
        request_json['approved_subject_ids'] = list(approved_preview.get('subject_ids') or [])
        request_json['approved_campus_codes_from_preview'] = sorted({str(item.get('campus') or '').strip().lower() for item in (approved_preview.get('classes') or []) if isinstance(item, dict) and str(item.get('campus') or '').strip()})
    except Exception as exc:
        raise HTTPException(status_code=403, detail=f'Không thể xác định phạm vi được phân quyền cho Auto map tất cả: {exc}')

    job = AcademicBulkOperationJob(
        job_type='subject_auto_map_all_sync',
        status='queued',
        term_id=payload.term_id,
        branch=branch_value,
        campus=campus_value,
        requested_by=user.user_id,
        progress_current=0,
        progress_total=100,
        progress_label='Đã đưa Auto map tất cả vào hàng đợi',
        request_json=request_json,
        result_json={},
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    from app.worker import academic_subject_auto_map_all_sync_task
    academic_subject_auto_map_all_sync_task.delay(job.id)
    message = 'Đã tạo job Auto map tất cả. Bạn có thể F5 hoặc chuyển màn hình; tiến trình vẫn chạy trong worker và hiển thị ở /jobs.'
    log_audit(
        db,
        action='academic.subject_course_mapping.auto_all_sync_job.enqueue',
        status='queued',
        message=message,
        user=user,
        target_type='academic_bulk_operation_job',
        target_id=job.id,
        metadata=json_safe_value(request_json),
    )
    return {
        'ok': True,
        'message': message,
        'job_id': job.id,
        'status': job.status,
        'term_id': payload.term_id,
        'branch': branch_value,
        'campus': campus_value,
        'subject_total': 0,
        'subject_mapped': 0,
        'subject_already_mapped': 0,
        'subject_failed': 0,
        'class_total': 0,
        'jobs_queued': 0,
        'jobs_reused': 0,
        'jobs_skipped': 0,
        'capped': False,
        'subject_results': [],
        'job_ids': [],
    }


@router.get('/bulk-operation-jobs', response_model=list[AcademicBulkOperationJobOut])
def list_academic_bulk_operation_jobs(
    status_filter: str | None = Query(None, alias='status'),
    limit: int = Query(50, ge=1, le=100),
    user: UserContext = Depends(_require_academic_view_permission),
    db: Session = Depends(get_db),
):
    query = db.query(AcademicBulkOperationJob)
    if status_filter and status_filter != 'all':
        if status_filter == 'active':
            query = query.filter(AcademicBulkOperationJob.status.in_(['queued', 'running']))
        else:
            query = query.filter(AcademicBulkOperationJob.status == status_filter)
    service = BusinessRBACService(db)
    candidates = query.order_by(AcademicBulkOperationJob.created_at.desc()).limit(max(limit * 5, 100)).all()
    visible: list[AcademicBulkOperationJob] = []
    for job in candidates:
        if service.can_access_academic_scope(user, campus=job.campus, requested_by=job.requested_by, request_json=job.request_json if isinstance(job.request_json, dict) else {}):
            visible.append(job)
        if len(visible) >= limit:
            break
    return visible


@router.get('/bulk-operation-jobs/{job_id}', response_model=AcademicBulkOperationJobOut)
def get_academic_bulk_operation_job(
    job_id: str,
    user: UserContext = Depends(_require_academic_view_permission),
    db: Session = Depends(get_db),
):
    job = db.get(AcademicBulkOperationJob, job_id)
    if not job:
        raise HTTPException(status_code=404, detail='Không tìm thấy job xử lý hàng loạt')
    BusinessRBACService(db).require_academic_scope(user, campus=job.campus, requested_by=job.requested_by, request_json=job.request_json if isinstance(job.request_json, dict) else {}, action='xem job xử lý hàng loạt')
    return job

@router.get('/subjects/{subject_id}/classes', response_model=AcademicClassListOut)
def list_subject_classes(
    subject_id: str,
    term_id: str | None = None,
    block_id: str | None = None,
    campus: str | None = None,
    branch: str | None = None,
    search: str | None = None,
    learning_status: str | None = Query(None, description='Lọc trạng thái học tập/cảnh báo'),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    user: UserContext = Depends(require_permission('view_questions')),
    db: Session = Depends(get_db),
):
    service = AcademicService(db)
    service.assert_can_access_subject(user, subject_id)
    return service.list_teacher_classes(
        user, term_id=term_id, block_id=block_id, subject_id=subject_id, campus=campus, branch=branch, search=search, learning_status=learning_status, page=page, page_size=page_size
    )


@router.post('/subjects/{subject_id}/course-mapping/auto', response_model=AcademicSubjectCourseAutoMapOut)
def auto_map_subject_course(
    subject_id: str,
    term_id: str = Query(...),
    branch: str | None = Query(None),
    user: UserContext = Depends(_require_academic_sync_permission),
    db: Session = Depends(get_db),
):
    result = AcademicService(db).auto_map_subject_course(user, term_id=term_id, subject_id=subject_id, branch=branch)
    log_audit(
        db,
        action='academic.subject_course_mapping.auto',
        status='success' if result.get('ok') else 'failed',
        message=result.get('message', ''),
        user=user,
        course_id=(result.get('mapping') or {}).get('openedx_course_id') if isinstance(result.get('mapping'), dict) else None,
        target_type='academic_subject',
        target_id=subject_id,
        metadata={'term_id': term_id, 'branch': branch, 'status': result.get('status')},
    )
    return result


@router.get('/course-mappings', response_model=AcademicCourseMappingListOut)
def list_academic_course_mappings(
    term_id: str | None = None,
    block_id: str | None = None,
    subject_id: str | None = None,
    search: str | None = None,
    active: bool | None = True,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    user: UserContext = Depends(require_permission('manage_settings')),
    db: Session = Depends(get_db),
):
    _require_academic_admin(db, user)
    return AcademicService(db).list_course_mappings(user, term_id=term_id, block_id=block_id, subject_id=subject_id, search=search, active=active, page=page, page_size=page_size)


@router.post('/course-mappings/validate', response_model=AcademicCourseMappingValidationOut)
def validate_academic_course_mapping(
    payload: AcademicCourseMappingValidateIn,
    user: UserContext = Depends(require_permission('manage_settings')),
    db: Session = Depends(get_db),
):
    _require_academic_admin(db, user)
    result = AcademicService(db).validate_course_mapping_payload(**payload.model_dump())
    log_audit(
        db,
        action='academic.course_mapping.validate',
        status='success' if result.get('ok') else 'failed',
        error_type=None if result.get('ok') else AuditErrorType.VALIDATION_ERROR,
        message=result.get('message', ''),
        user=user,
        course_id=payload.openedx_course_id,
        target_type='academic_course_mapping',
        metadata=result,
    )
    return result


@router.post('/course-mappings', response_model=AcademicCourseMappingOut)
def create_academic_course_mapping(
    payload: AcademicCourseMappingCreateIn,
    user: UserContext = Depends(require_permission('manage_settings')),
    db: Session = Depends(get_db),
):
    _require_academic_admin(db, user)
    try:
        item = AcademicService(db).create_or_update_course_mapping(user, payload.model_dump())
        log_audit(db, action='academic.course_mapping.save', status='success', message='Lưu mapping AP ↔ Open edX course thành công', user=user, course_id=item.get('openedx_course_id'), target_type='academic_course_mapping', target_id=item.get('id'), metadata={'validation_status': item.get('validation_status')})
        return item
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        log_audit(db, action='academic.course_mapping.save', status='failed', error_type=AuditErrorType.VALIDATION_ERROR, message='Không thể hoàn tất thao tác vận hành đào tạo.', user=user, target_type='academic_course_mapping')
        raise public_http_exception(
            status_code=400,
            code='ACADEMIC_MAPPING_FAILED',
            message='Không thể lưu mapping học vụ.',
            logger_name=__name__,
        ) from exc


@router.delete('/course-mappings/{mapping_id}', response_model=AcademicCourseMappingOut)
def deactivate_academic_course_mapping(
    mapping_id: str,
    user: UserContext = Depends(require_permission('manage_settings')),
    db: Session = Depends(get_db),
):
    _require_academic_admin(db, user)
    item = AcademicService(db).deactivate_course_mapping(user, mapping_id)
    log_audit(db, action='academic.course_mapping.deactivate', status='success', message='Đã tắt mapping course cấp môn/kỳ/block', user=user, course_id=item.get('openedx_course_id'), target_type='academic_course_mapping', target_id=mapping_id)
    return item


@router.get('/classes/{class_id}/course-mapping/proposal', response_model=AcademicClassCourseMappingProposalOut)
def get_class_course_mapping_proposal(
    class_id: str,
    user: UserContext = Depends(require_permission('view_questions')),
    db: Session = Depends(get_db),
):
    return AcademicService(db).class_course_mapping_proposal(user, class_id)


@router.post('/classes/{class_id}/course-mapping/validate', response_model=AcademicCourseMappingValidationOut)
def validate_class_course_mapping(
    class_id: str,
    payload: AcademicClassCourseMappingValidateIn,
    user: UserContext = Depends(require_permission('manage_settings')),
    db: Session = Depends(get_db),
):
    _require_academic_admin(db, user)
    result = AcademicService(db).validate_class_course_mapping(user, class_id, payload.model_dump())
    log_audit(db, action='academic.class_course_mapping.validate', status='success' if result.get('ok') else 'failed', error_type=None if result.get('ok') else AuditErrorType.VALIDATION_ERROR, message=result.get('message', ''), user=user, course_id=payload.openedx_course_id, target_type='academic_class', target_id=class_id, metadata=result)
    return result


@router.post('/classes/{class_id}/course-mapping', response_model=AcademicClassCourseMappingOut)
def save_class_course_mapping(
    class_id: str,
    payload: AcademicClassCourseMappingCreateIn,
    user: UserContext = Depends(require_permission('manage_settings')),
    db: Session = Depends(get_db),
):
    _require_academic_admin(db, user)
    try:
        item = AcademicService(db).create_or_update_class_course_mapping(user, class_id, payload.model_dump())
        log_audit(db, action='academic.class_course_mapping.save', status='success', message='Lưu mapping lớp AP sang Open edX course thành công', user=user, course_id=item.get('openedx_course_id'), target_type='academic_class', target_id=class_id, metadata={'validation_status': item.get('validation_status'), 'cohort': item.get('openedx_cohort_name')})
        return item
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        log_audit(db, action='academic.class_course_mapping.save', status='failed', error_type=AuditErrorType.VALIDATION_ERROR, message='Không thể hoàn tất thao tác vận hành đào tạo.', user=user, target_type='academic_class', target_id=class_id)
        raise public_http_exception(
            status_code=400,
            code='ACADEMIC_MAPPING_FAILED',
            message='Không thể lưu mapping học vụ.',
            logger_name=__name__,
        ) from exc


@router.delete('/classes/{class_id}/course-mapping', response_model=AcademicClassCourseMappingOut)
def deactivate_class_course_mapping(
    class_id: str,
    user: UserContext = Depends(require_permission('manage_settings')),
    db: Session = Depends(get_db),
):
    _require_academic_admin(db, user)
    item = AcademicService(db).deactivate_class_course_mapping(user, class_id)
    log_audit(db, action='academic.class_course_mapping.deactivate', status='success', message='Đã tắt mapping course riêng của lớp', user=user, course_id=item.get('openedx_course_id'), target_type='academic_class', target_id=class_id)
    return item


@router.get('/classes/{class_id}', response_model=AcademicClassOut)
def get_class_detail(
    class_id: str,
    user: UserContext = Depends(require_permission('view_questions')),
    db: Session = Depends(get_db),
):
    return AcademicService(db).get_class_detail(user, class_id)


@router.get('/classes/{class_id}/students', response_model=AcademicStudentListOut)
def list_class_students(
    class_id: str,
    search: str | None = None,
    learning_status: str | None = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    user: UserContext = Depends(require_permission('view_questions')),
    db: Session = Depends(get_db),
):
    return AcademicService(db).list_class_students(user, class_id, search=search, learning_status=learning_status, page=page, page_size=page_size)


@router.get('/classes/{class_id}/mapping-summary', response_model=AcademicMappingSummaryOut)
def get_class_mapping_summary(
    class_id: str,
    user: UserContext = Depends(require_permission('view_questions')),
    db: Session = Depends(get_db),
):
    return AcademicService(db).mapping_summary_for_class(user, class_id)


@router.get('/classes/{class_id}/learning-summary', response_model=AcademicLearningSummaryOut)
def get_class_learning_summary(
    class_id: str,
    user: UserContext = Depends(require_permission('view_questions')),
    db: Session = Depends(get_db),
):
    return AcademicService(db).learning_summary_for_class(user, class_id)






@router.get('/classes/{class_id}/quiz-deadline-overrides', response_model=list[AcademicQuizDeadlineOverrideOut])
def list_class_quiz_deadline_overrides(
    class_id: str,
    course_id: str | None = Query(None),
    user: UserContext = Depends(require_permission('view_questions')),
    db: Session = Depends(get_db),
):
    AcademicService(db).assert_can_access_class(user, class_id)
    query = db.query(AcademicQuizDeadlineOverride).filter(AcademicQuizDeadlineOverride.class_id == class_id)
    if course_id:
        query = query.filter((AcademicQuizDeadlineOverride.course_id == course_id) | (AcademicQuizDeadlineOverride.course_id.is_(None)))
    return query.order_by(AcademicQuizDeadlineOverride.quiz_number.asc().nullslast(), AcademicQuizDeadlineOverride.updated_at.desc()).all()


@router.put('/classes/{class_id}/quiz-deadline-overrides', response_model=list[AcademicQuizDeadlineOverrideOut])
def save_class_quiz_deadline_overrides(
    class_id: str,
    payload: AcademicQuizDeadlineOverrideBulkIn,
    user: UserContext = Depends(_require_training_write_permission),
    db: Session = Depends(get_db),
):
    AcademicService(db).assert_can_access_class(user, class_id)
    saved: list[AcademicQuizDeadlineOverride] = []
    now = datetime.utcnow()
    for item in payload.items:
        if not item.quiz_number:
            continue
        course_id = (item.course_id or '').strip() or None
        row = db.query(AcademicQuizDeadlineOverride).filter(
            AcademicQuizDeadlineOverride.class_id == class_id,
            AcademicQuizDeadlineOverride.course_id.is_(None) if course_id is None else AcademicQuizDeadlineOverride.course_id == course_id,
            AcademicQuizDeadlineOverride.quiz_number == int(item.quiz_number),
        ).first()
        if not row:
            row = AcademicQuizDeadlineOverride(
                class_id=class_id,
                course_id=course_id,
                quiz_number=int(item.quiz_number),
                created_by=user.user_id,
                created_at=now,
            )
            db.add(row)
        row.component_key = (item.component_key or '').strip() or None
        row.component_label = (item.component_label or f'Quiz {item.quiz_number}').strip()
        row.start_date = item.start_date
        row.deadline_date = item.deadline_date
        row.reason = (item.reason or '').strip()
        row.updated_by = user.user_id
        row.updated_at = now
        row.metadata_json = {'source': 'manual_ui'}
        saved.append(row)
    db.commit()
    for row in saved:
        db.refresh(row)
    log_audit(db, action='academic.quiz_deadline_override.save', status='success', message='Lưu deadline quiz thủ công thành công', user=user, target_type='academic_class', target_id=class_id, metadata={'count': len(saved)})
    return saved


@router.get('/classes/{class_id}/assignment-defense-scores', response_model=list[AcademicAssignmentDefenseScoreOut])
def list_class_assignment_defense_scores(
    class_id: str,
    course_id: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(200, ge=1, le=500),
    user: UserContext = Depends(require_permission('view_questions')),
    db: Session = Depends(get_db),
):
    AcademicService(db).assert_can_access_class(user, class_id)
    return AcademicAssignmentExternalWorkflowService(db).list_class_assignment_scores(
        class_id=class_id, course_id=course_id, page=page, page_size=page_size
    )



@router.put('/classes/{class_id}/assignment-defense-scores', response_model=list[AcademicAssignmentDefenseScoreOut])
def save_class_assignment_defense_scores(
    class_id: str,
    payload: AcademicAssignmentDefenseScoreBulkIn,
    user: UserContext = Depends(get_user_context),
    db: Session = Depends(get_db),
):
    AcademicService(db).assert_can_access_class(user, class_id)
    AcademicAssignmentExternalWorkflowService(db).reject_assignment_score_write()



@router.post('/classes/{class_id}/cms-sync-check/jobs', response_model=AcademicClassSyncJobOut)
def enqueue_class_cms_sync_check(
    class_id: str,
    payload: AcademicResolveClassUsersIn,
    user: UserContext = Depends(_require_academic_sync_permission),
    db: Session = Depends(get_db),
):
    return _enqueue_class_sync_job(db=db, user=user, class_id=class_id, job_type='cms_sync_check', force=payload.force, limit=payload.limit)


@router.post('/classes/{class_id}/cms-enrollment-sync/jobs', response_model=AcademicClassSyncJobOut)
def enqueue_class_cms_enrollment_sync(
    class_id: str,
    payload: AcademicEnrollmentSyncIn,
    user: UserContext = Depends(_require_academic_sync_permission),
    db: Session = Depends(get_db),
):
    return _enqueue_class_sync_job(db=db, user=user, class_id=class_id, job_type='cms_enrollment_sync', force=payload.force, limit=payload.limit, mode=payload.mode)


@router.post('/classes/{class_id}/learning-sync/jobs', response_model=AcademicClassSyncJobOut)
def enqueue_class_learning_sync(
    class_id: str,
    payload: AcademicLearningSyncIn,
    user: UserContext = Depends(_require_academic_sync_permission),
    db: Session = Depends(get_db),
):
    return _enqueue_class_sync_job(db=db, user=user, class_id=class_id, job_type='learning_sync', force=payload.force, limit=payload.limit)


@router.post('/classes/{class_id}/full-cms-sync/jobs', response_model=AcademicClassSyncJobOut)
def enqueue_class_full_cms_sync(
    class_id: str,
    payload: AcademicFullCmsSyncIn,
    user: UserContext = Depends(_require_academic_sync_permission),
    db: Session = Depends(get_db),
):
    return _enqueue_class_sync_job(
        db=db,
        user=user,
        class_id=class_id,
        job_type='full_cms_sync',
        force=payload.force,
        limit=payload.limit,
        mode=payload.mode,
        auto_map_course=payload.auto_map_course,
        sync_learning=payload.sync_learning,
    )



@router.get('/sync/class-jobs', response_model=list[AcademicClassSyncJobOut])
def list_recent_class_sync_jobs(
    class_id: str | None = Query(None),
    status_filter: str = Query('all', alias='status'),
    limit: int = Query(30, ge=1, le=100),
    user: UserContext = Depends(_require_academic_view_permission),
    db: Session = Depends(get_db),
):
    service = AcademicService(db)
    query = db.query(AcademicClassSyncJob)
    if class_id:
        service.assert_can_access_class(user, class_id)
        query = query.filter(AcademicClassSyncJob.class_id == class_id)
    if status_filter == 'active':
        query = query.filter(AcademicClassSyncJob.status.in_(['queued', 'running']))
    elif status_filter and status_filter != 'all':
        query = query.filter(AcademicClassSyncJob.status == status_filter)
    candidates = query.order_by(AcademicClassSyncJob.created_at.desc()).limit(max(limit, 30)).all()
    visible: list[AcademicClassSyncJob] = []
    for row in candidates:
        try:
            service.assert_can_access_class(user, row.class_id)
            visible.append(row)
        except Exception:
            continue
        if len(visible) >= limit:
            break
    return visible


@router.get('/classes/{class_id}/sync-jobs/{job_id}', response_model=AcademicClassSyncJobOut)
def get_class_sync_job(
    class_id: str,
    job_id: str,
    user: UserContext = Depends(require_permission('view_questions')),
    db: Session = Depends(get_db),
):
    AcademicService(db).assert_can_access_class(user, class_id)
    job = db.query(AcademicClassSyncJob).filter(AcademicClassSyncJob.id == job_id, AcademicClassSyncJob.class_id == class_id).first()
    if not job:
        raise HTTPException(status_code=404, detail='Không tìm thấy job đồng bộ lớp')
    return job


@router.get('/classes/{class_id}/sync-jobs', response_model=list[AcademicClassSyncJobOut])
def list_class_sync_jobs(
    class_id: str,
    limit: int = Query(10, ge=1, le=50),
    user: UserContext = Depends(require_permission('view_questions')),
    db: Session = Depends(get_db),
):
    AcademicService(db).assert_can_access_class(user, class_id)
    return db.query(AcademicClassSyncJob).filter(AcademicClassSyncJob.class_id == class_id).order_by(AcademicClassSyncJob.created_at.desc()).limit(limit).all()

@router.post('/classes/{class_id}/full-cms-sync', response_model=AcademicFullCmsSyncOut)
def sync_class_full_cms_flow(
    class_id: str,
    payload: AcademicFullCmsSyncIn,
    user: UserContext = Depends(_require_academic_sync_permission),
    db: Session = Depends(get_db),
):
    service = AcademicService(db)
    try:
        result = service.sync_class_full_cms_flow(
            user,
            class_id,
            force=payload.force,
            limit=payload.limit,
            mode=payload.mode,
            auto_map_course=payload.auto_map_course,
            sync_learning=payload.sync_learning,
        )
        log_audit(
            db,
            action='academic.full_cms_sync.class',
            status='success',
            message=result.get('message', 'Đồng bộ full CMS hoàn tất'),
            user=user,
            target_type='academic_class',
            target_id=class_id,
            course_id=result.get('openedx_course_id'),
            metadata={'counts': result.get('counts', {}), 'status': result.get('status')},
        )
        return result
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        log_audit(db, action='academic.full_cms_sync.class', status='failed', error_type=AuditErrorType.EXTERNAL_SERVICE_ERROR, message='Không thể hoàn tất thao tác vận hành đào tạo.', user=user, target_type='academic_class', target_id=class_id)
        raise HTTPException(status_code=502, detail=_safe_error_message('academic_external_sync_failed')) from exc


@router.post('/classes/{class_id}/cms-enrollment-sync', response_model=AcademicEnrollmentSyncOut)
def sync_class_cms_enrollment(
    class_id: str,
    payload: AcademicEnrollmentSyncIn,
    user: UserContext = Depends(_require_academic_sync_permission),
    db: Session = Depends(get_db),
):
    service = AcademicService(db)
    try:
        result = service.sync_class_course_enrollment(user, class_id, force=payload.force, limit=payload.limit, mode=payload.mode)
        log_audit(
            db,
            action='academic.cms_enrollment_sync.class',
            status='success',
            message='Tự enrollment sinh viên đã đồng bộ CMS vào Course CMS thành công',
            user=user,
            target_type='academic_class',
            target_id=class_id,
            metadata={'counts': result.get('counts', {}), 'updated': result.get('updated', 0), 'openedx_course_id': result.get('openedx_course_id')},
        )
        return result
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        log_audit(
            db,
            action='academic.cms_enrollment_sync.class',
            status='failed',
            error_type=AuditErrorType.EXTERNAL_SERVICE_ERROR,
            message='Không thể hoàn tất thao tác vận hành đào tạo.',
            user=user,
            target_type='academic_class',
            target_id=class_id,
        )
        raise HTTPException(status_code=502, detail=_safe_error_message('academic_external_sync_failed')) from exc

@router.post('/classes/{class_id}/learning-sync', response_model=AcademicLearningSyncOut)
def sync_class_learning_insight(
    class_id: str,
    payload: AcademicLearningSyncIn,
    user: UserContext = Depends(_require_academic_sync_permission),
    db: Session = Depends(get_db),
):
    service = AcademicService(db)
    try:
        result = service.sync_class_learning_insight(user, class_id, force=payload.force, limit=payload.limit)
        log_audit(
            db,
            action='academic.learning_sync.class',
            status='success',
            message='Cập nhật tiến độ/điểm CMS cho lớp thành công',
            user=user,
            target_type='academic_class',
            target_id=class_id,
            metadata={'counts': result.get('counts', {}), 'updated': result.get('updated', 0), 'openedx_course_id': result.get('openedx_course_id')},
        )
        return result
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        log_audit(
            db,
            action='academic.learning_sync.class',
            status='failed',
            error_type=AuditErrorType.EXTERNAL_SERVICE_ERROR,
            message='Không thể hoàn tất thao tác vận hành đào tạo.',
            user=user,
            target_type='academic_class',
            target_id=class_id,
        )
        raise HTTPException(status_code=502, detail=_safe_error_message('academic_external_sync_failed')) from exc


def _run_class_cms_sync_check(class_id: str, payload: AcademicResolveClassUsersIn, user: UserContext, db: Session) -> dict:
    service = AcademicService(db)
    try:
        result = service.resolve_class_openedx_users(user, class_id, force=payload.force, limit=payload.limit)
        log_audit(
            db,
            action='academic.cms_sync_check.class',
            status='success',
            message='Kiểm tra đồng bộ CMS theo AP username thành công',
            user=user,
            target_type='academic_class',
            target_id=class_id,
            metadata={'counts': result.get('counts', {}), 'updated': result.get('updated', 0)},
        )
        return result
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        log_audit(
            db,
            action='academic.cms_sync_check.class',
            status='failed',
            error_type=AuditErrorType.EXTERNAL_SERVICE_ERROR,
            message='Không thể hoàn tất thao tác vận hành đào tạo.',
            user=user,
            target_type='academic_class',
            target_id=class_id,
        )
        raise HTTPException(status_code=502, detail=_safe_error_message('academic_external_sync_failed')) from exc


@router.post('/classes/{class_id}/cms-sync-check', response_model=AcademicMappingResolveOut)
def check_class_cms_sync(
    class_id: str,
    payload: AcademicResolveClassUsersIn,
    user: UserContext = Depends(_require_academic_sync_permission),
    db: Session = Depends(get_db),
):
    return _run_class_cms_sync_check(class_id, payload, user, db)


@router.post('/classes/{class_id}/resolve-openedx-users', response_model=AcademicMappingResolveOut)
def resolve_class_openedx_users_legacy_alias(
    class_id: str,
    payload: AcademicResolveClassUsersIn,
    user: UserContext = Depends(_require_academic_sync_permission),
    db: Session = Depends(get_db),
):
    # Backward-compatible alias. UI and docs use /cms-sync-check from v25.9.16.2.23.
    return _run_class_cms_sync_check(class_id, payload, user, db)


@router.get('/classes/{class_id}/identity-reconciliation', response_model=AcademicIdentityReconciliationOut)
def get_class_identity_reconciliation(
    class_id: str,
    status_filter: str | None = Query('all', description='ALL/OK/LEGACY_AP_USERNAME/MISSING_MAPPING/MISSING_ROLLNUMBER/DUPLICATE_ROLLNUMBER/CMS_USERNAME_MISMATCH'),
    page: int = Query(1, ge=1),
    page_size: int = Query(200, ge=1, le=500),
    user: UserContext = Depends(_require_academic_sync_permission),
    db: Session = Depends(get_db),
):
    """Dry-run RollNumber identity audit for a class.

    This endpoint is intentionally read-only.  It helps admins identify legacy
    CMS users created from AP username/email before the RollNumber username
    policy is used for broad enrollment and learning sync.
    """
    return AcademicService(db).identity_reconciliation_for_class(
        user,
        class_id,
        status_filter=status_filter,
        page=page,
        page_size=page_size,
    )


@router.post('/classes/{class_id}/identity-reconciliation/uat-cleanup', response_model=AcademicIdentityCleanupOut)
def cleanup_class_identity_reconciliation_uat(
    class_id: str,
    payload: AcademicIdentityCleanupIn,
    user: UserContext = Depends(_require_academic_sync_permission),
    db: Session = Depends(get_db),
):
    """UAT-only cleanup for wrong RollNumber identity mappings.

    This is intentionally explicit and guarded. It deletes AI Server mapping rows
    and stale class learning snapshots that point to legacy AP usernames so the
    next CMS sync can recreate/check users by RollNumber. It does not delete
    Open edX Django users.
    """
    return AcademicService(db).cleanup_identity_reconciliation_for_class(user, class_id, payload.model_dump())


@router.get('/identity/rollnumber-migration', response_model=AcademicIdentityMigrationOut)
def get_rollnumber_identity_migration_report(
    class_id: str | None = Query(None),
    term_id: str | None = Query(None),
    campus: str | None = Query(None),
    branch: str | None = Query(None),
    subject_id: str | None = Query(None),
    status_filter: str | None = Query('all', description='ALL/OK/BLOCKERS/WARNINGS/LEGACY_AP_USERNAME/MISSING_MAPPING/...'),
    page: int = Query(1, ge=1),
    page_size: int = Query(200, ge=1, le=500),
    user: UserContext = Depends(_require_academic_sync_permission),
    db: Session = Depends(get_db),
):
    """Read-only RollNumber identity migration assistant.

    This endpoint is for UAT/pilot migration planning only.  It scans AP roster
    rows under the caller scope, compares AP username/email with RollNumber-based
    canonical CMS username, and returns blocker/warning rows before broad CMS sync.
    It never mutates mapping rows or Open edX users.
    """
    return AcademicService(db).rollnumber_identity_migration_report(
        user,
        class_id=class_id,
        term_id=term_id,
        campus=campus,
        branch=branch,
        subject_id=subject_id,
        status_filter=status_filter,
        page=page,
        page_size=page_size,
    )


@router.post('/openedx-user-mappings/import', response_model=AcademicManualMappingImportOut)
def import_openedx_user_mappings(
    payload: AcademicManualMappingImportIn,
    user: UserContext = Depends(require_permission('manage_settings')),
    db: Session = Depends(get_db),
):
    _require_academic_admin(db, user)
    result = AcademicService(db).import_openedx_user_mappings([item.model_dump() for item in payload.records], requested_by=user.user_id)
    log_audit(
        db,
        action='academic.openedx_user_mapping.import',
        status='success',
        message='Import mapping AP username sang Open edX user thành công',
        user=user,
        target_type='openedx_user_mappings',
        target_id='bulk',
        metadata={'counters': result.get('counters', {}), 'total': result.get('total', 0)},
    )
    return result



@router.get('/campuses', response_model=list[AcademicCampusOut])
def list_academic_campuses(
    branch: str | None = Query('poly'),
    active: bool | None = True,
    user: UserContext = Depends(_require_academic_view_permission),
    db: Session = Depends(get_db),
):
    service = AcademicService(db)
    decision = service.access_decision(user)
    query = db.query(AcademicCampus)
    if branch:
        query = query.filter(AcademicCampus.branch == branch)
    if active is not None:
        query = query.filter(AcademicCampus.active.is_(active))
    if not decision.unrestricted:
        access_conditions = []
        if decision.teacher_ids:
            teacher_campuses = db.query(AcademicClass.campus).join(
                AcademicTeacherAssignment, AcademicTeacherAssignment.class_id == AcademicClass.id
            ).filter(AcademicTeacherAssignment.teacher_id.in_(decision.teacher_ids))
            if branch:
                teacher_campuses = teacher_campuses.filter(AcademicClass.branch == branch)
            access_conditions.append(AcademicCampus.campus_code.in_(teacher_campuses.distinct()))
        if decision.subject_codes:
            subject_campuses = db.query(AcademicClass.campus).join(
                AcademicSubject, AcademicSubject.id == AcademicClass.subject_id
            ).filter(func.lower(AcademicSubject.subject_code).in_(decision.subject_codes))
            if branch:
                subject_campuses = subject_campuses.filter(AcademicClass.branch == branch)
            access_conditions.append(AcademicCampus.campus_code.in_(subject_campuses.distinct()))
        if decision.campus_codes:
            access_conditions.append(func.lower(AcademicCampus.campus_code).in_(decision.campus_codes))
        if not access_conditions:
            query = query.filter(False)
        else:
            query = query.filter(or_(*access_conditions))
    return query.order_by(AcademicCampus.sort_order.asc(), AcademicCampus.campus_code.asc()).all()


@router.post('/campuses', response_model=AcademicCampusOut)
def upsert_academic_campus(
    payload: AcademicCampusUpsertIn,
    user: UserContext = Depends(_require_academic_catalog_admin),
    db: Session = Depends(get_db),
):
    code = (payload.campus_code or '').strip().lower()
    branch = (payload.branch or 'poly').strip().lower()
    if not code:
        raise HTTPException(status_code=400, detail='Thiếu mã cơ sở AP')
    campus = db.query(AcademicCampus).filter(AcademicCampus.campus_code == code, AcademicCampus.branch == branch).first()
    if not campus:
        campus = AcademicCampus(campus_code=code, branch=branch, created_at=func.now(), updated_at=func.now())
        db.add(campus)
    campus.campus_name = payload.campus_name.strip() or code.upper()
    campus.active = payload.active
    campus.sort_order = payload.sort_order
    campus.metadata_json = {**(campus.metadata_json or {}), 'source': 'manual_ui', 'updated_from_ui': True}
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail={'code': 'ACADEMIC_CAMPUS_CONFLICT', 'message': 'Mã cơ sở đã tồn tại trong hệ này.'},
        ) from exc
    db.refresh(campus)
    log_audit(db, action='academic.campus.upsert', status='success', message='Lưu cơ sở AP thành công', user=user, target_type='academic_campus', target_id=campus.id, metadata={'campus_code': code, 'branch': branch})
    return campus


@router.patch('/campuses/{campus_id}', response_model=AcademicCampusOut)
def update_academic_campus(
    campus_id: str,
    payload: AcademicCampusUpdateIn,
    user: UserContext = Depends(_require_academic_catalog_admin),
    db: Session = Depends(get_db),
):
    campus = db.query(AcademicCampus).filter(AcademicCampus.id == campus_id).first()
    if not campus:
        raise HTTPException(status_code=404, detail='Không tìm thấy cơ sở')
    old_code = str(campus.campus_code or '')
    old_branch = str(campus.branch or '')
    code = (payload.campus_code if payload.campus_code is not None else campus.campus_code or '').strip().lower()
    branch = (payload.branch if payload.branch is not None else campus.branch or 'poly').strip().lower()
    if not code:
        raise HTTPException(status_code=422, detail='Thiếu mã cơ sở AP')
    conflict = db.query(AcademicCampus).filter(
        AcademicCampus.campus_code == code,
        AcademicCampus.branch == branch,
        AcademicCampus.id != campus.id,
    ).first()
    if conflict:
        raise HTTPException(
            status_code=409,
            detail={'code': 'ACADEMIC_CAMPUS_CONFLICT', 'message': 'Mã cơ sở đã tồn tại trong hệ này.'},
        )
    identity_changed = code != old_code.lower() or branch != old_branch.lower()
    if identity_changed:
        linked_classes = int(db.query(func.count(AcademicClass.id)).filter(
            func.lower(AcademicClass.campus) == old_code.lower(),
            func.lower(func.coalesce(AcademicClass.branch, 'poly')) == (old_branch or 'poly').lower(),
        ).scalar() or 0)
        linked_roles = int(db.query(func.count(UserRoleAssignment.id)).filter(
            UserRoleAssignment.scope_type == 'CAMPUS',
            func.lower(UserRoleAssignment.scope_id) == old_code.lower(),
            UserRoleAssignment.revoked_at.is_(None),
        ).scalar() or 0)
        if linked_classes or linked_roles:
            raise HTTPException(
                status_code=409,
                detail={
                    'code': 'ACADEMIC_CAMPUS_IDENTITY_IN_USE',
                    'message': 'Không đổi mã hoặc hệ của cơ sở đang được sử dụng. Bạn vẫn có thể sửa tên, trạng thái và thứ tự.',
                    'details': {'linked_classes': linked_classes, 'active_role_assignments': linked_roles},
                },
            )
    campus.campus_code = code
    campus.branch = branch
    if payload.campus_name is not None:
        campus.campus_name = payload.campus_name.strip() or code.upper()
    if payload.active is not None:
        campus.active = payload.active
    if payload.sort_order is not None:
        campus.sort_order = payload.sort_order
    campus.metadata_json = {
        **(campus.metadata_json or {}),
        'source': 'manual_ui',
        'updated_from_ui': True,
        'previous_identity': {'campus_code': old_code, 'branch': old_branch},
    }
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail={'code': 'ACADEMIC_CAMPUS_CONFLICT', 'message': 'Mã cơ sở đã tồn tại trong hệ này.'},
        ) from exc
    db.refresh(campus)
    log_audit(
        db,
        action='academic.campus.update',
        status='success',
        message='Cập nhật cơ sở AP thành công',
        user=user,
        target_type='academic_campus',
        target_id=campus.id,
        metadata={'campus_code': campus.campus_code, 'branch': campus.branch, 'previous_code': old_code, 'previous_branch': old_branch},
    )
    return campus


@router.post('/campuses/seed-from-env', response_model=list[AcademicCampusOut])
def seed_academic_campuses_from_env(
    branch: str = Query('poly'),
    user: UserContext = Depends(_require_academic_catalog_admin),
    db: Session = Depends(get_db),
):
    items = AcademicImportService(db).seed_campuses_from_settings(branch=branch)
    log_audit(db, action='academic.campus.seed_from_env', status='success', message='Seed cơ sở AP từ env thành công', user=user, target_type='academic_campus', target_id='bulk', metadata={'branch': branch, 'count': len(items)})
    return items


@router.post('/campuses/sync-from-ap', response_model=list[AcademicCampusOut])
def sync_academic_campuses_from_ap(
    branch: str = Query('poly'),
    user: UserContext = Depends(_require_academic_catalog_admin),
    db: Session = Depends(get_db),
):
    return AcademicAPSyncWorkflowService(db).sync_campuses_from_ap(branch=branch, user=user)


@router.delete('/campuses/{campus_id}', response_model=AcademicCampusOut)
def delete_academic_campus(
    campus_id: str,
    user: UserContext = Depends(_require_academic_catalog_admin),
    db: Session = Depends(get_db),
):
    campus = db.query(AcademicCampus).filter(AcademicCampus.id == campus_id).first()
    if not campus:
        raise HTTPException(status_code=404, detail='Không tìm thấy cơ sở')
    campus.active = False
    meta = dict(campus.metadata_json or {})
    meta.update({'deleted_from_ui': True})
    campus.metadata_json = meta
    db.commit()
    db.refresh(campus)
    log_audit(db, action='academic.campus.delete', status='success', message='Đã xóa/ẩn cơ sở AP', user=user, target_type='academic_campus', target_id=campus.id, metadata={'campus_code': campus.campus_code, 'branch': campus.branch})
    return campus


@router.get('/sync/ap/options', response_model=AcademicAPSyncOptionsOut)
def get_ap_sync_options(
    term_name: str = Query('', description='Tên kỳ AP, ví dụ Summer 2026.'),
    branch: str = Query('poly'),
    campus: str | None = Query(None, description='Giữ tương thích UI cũ; danh sách môn lấy từ AP CMS /api/cms/get-subject-cms. Nếu có term_name thì backend gửi thêm term_name; nếu AP CMS cần campus_code tạm thời, cấu hình static trong env ACADEMIC_AP_CMS_GET_SUBJECT_ENDPOINT.'),
    include_subjects: bool = Query(True),
    user: UserContext = Depends(_require_academic_catalog_admin),
    db: Session = Depends(get_db),
):
    return AcademicAPSyncWorkflowService(db).get_sync_options(term_name=term_name or None, branch=branch, campus=campus, include_subjects=include_subjects)


@router.post('/sync/from-json', response_model=AcademicImportResultOut)
def sync_from_json(
    payload: AcademicImportFromJsonIn,
    user: UserContext = Depends(require_permission('manage_settings')),
    db: Session = Depends(get_db),
):
    _require_academic_admin(db, user)
    return AcademicAPSyncWorkflowService(db).sync_from_json(payload, user=user)




@router.post('/sync/ap/jobs', response_model=AcademicImportResultOut)
def enqueue_sync_from_ap_job(
    payload: AcademicAPSyncIn,
    user: UserContext = Depends(_require_academic_catalog_admin),
    db: Session = Depends(get_db),
):
    return AcademicAPSyncWorkflowService(db).enqueue_sync_from_ap_job(payload, user=user)



@router.get('/sync/ap/jobs', response_model=list[AcademicSyncRunOut])
def list_ap_sync_jobs(
    term_name: str = Query(''),
    branch: str = Query(''),
    status_filter: str = Query('active', alias='status'),
    limit: int = Query(10, ge=1, le=50),
    user: UserContext = Depends(_require_academic_catalog_admin),
    db: Session = Depends(get_db),
):
    return AcademicAPSyncWorkflowService(db).list_sync_jobs(term_name=term_name, branch=branch, status_filter=status_filter, limit=limit)



@router.get('/sync/ap/jobs/{run_id}', response_model=AcademicSyncRunOut)
def get_ap_sync_job(
    run_id: str,
    user: UserContext = Depends(_require_academic_catalog_admin),
    db: Session = Depends(get_db),
):
    return AcademicAPSyncWorkflowService(db).get_sync_job(run_id)



@router.post('/sync/ap', response_model=AcademicImportResultOut)
def sync_from_ap(
    payload: AcademicAPSyncIn,
    user: UserContext = Depends(require_permission('manage_settings')),
    db: Session = Depends(get_db),
):
    _require_academic_admin(db, user)
    return AcademicAPSyncWorkflowService(db).sync_from_ap(payload, user=user)
