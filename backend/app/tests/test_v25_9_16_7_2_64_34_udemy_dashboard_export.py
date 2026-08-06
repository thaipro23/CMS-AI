from datetime import date, datetime
from io import BytesIO
from pathlib import Path
import os
import sys
import types

if 'jose' not in sys.modules:
    jose_stub = types.ModuleType('jose')
    jose_stub.JWTError = type('JWTError', (Exception,), {})
    jose_stub.jwt = types.SimpleNamespace(decode=lambda *args, **kwargs: {})
    sys.modules['jose'] = jose_stub

os.environ.setdefault('DATABASE_URL', 'sqlite+pysqlite:///:memory:')

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.models.academic import (
    AcademicBlock,
    AcademicBulkOperationJob,
    AcademicClass,
    AcademicClassCourseMapping,
    AcademicClassStudent,
    AcademicCourseMapping,
    AcademicStudent,
    AcademicSubject,
    AcademicSubjectDelivery,
    AcademicTeacher,
    AcademicTeacherAssignment,
    AcademicTerm,
    UdemyProgressImportBatch,
    UdemyStudentProgress,
    UdemySubjectPlan,
    UdemySubjectPlanMilestone,
    OpenEdXUserMapping,
)
from app.core.rbac import UserContext
from app.services.academic.udemy_progress import UdemyProgressService
from app.services.academic_service import AcademicService
from app.api.routes.academic import _create_training_teacher_report_workbook

ROOT = Path(__file__).resolve().parents[3]


def _session():
    engine = create_engine('sqlite:///:memory:')
    for table in [
        AcademicTerm.__table__, AcademicBlock.__table__, AcademicSubject.__table__,
        AcademicSubjectDelivery.__table__, AcademicStudent.__table__, AcademicClass.__table__,
        AcademicClassStudent.__table__, OpenEdXUserMapping.__table__,
        AcademicClassCourseMapping.__table__, AcademicCourseMapping.__table__,
        AcademicTeacher.__table__, AcademicTeacherAssignment.__table__, AcademicBulkOperationJob.__table__,
        UdemySubjectPlan.__table__, UdemySubjectPlanMilestone.__table__,
        UdemyProgressImportBatch.__table__, UdemyStudentProgress.__table__,
    ]:
        table.create(engine)
    return sessionmaker(bind=engine)()


def _seed(db):
    term = AcademicTerm(id='term-su26', term_code='SU26', term_name='Summer 2026', branch='poly', active=True)
    block = AcademicBlock(id='block-1', term_id=term.id, block_code='Block 1', block_name='Block 1', sort_order=1, active=True)
    subject = AcademicSubject(id='subject-sof3032', subject_code='SOF3032', subject_name='Java nâng cao', branch='poly', active=True)
    delivery = AcademicSubjectDelivery(id='delivery-sof3032', subject_id=subject.id, term_id=term.id, block_id=block.id, branch='poly', learning_platform='udemy', active=True)
    class_a = AcademicClass(id='class-a', term_id=term.id, block_id=block.id, subject_id=subject.id, class_code='SOF3032.01', class_name='SOF3032.01', campus='ph', branch='poly', active=True)
    class_b = AcademicClass(id='class-b', term_id=term.id, block_id=block.id, subject_id=subject.id, class_code='SOF3032.02', class_name='SOF3032.02', campus='hn', branch='poly', active=True)
    student_a = AcademicStudent(id='student-a', student_code='PH00001', username='PH00001', email='a@fpt.edu.vn', full_name='Sinh viên A', campus='ph', branch='poly', active=True)
    student_b = AcademicStudent(id='student-b', student_code='HN00002', username='HN00002', email='b@fpt.edu.vn', full_name='Sinh viên B', campus='hn', branch='poly', active=True)
    teacher = AcademicTeacher(id='teacher-a', username='teacher.a', email='teacher.a@fpt.edu.vn', full_name='Giảng viên A', campus='ph', branch='poly', active=True)
    assignment = AcademicTeacherAssignment(id='assignment-a', teacher_id=teacher.id, class_id=class_a.id, subject_id=subject.id, term_id=term.id, block_id=block.id, campus='ph', branch='poly')
    plan = UdemySubjectPlan(id='plan-1', subject_delivery_id=delivery.id, version=1, item_count=10, active=True, source='manual')
    milestone = UdemySubjectPlanMilestone(id='milestone-1', plan_id=plan.id, week_number=2, deadline_date=date(2026, 1, 1), required_progress_percent=60, sort_order=2)
    batch = UdemyProgressImportBatch(id='batch-1', subject_delivery_id=delivery.id, idempotency_key='batch34-key', file_name='SOF3032_report.xlsx', file_hash='0' * 64, status='completed', parser_format='aggregate', processed_rows=3, matched_rows=2, requested_by='admin', created_at=datetime(2026, 1, 2), finished_at=datetime(2026, 1, 2))
    progress_a = UdemyStudentProgress(id='progress-a', subject_delivery_id=delivery.id, class_id=class_a.id, student_id=student_a.id, email='a@fpt.edu.vn', normalized_email='a@fpt.edu.vn', display_name='Sinh viên A', progress_percent=40, is_late=False, current_plan_week=2, required_progress_percent=60, current_deadline_date=date(2026, 1, 1), match_status='matched_roster', source_format='aggregate', last_import_batch_id=batch.id, last_imported_at=datetime(2026, 1, 2))
    progress_b = UdemyStudentProgress(id='progress-b', subject_delivery_id=delivery.id, class_id=class_b.id, student_id=student_b.id, email='b@fpt.edu.vn', normalized_email='b@fpt.edu.vn', display_name='Sinh viên B', progress_percent=80, is_late=True, current_plan_week=2, required_progress_percent=60, current_deadline_date=date(2026, 1, 1), match_status='matched_roster', source_format='aggregate', last_import_batch_id=batch.id, last_imported_at=datetime(2026, 1, 2))
    progress_unknown = UdemyStudentProgress(id='progress-u', subject_delivery_id=delivery.id, class_id=None, student_id=None, email='unknown@fpt.edu.vn', normalized_email='unknown@fpt.edu.vn', display_name='Chưa khớp', progress_percent=90, is_late=False, current_plan_week=2, required_progress_percent=60, current_deadline_date=date(2026, 1, 1), match_status='unmatched', source_format='aggregate', last_import_batch_id=batch.id, last_imported_at=datetime(2026, 1, 2))
    db.add_all([term, block, subject, delivery, class_a, class_b, student_a, student_b, teacher, assignment, plan, milestone, batch, progress_a, progress_b, progress_unknown])
    db.commit()
    return delivery


def test_dashboard_list_filters_scope_and_export():
    db = _session()
    delivery = _seed(db)
    service = UdemyProgressService(db)

    dashboard = service.dashboard(delivery.id)
    assert dashboard['summary']['total_students'] == 3
    assert dashboard['summary']['late_students'] == 1
    assert dashboard['summary']['on_track_students'] == 1
    # Status is recalculated against the current due milestone, not trusted from stale snapshot flags.
    assert dashboard['summary']['unmatched_students'] == 1
    assert dashboard['summary']['required_progress_percent'] == 60
    assert len(dashboard['recent_imports']) == 1

    scoped = service.dashboard(delivery.id, allowed_class_ids={'class-a'}, scope_label='lớp được AP phân công')
    assert scoped['summary']['total_students'] == 1
    assert scoped['summary']['late_students'] == 1
    assert scoped['summary']['scope_label'] == 'lớp được AP phân công'
    assert scoped['recent_imports'] == []

    late = service.list_students(delivery.id, status_filter='late')
    assert late['total'] == 1
    assert late['items'][0]['student_code'] == 'PH00001'
    assert late['items'][0]['teacher_names'] == ['Giảng viên A']
    assert late['items'][0]['variance_percent'] == -20

    scoped_rows = service.list_students(delivery.id, allowed_class_ids={'class-a'})
    assert scoped_rows['total'] == 1
    assert all(item['class_id'] == 'class-a' for item in scoped_rows['items'])

    raw = service.export_workbook(delivery.id, status_filter='late')
    wb = load_workbook(BytesIO(raw), data_only=True)
    assert wb.sheetnames == ['TongQuan', 'TienDoSinhVien', 'CanhBao', 'LichSuImport', 'HuongDan']
    assert wb['TienDoSinhVien'].max_row == 2
    assert wb['TienDoSinhVien']['B2'].value == 'PH00001'
    assert wb['CanhBao'].max_row == 2
    wb.close()



def test_class_detail_routes_udemy_classes_to_udemy_dashboard():
    db = _session()
    delivery = _seed(db)
    user = UserContext(user_id='admin', role='admin', permissions={'manage_settings'}, raw_claims={'ai_system_admin': True})

    detail = AcademicService(db).get_class_detail(user, 'class-a')

    assert detail['learning_platform'] == 'udemy'
    assert detail['subject_delivery_id'] == delivery.id
    assert detail['udemy_progress_student_count'] == 1
    assert detail['udemy_progress_late_count'] == 1
    assert detail['udemy_progress_average_percent'] == 40.0
    assert detail['udemy_progress_last_imported_at'] == datetime(2026, 1, 2)



def test_teacher_management_udemy_context_and_export_contract():
    db = _session()
    delivery = _seed(db)
    class_a = db.query(AcademicClass).filter(AcademicClass.id == 'class-a').one()
    workflow = AcademicService(db)._training_teacher_report_workflow()

    context = workflow._teacher_udemy_context([class_a])
    row = context[class_a.id]
    assert row['learning_platform'] == 'udemy'
    assert row['subject_delivery_id'] == delivery.id
    assert row['progress_student_count'] == 1
    assert row['late_student_count'] == 1
    assert row['average_progress_percent'] == 40.0
    assert row['required_progress_percent'] == 60

    report = {
        'items': [{
            'teacher_id': 'teacher-a', 'teacher_name': 'Giảng viên A', 'teacher_username': 'teacher.a',
            'teacher_email': 'teacher.a@fpt.edu.vn', 'branch': 'poly', 'campus': 'ph',
            'subject_count': 1, 'subject_codes': ['SOF3032'], 'class_count': 1,
            'cms_class_count': 0, 'udemy_class_count': 1, 'student_count': 1,
            'cms_student_count': 0, 'udemy_student_count': 1, 'unique_student_count': 1,
            'relearn_student_count': 0, 'total_relearn_count': 0, 'cms_synced_count': 0,
            'learning_enrolled_count': 0, 'learning_active_count': 0,
            'udemy_progress_student_count': 1, 'udemy_progress_late_count': 1,
            'udemy_progress_average_percent': 40.0, 'udemy_progress_last_imported_at': datetime(2026, 1, 2),
            'classes_without_course_count': 0, 'risk_student_count': 1, 'status_counts': {},
            'learning_alerts': ['1 SV Udemy chậm tiến độ'],
            'classes': [{
                'class_id': class_a.id, 'class_code': class_a.class_code, 'class_name': class_a.class_name,
                'branch': 'poly', 'campus': 'ph', 'term_name': 'Summer 2026', 'block_name': 'Block 1',
                'subject_code': 'SOF3032', 'subject_name': 'Java nâng cao', 'learning_platform': 'udemy',
                'subject_delivery_id': delivery.id, 'student_count': 1, 'cms_synced_count': 0,
                'learning_enrolled_count': 0, 'learning_active_count': 0, 'status_counts': {},
                'udemy_progress_student_count': 1, 'udemy_progress_late_count': 1,
                'udemy_progress_average_percent': 40.0, 'udemy_progress_required_percent': 60,
                'udemy_progress_current_week': 2, 'udemy_progress_deadline_date': date(2026, 1, 1),
                'udemy_progress_last_imported_at': datetime(2026, 1, 2),
                'learning_alerts': ['1 SV Udemy chậm tiến độ'],
            }],
        }],
        'student_watch_rows': [],
    }
    wb = _create_training_teacher_report_workbook(report)
    assert 'UdemyChamTienDo' in wb.sheetnames
    overview_headers = [cell.value for cell in wb['TongQuanGV'][1]]
    assert 'Lớp Udemy' in overview_headers
    assert 'Tiến độ Udemy TB (%)' in overview_headers
    class_headers = [cell.value for cell in wb['ChiTietLop'][1]]
    assert 'Nền tảng' in class_headers
    assert 'Mốc Udemy hiện tại' in class_headers
    assert wb['UdemyChamTienDo'].max_row == 2
    wb.close()

def test_batch34_cross_layer_contracts():
    service = (ROOT / 'backend/app/services/academic/udemy_progress.py').read_text(encoding='utf-8')
    routes = (ROOT / 'backend/app/api/routes/academic.py').read_text(encoding='utf-8')
    schemas = (ROOT / 'backend/app/schemas/academic.py').read_text(encoding='utf-8')
    api = (ROOT / 'frontend/lib/api.ts').read_text(encoding='utf-8')
    page = (ROOT / 'frontend/app/subject-management/[deliveryId]/udemy/page.tsx').read_text(encoding='utf-8')
    subjects = (ROOT / 'frontend/app/subject-management/page.tsx').read_text(encoding='utf-8')
    teacher_service = (ROOT / 'backend/app/services/academic/teacher_report.py').read_text(encoding='utf-8')
    teacher_page = (ROOT / 'frontend/app/teacher-management/page.tsx').read_text(encoding='utf-8')
    types = (ROOT / 'frontend/types/index.ts').read_text(encoding='utf-8')

    assert 'def export_workbook(' in service
    assert 'allowed_class_ids' in service
    assert "@router.get('/subject-deliveries/{delivery_id}/udemy-progress/dashboard'" in routes
    assert "@router.get('/subject-deliveries/{delivery_id}/udemy-progress/students'" in routes
    assert "@router.get('/subject-deliveries/{delivery_id}/udemy-progress/export.xlsx')" in routes
    assert 'class UdemyProgressDashboardOut' in schemas
    assert 'downloadUdemyProgressExport' in api
    assert 'Sinh viên cần xử lý' in page
    assert 'Lịch sử import gần nhất' in page
    assert '/udemy`}>Xem tiến độ' in subjects
    assert '_teacher_udemy_context' in teacher_service
    assert "status_filter == 'udemy_late'" in teacher_service
    assert 'Có SV Udemy chậm tiến độ' in teacher_page
    assert 'Làm mới số liệu' in teacher_page
    assert 'createAcademicTrainingTeacherCacheJob' in teacher_page
    assert 'udemy_progress_average_percent' in types
    assert "wb.create_sheet('UdemyChamTienDo')" in routes
