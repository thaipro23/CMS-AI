from __future__ import annotations

import csv
import math
import re
import time
import unicodedata
import zipfile
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.json_safe import json_safe_value
from app.core.timezone import vn_now
from app.models.academic import (
    AcademicBlock,
    AcademicClass,
    AcademicClassStudent,
    AcademicStudent,
    AcademicSubject,
    AcademicSubjectDelivery,
    AcademicTeacher,
    AcademicTeacherAssignment,
    AcademicTerm,
    UdemyProgressImportBatch,
    UdemyProgressUnmatchedRow,
    UdemyStudentProgress,
    UdemySubjectPlan,
    UdemySubjectPlanMilestone,
)


@dataclass
class ParsedProgressRecord:
    email: str
    normalized_email: str
    display_name: str | None
    progress_percent: float
    class_code: str | None
    row_numbers: list[int] = field(default_factory=list)
    source_row_count: int = 1
    metadata: dict[str, Any] = field(default_factory=dict)


class UdemyProgressService:
    """Safe parser and snapshot writer for Udemy progress exports.

    Supported formats:
    - Raw Udemy item-level export in .xlsx or UTF-8 .csv. Columns are resolved
      by header name, so optional/new columns such as ``ID bên ngoài`` may be
      inserted or reordered without shifting Email or completion percentage.
    - Legacy 25-column item export as a compatibility fallback.
    - ACMS aggregate export containing Email, Name and Tiến độ hiện tại.
    """

    MAX_FILE_BYTES = int(settings.academic_udemy_import_max_file_bytes)
    MAX_TOTAL_UPLOAD_BYTES = int(settings.academic_udemy_import_max_total_bytes)
    MAX_FILES = int(settings.academic_udemy_import_max_files)
    MAX_ROWS = int(settings.academic_udemy_import_max_rows)
    MAX_XLSX_ENTRIES = int(settings.academic_udemy_xlsx_max_entries)
    MAX_XLSX_UNCOMPRESSED_BYTES = int(settings.academic_udemy_xlsx_max_uncompressed_bytes)
    MAX_XLSX_COMPRESSION_RATIO = int(settings.academic_udemy_xlsx_max_compression_ratio)
    ALLOWED_CONTENT_TYPES = {
        '',
        'application/octet-stream',
        'application/zip',
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'text/csv',
        'application/csv',
        'application/vnd.ms-excel',
        'text/plain',
    }
    ALLOWED_EXTENSIONS = {'.xlsx', '.csv'}
    EMAIL_RE = re.compile(r'^[^\s@]+@[^\s@]+\.[^\s@]+$')

    HEADER_ALIASES = {
        'email': {
            'email', 'email address', 'user email', 'learner email', 'student email', 'e mail',
            'email cua nguoi dung', 'email người dùng', 'email nguoi dung',
        },
        'name': {
            'name', 'full name', 'user name', 'learner name', 'student name', 'ten', 'ho ten',
            'ten cua nguoi dung', 'ten người dùng', 'ten nguoi dung',
        },
        'last_name': {'last name', 'surname', 'family name', 'ho cua nguoi dung', 'ho nguoi dung'},
        'progress': {
            'progress', 'progress percent', 'progress percentage', 'completion', 'completion percent',
            'item progress', 'item completion', 'item completion percent', 'item completion percentage',
            'path item completion', 'path item completion percent', 'path item completion percentage',
            'tien do', 'tien do hien tai', 'phan tram tien do',
            'ty le hoan thanh muc trong lo trinh', 'ti le hoan thanh muc trong lo trinh',
        },
        'class_code': {'class', 'class code', 'class name', 'lop', 'ma lop'},
        'subject_code': {'subject', 'subject code', 'course code', 'ma mon', 'ma mon hoc'},
    }
    AGGREGATE_PROGRESS_HEADERS = {
        'tien do hien tai', 'overall progress', 'course progress', 'completion percent', 'progress percent',
    }

    def __init__(self, db: Session):
        self.db = db

    @staticmethod
    def normalize_text(value: Any) -> str:
        text = unicodedata.normalize('NFKD', str(value or '').strip().lower())
        text = ''.join(char for char in text if not unicodedata.combining(char)).replace('đ', 'd')
        return re.sub(r'[^a-z0-9]+', ' ', text).strip()

    @classmethod
    def normalize_email(cls, value: Any) -> str:
        return str(value or '').strip().lower()

    @classmethod
    def _canonical_header(cls, value: Any) -> str | None:
        normalized = cls.normalize_text(value)
        for canonical, aliases in cls.HEADER_ALIASES.items():
            if normalized in {cls.normalize_text(item) for item in aliases}:
                return canonical
        return None

    @classmethod
    def validate_upload_metadata(cls, *, filename: str, content_type: str | None) -> None:
        suffix = Path(str(filename or '')).suffix.lower()
        if suffix not in cls.ALLOWED_EXTENSIONS:
            raise ValueError('Chỉ chấp nhận file tiến độ Udemy .xlsx hoặc .csv.')
        clean_type = str(content_type or '').split(';', 1)[0].strip().lower()
        if clean_type not in cls.ALLOWED_CONTENT_TYPES:
            raise ValueError('Content-Type của file không phù hợp với .xlsx/.csv.')

    @classmethod
    def _validate_common_size(cls, raw: bytes) -> None:
        if not raw:
            raise ValueError('File Udemy rỗng.')
        if len(raw) > cls.MAX_FILE_BYTES:
            limit_mb = max(1, cls.MAX_FILE_BYTES // (1024 * 1024))
            raise ValueError(f'Mỗi file Udemy tối đa {limit_mb} MB.')

    @classmethod
    def _validate_xlsx(cls, raw: bytes) -> None:
        cls._validate_common_size(raw)
        if not raw.startswith(b'PK'):
            raise ValueError('Nội dung file không phải workbook .xlsx hợp lệ.')
        try:
            with zipfile.ZipFile(BytesIO(raw)) as archive:
                entries = archive.infolist()
                if len(entries) > cls.MAX_XLSX_ENTRIES:
                    raise ValueError('File Excel có quá nhiều thành phần, không an toàn để xử lý.')
                names = {item.filename.replace('\\', '/') for item in entries}
                if '[Content_Types].xml' not in names or 'xl/workbook.xml' not in names:
                    raise ValueError('Workbook thiếu cấu trúc OpenXML bắt buộc.')
                total = 0
                for item in entries:
                    normalized = item.filename.replace('\\', '/')
                    parts = [part for part in normalized.split('/') if part not in {'', '.'}]
                    if normalized.startswith('/') or '..' in parts:
                        raise ValueError('Workbook chứa đường dẫn nội bộ không an toàn.')
                    if item.flag_bits & 0x1:
                        raise ValueError('Workbook được mã hóa bằng mật khẩu nên không thể xử lý.')
                    file_size = max(0, int(item.file_size or 0))
                    compressed_size = max(0, int(item.compress_size or 0))
                    total += file_size
                    if compressed_size and file_size > max(1024 * 1024, compressed_size * cls.MAX_XLSX_COMPRESSION_RATIO):
                        raise ValueError('Workbook có tỷ lệ nén bất thường, có nguy cơ zip bomb.')
                if total > cls.MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise ValueError('Dữ liệu giải nén của file Excel vượt giới hạn an toàn.')
        except zipfile.BadZipFile as exc:
            raise ValueError('Không đọc được file Excel. Hãy dùng file .xlsx hợp lệ.') from exc

    @classmethod
    def _decode_csv(cls, raw: bytes) -> str:
        cls._validate_common_size(raw)
        if b'\x00' in raw:
            raise ValueError('File CSV chứa byte rỗng và không phải CSV văn bản hợp lệ.')
        try:
            return raw.decode('utf-8-sig')
        except UnicodeDecodeError as exc:
            raise ValueError('File CSV phải dùng mã hóa UTF-8.') from exc

    @classmethod
    def _csv_dialect(cls, text: str):
        sample = text[:65536]
        try:
            return csv.Sniffer().sniff(sample, delimiters=',;\t')
        except csv.Error:
            return csv.excel

    @classmethod
    def validate_upload_content(cls, *, filename: str, raw: bytes) -> None:
        suffix = Path(str(filename or '')).suffix.lower()
        if suffix == '.xlsx':
            cls._validate_xlsx(raw)
            return
        if suffix == '.csv':
            text = cls._decode_csv(raw)
            reader = csv.reader(StringIO(text), dialect=cls._csv_dialect(text))
            rows = []
            for _ in range(15):
                try:
                    rows.append(next(reader))
                except StopIteration:
                    break
            cls._find_header_rows(rows)
            return
        raise ValueError('Chỉ chấp nhận file tiến độ Udemy .xlsx hoặc .csv.')

    @classmethod
    def cleanup_expired_artifacts(
        cls,
        *,
        root: Path | None = None,
        now_ts: float | None = None,
        protected_import_job_ids: set[str] | None = None,
        protected_export_file_names: set[str] | None = None,
    ) -> dict[str, int]:
        """Delete expired Udemy artifacts while preserving active job inputs.

        Job records remain in PostgreSQL for audit. The caller supplies active
        import parent IDs and export filenames so a delayed worker never loses
        the source artifact it still needs. Retry endpoints return HTTP 410 only
        after the configured retention actually expires.
        """
        storage_root = (root or Path(settings.local_storage_path or '/app/.runtime')).expanduser().resolve()
        now_value = float(now_ts if now_ts is not None else time.time())
        import_cutoff = now_value - max(1, int(settings.academic_udemy_file_retention_hours)) * 3600
        export_cutoff = now_value - max(1, int(settings.academic_udemy_export_file_retention_hours)) * 3600
        deleted_files = 0
        deleted_dirs = 0
        scanned = 0
        max_scan = 10_000
        protected_import_ids = {str(item) for item in (protected_import_job_ids or set()) if str(item).strip()}
        protected_export_names = {str(item) for item in (protected_export_file_names or set()) if str(item).strip()}

        def remove_old_files(base: Path, cutoff: float, *, artifact_kind: str) -> None:
            nonlocal deleted_files, deleted_dirs, scanned
            if not base.exists():
                return
            for path in base.rglob('*'):
                if scanned >= max_scan:
                    break
                scanned += 1
                try:
                    relative = path.relative_to(base)
                    if artifact_kind == 'import' and relative.parts and relative.parts[0] in protected_import_ids:
                        continue
                    if artifact_kind == 'export' and path.name in protected_export_names:
                        continue
                    if path.is_file() and path.stat().st_mtime < cutoff:
                        path.unlink(missing_ok=True)
                        deleted_files += 1
                except OSError:
                    continue
            for directory in sorted((item for item in base.rglob('*') if item.is_dir()), key=lambda item: len(item.parts), reverse=True):
                try:
                    relative = directory.relative_to(base)
                    if artifact_kind == 'import' and relative.parts and relative.parts[0] in protected_import_ids:
                        continue
                    directory.rmdir()
                    deleted_dirs += 1
                except OSError:
                    pass

        remove_old_files(storage_root / 'udemy-progress-imports', import_cutoff, artifact_kind='import')
        remove_old_files(storage_root / 'udemy-progress-exports', export_cutoff, artifact_kind='export')
        return {'deleted_files': deleted_files, 'deleted_dirs': deleted_dirs, 'scanned': scanned}

    def resolve_delivery(
        self,
        *,
        term_id: str,
        block_id: str,
        branch: str,
        filename: str,
        delivery_id: str | None = None,
    ) -> tuple[AcademicSubjectDelivery, AcademicSubject]:
        branch_value = str(branch or 'poly').strip().lower()
        if branch_value not in {'poly', 'ptcd'}:
            raise ValueError('Hệ chỉ nhận poly hoặc ptcd.')
        term = self.db.get(AcademicTerm, term_id)
        block = self.db.get(AcademicBlock, block_id)
        if not term:
            raise ValueError('Không tìm thấy học kỳ đã chọn.')
        if not block or block.term_id != term.id:
            raise ValueError('Block không thuộc học kỳ đã chọn.')

        if delivery_id:
            delivery = self.db.get(AcademicSubjectDelivery, delivery_id)
            if not delivery or not delivery.active:
                raise ValueError('Không tìm thấy môn Udemy đã chọn.')
            if delivery.term_id != term_id or delivery.block_id != block_id or str(delivery.branch).lower() != branch_value:
                raise ValueError('Môn Udemy không thuộc đúng học kỳ, Block và hệ hiện tại.')
            subject = self.db.get(AcademicSubject, delivery.subject_id)
        else:
            stem = Path(filename).stem.strip()
            match = re.match(r'^([A-Za-z]{2,}[A-Za-z0-9]*)[_\-\s]', stem)
            subject_code = (match.group(1) if match else stem.split('_', 1)[0]).strip().upper()
            if not re.fullmatch(r'[A-Z][A-Z0-9]{2,31}', subject_code):
                raise ValueError(f'Không xác định được mã môn từ tên file “{filename}”. Dùng dạng SOF3032_*.xlsx hoặc import tại đúng dòng môn.')
            subject = self.db.query(AcademicSubject).filter(
                func.upper(AcademicSubject.subject_code) == subject_code,
                func.lower(func.coalesce(AcademicSubject.branch, branch_value)) == branch_value,
                AcademicSubject.active.is_(True),
            ).first()
            if not subject:
                raise ValueError(f'Không tìm thấy môn {subject_code} trong hệ {branch_value.upper()}.')
            delivery = self.db.query(AcademicSubjectDelivery).filter(
                AcademicSubjectDelivery.subject_id == subject.id,
                AcademicSubjectDelivery.term_id == term_id,
                AcademicSubjectDelivery.block_id == block_id,
                func.lower(AcademicSubjectDelivery.branch) == branch_value,
                AcademicSubjectDelivery.active.is_(True),
            ).first()
        if not delivery or not subject:
            raise ValueError('Không tìm thấy bản ghi triển khai môn trong phạm vi đã chọn.')
        if delivery.learning_platform != 'udemy':
            current = 'Chưa chọn' if delivery.learning_platform is None else str(delivery.learning_platform).upper()
            raise ValueError(f'Môn {subject.subject_code} đang ở nền tảng {current}; chỉ môn Udemy mới được import tiến độ.')
        return delivery, subject

    @classmethod
    def _parse_percent(cls, value: Any, number_format: str | None = None) -> tuple[float, bool]:
        if value is None or str(value).strip() == '':
            raise ValueError('Thiếu tiến độ.')
        clamped = False
        if isinstance(value, str):
            text = value.strip().replace(',', '.')
            if text.upper() in {'#N/A', '#VALUE!', 'N/A', 'NA', 'NULL'}:
                raise ValueError('Tiến độ không hợp lệ.')
            percent_suffix = text.endswith('%')
            if percent_suffix:
                text = text[:-1].strip()
            try:
                number = float(text)
            except ValueError as exc:
                raise ValueError('Tiến độ phải là số.') from exc
            if not percent_suffix and 0 <= number <= 1 and '%' in str(number_format or ''):
                number *= 100
        else:
            try:
                number = float(value)
            except (TypeError, ValueError) as exc:
                raise ValueError('Tiến độ phải là số.') from exc
            if 0 <= number <= 1 and '%' in str(number_format or ''):
                number *= 100
        if not math.isfinite(number):
            raise ValueError('Tiến độ không hữu hạn.')
        if number < 0:
            raise ValueError('Tiến độ không được âm.')
        if number > 100:
            number = 100.0
            clamped = True
        return round(number, 4), clamped

    @classmethod
    def _classify_header(cls, found: dict[str, int], raw_progress_header: str, column_count: int) -> str:
        aggregate = raw_progress_header in {cls.normalize_text(item) for item in cls.AGGREGATE_PROGRESS_HEADERS}
        if column_count < 20:
            aggregate = True
        return 'aggregate' if aggregate else 'item_rows'

    @classmethod
    def _map_header_values(cls, values: list[Any]) -> tuple[dict[str, int], str]:
        found: dict[str, int] = {}
        raw_progress_header = ''
        for col_no, raw in enumerate(values[:200], start=1):
            canonical = cls._canonical_header(raw)
            if canonical and canonical not in found:
                found[canonical] = col_no
                if canonical == 'progress':
                    raw_progress_header = cls.normalize_text(raw)
        return found, raw_progress_header

    @classmethod
    def _find_header_rows(cls, rows: list[list[Any]]) -> tuple[int, dict[str, int], str]:
        for row_no, values in enumerate(rows[:15], start=1):
            found, raw_progress_header = cls._map_header_values(list(values))
            if 'email' in found and 'progress' in found:
                return row_no, found, cls._classify_header(found, raw_progress_header, len(values))
        # Compatibility only for historical exports whose headers cannot be
        # normalized. New/current Udemy exports are always resolved by header.
        if rows:
            column_count = len(rows[0])
            if column_count == 25:
                return 1, {'email': 3, 'name': 2, 'progress': 17}, 'legacy_25_item_rows'
            if column_count == 26:
                return 1, {'email': 3, 'name': 2, 'progress': 18}, 'legacy_26_item_rows'
        raise ValueError('Không nhận diện được cột Email và Tỷ lệ hoàn thành. Hãy dùng file export gốc từ Udemy hoặc file tổng hợp tiến độ.')

    @classmethod
    def _find_header(cls, ws) -> tuple[int, dict[str, int], str]:
        rows: list[list[Any]] = []
        for row_no in range(1, min(15, ws.max_row or 15) + 1):
            rows.append([ws.cell(row=row_no, column=col_no).value for col_no in range(1, min(ws.max_column or 1, 200) + 1)])
        return cls._find_header_rows(rows)

    @classmethod
    def _finalize_parsed_rows(
        cls,
        *,
        rows: Any,
        header_row: int,
        columns: dict[str, int],
        parser_format: str,
        active_plan: UdemySubjectPlan | None,
        cells: bool,
    ) -> dict[str, Any]:
        grouped: dict[str, dict[str, Any]] = {}
        issues: list[dict[str, Any]] = []
        read_rows = 0
        for row_no, row in enumerate(rows, start=header_row + 1):
            if read_rows >= cls.MAX_ROWS:
                issues.append({'row_number': row_no, 'reason_code': 'ROW_LIMIT', 'reason_message': f'File vượt giới hạn {cls.MAX_ROWS} dòng.'})
                break
            read_rows += 1

            def item_at(canonical: str):
                col = columns.get(canonical)
                return row[col - 1] if col and col - 1 < len(row) else None

            def value_at(canonical: str):
                item = item_at(canonical)
                return getattr(item, 'value', None) if cells and item is not None else item

            email_raw = value_at('email')
            progress_raw = value_at('progress')
            name_raw = value_at('name')
            last_name_raw = value_at('last_name')
            class_raw = value_at('class_code')
            name_parts = [str(value).strip() for value in (name_raw, last_name_raw) if value not in {None, ''}]
            name = ' '.join(name_parts) or None
            class_code = str(class_raw).strip() if class_raw not in {None, ''} else None
            if email_raw in {None, ''} and progress_raw in {None, ''} and not name:
                continue
            normalized_email = cls.normalize_email(email_raw)
            if not normalized_email or not cls.EMAIL_RE.match(normalized_email):
                issues.append({
                    'row_number': row_no, 'email': str(email_raw or ''), 'display_name': name,
                    'raw_progress': str(progress_raw or ''), 'reason_code': 'INVALID_EMAIL',
                    'reason_message': 'Email trống hoặc không đúng định dạng.',
                })
                continue
            progress_item = item_at('progress')
            number_format = getattr(progress_item, 'number_format', None) if cells and progress_item is not None else None
            try:
                progress, clamped = cls._parse_percent(progress_raw, number_format)
            except ValueError as exc:
                issues.append({
                    'row_number': row_no, 'email': normalized_email, 'display_name': name,
                    'raw_progress': str(progress_raw or ''), 'reason_code': 'INVALID_PROGRESS',
                    'reason_message': str(exc),
                })
                continue
            item = grouped.setdefault(normalized_email, {
                'email': str(email_raw).strip(), 'normalized_email': normalized_email,
                'display_name': name, 'class_code': class_code, 'values': [], 'row_numbers': [], 'clamped_rows': 0,
            })
            if name and not item.get('display_name'):
                item['display_name'] = name
            if class_code and not item.get('class_code'):
                item['class_code'] = class_code
            item['values'].append(progress)
            item['row_numbers'].append(row_no)
            item['clamped_rows'] += int(clamped)

        records: list[ParsedProgressRecord] = []
        warnings: list[str] = []
        item_count = int(active_plan.item_count) if active_plan else 0
        for item in grouped.values():
            values = item['values']
            if parser_format in {'item_rows', 'legacy_25_item_rows', 'legacy_26_item_rows'}:
                denominator = item_count or len(values)
                if not item_count:
                    warnings.append('Môn chưa có kế hoạch Udemy; tiến độ item được ước tính theo số dòng quan sát và chưa thể đánh giá chậm tiến độ.')
                progress = math.floor(sum(values) / max(1, denominator))
                progress = max(0.0, min(100.0, float(progress)))
            else:
                progress = max(values)
            records.append(ParsedProgressRecord(
                email=item['email'], normalized_email=item['normalized_email'], display_name=item.get('display_name'),
                progress_percent=round(progress, 2), class_code=item.get('class_code'),
                row_numbers=list(item['row_numbers']), source_row_count=len(values),
                metadata={'clamped_rows': item['clamped_rows'], 'duplicate_rows': max(0, len(values) - 1)},
            ))
        return {
            'parser_format': parser_format,
            'header_row': header_row,
            'read_rows': read_rows,
            'records': records,
            'issues': issues,
            'warnings': sorted(set(warnings)),
        }

    def parse_path(self, path: Path, *, active_plan: UdemySubjectPlan | None) -> dict[str, Any]:
        raw = path.read_bytes()
        suffix = path.suffix.lower()
        if suffix == '.csv':
            text = self._decode_csv(raw)
            dialect = self._csv_dialect(text)
            sample_reader = csv.reader(StringIO(text), dialect=dialect)
            sample_rows: list[list[Any]] = []
            for _ in range(15):
                try:
                    sample_rows.append(next(sample_reader))
                except StopIteration:
                    break
            header_row, columns, parser_format = self._find_header_rows(sample_rows)
            reader = csv.reader(StringIO(text), dialect=dialect)
            for _ in range(header_row):
                try:
                    next(reader)
                except StopIteration:
                    break
            return self._finalize_parsed_rows(
                rows=reader, header_row=header_row, columns=columns, parser_format=parser_format,
                active_plan=active_plan, cells=False,
            )

        if suffix != '.xlsx':
            raise ValueError('Chỉ chấp nhận file tiến độ Udemy .xlsx hoặc .csv.')
        self._validate_xlsx(raw)
        try:
            wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError('Không đọc được nội dung workbook.') from exc
        if not wb.worksheets:
            wb.close()
            raise ValueError('File không có sheet dữ liệu.')
        ws = wb.worksheets[0]
        header_row, columns, parser_format = self._find_header(ws)
        try:
            return self._finalize_parsed_rows(
                rows=ws.iter_rows(min_row=header_row + 1, values_only=False),
                header_row=header_row, columns=columns, parser_format=parser_format,
                active_plan=active_plan, cells=True,
            )
        finally:
            wb.close()

    def _active_plan_context(self, delivery_id: str, as_of: date) -> tuple[UdemySubjectPlan | None, dict[str, Any] | None]:
        plan = self.db.query(UdemySubjectPlan).filter(
            UdemySubjectPlan.subject_delivery_id == delivery_id,
            UdemySubjectPlan.active.is_(True),
        ).order_by(UdemySubjectPlan.version.desc()).first()
        if not plan:
            return None, None
        milestones = self.db.query(UdemySubjectPlanMilestone).filter(
            UdemySubjectPlanMilestone.plan_id == plan.id,
        ).order_by(UdemySubjectPlanMilestone.deadline_date.asc(), UdemySubjectPlanMilestone.week_number.asc()).all()
        current = None
        for milestone in milestones:
            if milestone.deadline_date <= as_of:
                current = milestone
            else:
                break
        if not current:
            return plan, {'week_number': None, 'deadline_date': None, 'required_progress_percent': 0.0}
        return plan, {
            'week_number': current.week_number,
            'deadline_date': current.deadline_date,
            'required_progress_percent': float(current.required_progress_percent),
        }

    def _roster_maps(self, delivery: AcademicSubjectDelivery) -> tuple[dict[str, list[dict[str, Any]]], dict[str, AcademicClass]]:
        classes = self.db.query(AcademicClass).filter(
            AcademicClass.subject_id == delivery.subject_id,
            AcademicClass.term_id == delivery.term_id,
            AcademicClass.block_id == delivery.block_id,
            func.lower(func.coalesce(AcademicClass.branch, delivery.branch)) == str(delivery.branch).lower(),
            AcademicClass.active.is_(True),
        ).all()
        class_by_code = {str(item.class_code or '').strip().lower(): item for item in classes if str(item.class_code or '').strip()}
        class_ids = [item.id for item in classes]
        roster: dict[str, list[dict[str, Any]]] = defaultdict(list)
        if class_ids:
            rows = self.db.query(AcademicClassStudent, AcademicStudent, AcademicClass).join(
                AcademicStudent, AcademicStudent.id == AcademicClassStudent.student_id,
            ).join(AcademicClass, AcademicClass.id == AcademicClassStudent.class_id).filter(
                AcademicClassStudent.class_id.in_(class_ids),
                AcademicStudent.active.is_(True),
            ).all()
            for link, student, class_row in rows:
                email = self.normalize_email(student.email)
                if not email:
                    continue
                roster[email].append({'link': link, 'student': student, 'class': class_row})
        return roster, class_by_code

    def _resolve_student(self, record: ParsedProgressRecord, roster: dict[str, list[dict[str, Any]]]) -> tuple[str, AcademicStudent | None, AcademicClass | None, str | None]:
        candidates = roster.get(record.normalized_email, [])
        if record.class_code and candidates:
            code = record.class_code.strip().lower()
            candidates = [item for item in candidates if str(item['class'].class_code or '').strip().lower() == code]
        unique_pairs = {(item['student'].id, item['class'].id): item for item in candidates}
        if len(unique_pairs) == 1:
            selected = next(iter(unique_pairs.values()))
            return 'matched_roster', selected['student'], selected['class'], None
        if len(unique_pairs) > 1:
            return 'ambiguous', None, None, 'Email khớp nhiều lớp trong cùng môn/kỳ/Block. Hãy bổ sung cột Lớp hoặc kiểm tra roster AP.'
        global_candidates = self.db.query(AcademicStudent).filter(
            func.lower(func.trim(AcademicStudent.email)) == record.normalized_email,
            AcademicStudent.active.is_(True),
        ).limit(3).all()
        if len(global_candidates) == 1:
            return 'matched_student_outside_roster', global_candidates[0], None, 'Email có trong danh mục sinh viên nhưng chưa nằm trong roster AP của môn hiện tại.'
        if len(global_candidates) > 1:
            return 'ambiguous', None, None, 'Email trùng nhiều sinh viên trong dữ liệu học vụ.'
        return 'unmatched', None, None, 'Không tìm thấy email trong roster AP hoặc danh mục sinh viên.'

    def process_batch(self, batch: UdemyProgressImportBatch) -> dict[str, Any]:
        delivery = self.db.get(AcademicSubjectDelivery, batch.subject_delivery_id)
        subject = self.db.get(AcademicSubject, delivery.subject_id) if delivery else None
        if not delivery or not delivery.active or not subject:
            raise ValueError('Môn Udemy không còn tồn tại.')
        if delivery.learning_platform != 'udemy':
            raise ValueError('Môn đã được chuyển khỏi nền tảng Udemy; job được dừng để không ghi sai dữ liệu.')
        path = Path(str(batch.file_path or '')).expanduser().resolve()
        root = Path(settings.local_storage_path or '/app/.runtime').expanduser().resolve()
        try:
            path.relative_to(root)
        except ValueError as exc:
            raise ValueError('Đường dẫn file import nằm ngoài vùng lưu trữ cho phép.') from exc
        if not path.is_file():
            raise ValueError('File import không còn tồn tại trong vùng lưu trữ.')

        now = datetime.utcnow()
        plan, milestone = self._active_plan_context(delivery.id, now.date())
        parsed = self.parse_path(path, active_plan=plan)
        roster, _class_by_code = self._roster_maps(delivery)
        self.db.query(UdemyProgressUnmatchedRow).filter(UdemyProgressUnmatchedRow.batch_id == batch.id).delete(synchronize_session=False)

        counters = defaultdict(int)
        for issue in parsed['issues']:
            self.db.add(UdemyProgressUnmatchedRow(
                batch_id=batch.id, subject_delivery_id=delivery.id,
                row_number=issue.get('row_number'), email=issue.get('email'), display_name=issue.get('display_name'),
                raw_progress=issue.get('raw_progress'), normalized_progress=None,
                reason_code=issue.get('reason_code') or 'INVALID_ROW', reason_message=issue.get('reason_message') or 'Dòng không hợp lệ.',
                raw_json=json_safe_value({'source_file': batch.file_name}),
            ))
            counters['failed'] += 1

        for record in parsed['records']:
            match_status, student, class_row, diagnostic = self._resolve_student(record, roster)
            required = float(milestone['required_progress_percent']) if milestone else None
            is_late = None if milestone is None else bool(record.progress_percent < required)
            snapshot = self.db.query(UdemyStudentProgress).filter(
                UdemyStudentProgress.subject_delivery_id == delivery.id,
                UdemyStudentProgress.normalized_email == record.normalized_email,
            ).first()
            if not snapshot:
                snapshot = UdemyStudentProgress(
                    subject_delivery_id=delivery.id,
                    email=record.email,
                    normalized_email=record.normalized_email,
                    last_import_batch_id=batch.id,
                )
            snapshot.class_id = class_row.id if class_row else None
            snapshot.student_id = student.id if student else None
            snapshot.email = record.email
            snapshot.display_name = record.display_name or (student.full_name if student else None)
            snapshot.progress_percent = record.progress_percent
            snapshot.is_late = is_late
            snapshot.current_plan_week = milestone.get('week_number') if milestone else None
            snapshot.required_progress_percent = required
            snapshot.current_deadline_date = milestone.get('deadline_date') if milestone else None
            snapshot.match_status = match_status
            snapshot.source_format = parsed['parser_format']
            snapshot.last_import_batch_id = batch.id
            snapshot.last_imported_at = now
            snapshot.updated_at = now
            snapshot.metadata_json = json_safe_value({
                'source_file': batch.file_name,
                'source_rows': record.row_numbers,
                'source_row_count': record.source_row_count,
                'parser_format': parsed['parser_format'],
                'plan_id': plan.id if plan else None,
                'plan_version': plan.version if plan else None,
                'diagnostic': diagnostic,
                **record.metadata,
            })
            self.db.add(snapshot)
            if match_status == 'matched_roster':
                counters['matched'] += 1
            elif match_status == 'matched_student_outside_roster':
                counters['outside_roster'] += 1
            elif match_status == 'ambiguous':
                counters['ambiguous'] += 1
            else:
                counters['unmatched'] += 1
            if match_status != 'matched_roster':
                self.db.add(UdemyProgressUnmatchedRow(
                    batch_id=batch.id, subject_delivery_id=delivery.id,
                    row_number=record.row_numbers[0] if record.row_numbers else None,
                    email=record.email, display_name=record.display_name,
                    raw_progress=str(record.progress_percent), normalized_progress=record.progress_percent,
                    reason_code=match_status.upper(), reason_message=diagnostic or match_status,
                    raw_json=json_safe_value({'class_code': record.class_code, 'source_rows': record.row_numbers}),
                ))

        batch.parser_format = parsed['parser_format']
        batch.total_rows = int(parsed['read_rows'])
        batch.processed_rows = len(parsed['records'])
        batch.matched_rows = counters['matched']
        batch.outside_roster_rows = counters['outside_roster']
        batch.unmatched_rows = counters['unmatched']
        batch.ambiguous_rows = counters['ambiguous']
        batch.failed_rows = counters['failed']
        batch.result_json = json_safe_value({
            'ok': True,
            'subject_code': subject.subject_code,
            'subject_name': subject.subject_name,
            'parser_format': parsed['parser_format'],
            'total_source_rows': parsed['read_rows'],
            'student_records': len(parsed['records']),
            'matched_rows': counters['matched'],
            'outside_roster_rows': counters['outside_roster'],
            'unmatched_rows': counters['unmatched'],
            'ambiguous_rows': counters['ambiguous'],
            'failed_rows': counters['failed'],
            'late_rows': self.db.query(UdemyStudentProgress).filter(
                UdemyStudentProgress.subject_delivery_id == delivery.id,
                UdemyStudentProgress.is_late.is_(True),
            ).count(),
            'plan_id': plan.id if plan else None,
            'plan_version': plan.version if plan else None,
            'current_milestone': json_safe_value(milestone),
            'warnings': parsed['warnings'],
        })
        batch.status = 'completed'
        batch.error_message = None
        batch.finished_at = now
        batch.updated_at = now
        self.db.add(batch)
        self.db.flush()
        report_path = self._write_error_report(batch, subject)
        batch.error_report_path = str(report_path) if report_path else None
        self.db.add(batch)
        self.db.commit()
        self.db.refresh(batch)
        return dict(batch.result_json or {})

    def _write_error_report(self, batch: UdemyProgressImportBatch, subject: AcademicSubject) -> Path | None:
        rows = self.db.query(UdemyProgressUnmatchedRow).filter(
            UdemyProgressUnmatchedRow.batch_id == batch.id,
        ).order_by(UdemyProgressUnmatchedRow.row_number.asc().nullslast(), UdemyProgressUnmatchedRow.created_at.asc()).all()
        if not rows:
            return None
        root = Path(settings.local_storage_path or '/app/.runtime').expanduser().resolve()
        out_dir = root / 'udemy-progress-imports' / 'errors'
        out_dir.mkdir(parents=True, exist_ok=True)
        path = out_dir / f'udemy-progress-errors-{subject.subject_code}-{batch.id[:8]}.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.title = 'Lỗi import'
        headers = ['Dòng', 'Mã môn', 'Email', 'Họ tên', 'Tiến độ', 'Mã lỗi', 'Nội dung']
        ws.append(headers)
        for row in rows:
            ws.append([
                row.row_number, subject.subject_code, row.email or '', row.display_name or '',
                row.normalized_progress if row.normalized_progress is not None else row.raw_progress or '',
                row.reason_code, row.reason_message,
            ])
        fill = PatternFill('solid', fgColor='17365D')
        for cell in ws[1]:
            cell.fill = fill
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        widths = [10, 14, 34, 28, 14, 28, 70]
        for index, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + index)].width = width
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        wb.save(path)
        wb.close()
        return path

    def batch_to_dict(self, batch: UdemyProgressImportBatch) -> dict[str, Any]:
        delivery = self.db.get(AcademicSubjectDelivery, batch.subject_delivery_id)
        subject = self.db.get(AcademicSubject, delivery.subject_id) if delivery else None
        return {
            'id': batch.id,
            'parent_job_id': batch.parent_job_id,
            'subject_delivery_id': batch.subject_delivery_id,
            'subject_code': subject.subject_code if subject else None,
            'subject_name': subject.subject_name if subject else None,
            'duplicate_of_batch_id': batch.duplicate_of_batch_id,
            'file_name': batch.file_name,
            'file_hash': batch.file_hash,
            'file_size_bytes': batch.file_size_bytes,
            'parser_format': batch.parser_format,
            'status': batch.status,
            'force_reimport': batch.force_reimport,
            'total_rows': batch.total_rows,
            'processed_rows': batch.processed_rows,
            'matched_rows': batch.matched_rows,
            'outside_roster_rows': batch.outside_roster_rows,
            'unmatched_rows': batch.unmatched_rows,
            'ambiguous_rows': batch.ambiguous_rows,
            'failed_rows': batch.failed_rows,
            'requested_by': batch.requested_by,
            'result_json': batch.result_json if isinstance(batch.result_json, dict) else {},
            'error_message': batch.error_message,
            'error_report_available': bool(batch.error_report_path),
            'started_at': batch.started_at,
            'finished_at': batch.finished_at,
            'created_at': batch.created_at,
            'updated_at': batch.updated_at,
        }


    def _delivery_context(self, delivery_id: str) -> tuple[AcademicSubjectDelivery, AcademicSubject, AcademicTerm | None, AcademicBlock | None]:
        delivery = self.db.get(AcademicSubjectDelivery, delivery_id)
        if not delivery or not delivery.active:
            raise HTTPException(status_code=404, detail='Không tìm thấy môn Udemy.')
        subject = self.db.get(AcademicSubject, delivery.subject_id)
        if not subject:
            raise HTTPException(status_code=404, detail='Không tìm thấy môn học.')
        if delivery.learning_platform != 'udemy':
            raise HTTPException(status_code=409, detail='Môn này không được cấu hình trên nền tảng Udemy.')
        return delivery, subject, self.db.get(AcademicTerm, delivery.term_id), self.db.get(AcademicBlock, delivery.block_id)

    @staticmethod
    def _status_for_snapshot(
        snapshot: UdemyStudentProgress,
        *,
        current_required_percent: float | None = None,
    ) -> tuple[str, str]:
        if snapshot.match_status == 'ambiguous':
            return 'ambiguous', 'Cần đối chiếu'
        if snapshot.match_status == 'unmatched':
            return 'unmatched', 'Chưa khớp AP'
        if snapshot.match_status == 'matched_student_outside_roster':
            return 'outside_roster', 'Ngoài roster AP'
        if current_required_percent is None:
            return 'no_plan', 'Chưa có mốc đến hạn'
        if float(snapshot.progress_percent or 0) < current_required_percent:
            return 'late', 'Chậm tiến độ'
        return 'on_track', 'Đạt tiến độ'

    def _current_dashboard_milestone(
        self,
        delivery_id: str,
    ) -> tuple[UdemySubjectPlan | None, dict[str, Any] | None]:
        plan, context = self._active_plan_context(delivery_id, vn_now().date())
        # A plan may exist while its first deadline has not arrived. In that case
        # the latest progress snapshot is visible, but the system must not label
        # learners as late/on-track yet.
        if not context or context.get('deadline_date') is None:
            return plan, None
        return plan, context

    def _scoped_snapshot_query(self, delivery_id: str, allowed_class_ids: set[str] | None = None):
        query = self.db.query(UdemyStudentProgress).filter(
            UdemyStudentProgress.subject_delivery_id == delivery_id,
        )
        if allowed_class_ids is not None:
            # Scoped teachers/campus owners must never see unresolved rows that cannot
            # be tied to one of their AP classes.
            if not allowed_class_ids:
                return query.filter(UdemyStudentProgress.id == '__no_accessible_rows__')
            query = query.filter(UdemyStudentProgress.class_id.in_(sorted(allowed_class_ids)))
        return query

    def _class_teacher_map(self, class_ids: list[str]) -> dict[str, list[str]]:
        if not class_ids:
            return {}
        rows = self.db.query(AcademicTeacherAssignment, AcademicTeacher).join(
            AcademicTeacher, AcademicTeacher.id == AcademicTeacherAssignment.teacher_id,
        ).filter(AcademicTeacherAssignment.class_id.in_(class_ids)).all()
        result: dict[str, list[str]] = defaultdict(list)
        for assignment, teacher in rows:
            name = str(teacher.full_name or teacher.username or '').strip()
            if name and name not in result[assignment.class_id]:
                result[assignment.class_id].append(name)
        return dict(result)

    def dashboard(
        self,
        delivery_id: str,
        *,
        allowed_class_ids: set[str] | None = None,
        scope_label: str = 'Toàn bộ môn',
    ) -> dict[str, Any]:
        delivery, subject, term, block = self._delivery_context(delivery_id)
        query = self._scoped_snapshot_query(delivery_id, allowed_class_ids)
        rows = query.all()
        active_plan, milestone = self._current_dashboard_milestone(delivery_id)
        current_required = float(milestone['required_progress_percent']) if milestone else None
        statuses = defaultdict(int)
        matched = outside = ambiguous = unmatched = 0
        total_progress = 0.0
        latest = None
        for row in rows:
            status_key, _ = self._status_for_snapshot(row, current_required_percent=current_required)
            statuses[status_key] += 1
            matched += int(row.match_status == 'matched_roster')
            outside += int(row.match_status == 'matched_student_outside_roster')
            ambiguous += int(row.match_status == 'ambiguous')
            unmatched += int(row.match_status == 'unmatched')
            total_progress += float(row.progress_percent or 0)
            if row.last_imported_at and (latest is None or row.last_imported_at > latest):
                latest = row.last_imported_at

        class_query = self.db.query(AcademicClass).filter(
            AcademicClass.subject_id == delivery.subject_id,
            AcademicClass.term_id == delivery.term_id,
            AcademicClass.block_id == delivery.block_id,
            func.lower(func.coalesce(AcademicClass.branch, delivery.branch)) == str(delivery.branch).lower(),
            AcademicClass.active.is_(True),
        )
        if allowed_class_ids is not None:
            class_query = class_query.filter(AcademicClass.id.in_(sorted(allowed_class_ids or {'__none__'})))
        classes = class_query.order_by(AcademicClass.class_code.asc()).all()

        recent_batches = self.db.query(UdemyProgressImportBatch).filter(
            UdemyProgressImportBatch.subject_delivery_id == delivery_id,
        ).order_by(UdemyProgressImportBatch.created_at.desc()).limit(10).all()
        return {
            'delivery': {
                'id': delivery.id,
                'subject_id': subject.id,
                'subject_code': subject.subject_code,
                'subject_name': subject.subject_name,
                'term_id': delivery.term_id,
                'term_name': term.term_name if term else '',
                'block_id': delivery.block_id,
                'block_name': block.block_name if block else '',
                'branch': delivery.branch,
                'learning_platform': delivery.learning_platform,
            },
            'summary': {
                'subject_delivery_id': delivery_id,
                'total_students': len(rows),
                'matched_students': matched,
                'outside_roster_students': outside,
                'ambiguous_students': ambiguous,
                'unmatched_students': unmatched,
                'late_students': statuses['late'],
                'on_track_students': statuses['on_track'],
                'no_plan_students': statuses['no_plan'],
                'average_progress_percent': round(total_progress / len(rows), 2) if rows else None,
                'required_progress_percent': current_required,
                'current_plan_week': milestone.get('week_number') if milestone else None,
                'current_deadline_date': milestone.get('deadline_date') if milestone else None,
                'last_imported_at': latest,
                'class_count': len(classes),
                'scope_label': scope_label,
            },
            'active_plan': {
                'id': active_plan.id,
                'version': active_plan.version,
                'item_count': active_plan.item_count,
                'source': active_plan.source,
                'imported_at': active_plan.imported_at,
            } if active_plan else None,
            'classes': [
                {
                    'id': item.id,
                    'class_code': item.class_code,
                    'class_name': item.class_name,
                    'campus': item.campus,
                }
                for item in classes
            ],
            'recent_imports': [self.batch_to_dict(item) for item in recent_batches] if allowed_class_ids is None else [],
        }

    def list_students(
        self,
        delivery_id: str,
        *,
        allowed_class_ids: set[str] | None = None,
        q: str | None = None,
        class_id: str | None = None,
        status_filter: str | None = None,
        page: int = 1,
        page_size: int = 50,
        sort_by: str = 'student',
        sort_dir: str = 'asc',
    ) -> dict[str, Any]:
        self._delivery_context(delivery_id)
        _active_plan, milestone = self._current_dashboard_milestone(delivery_id)
        current_required = float(milestone['required_progress_percent']) if milestone else None
        query = self._scoped_snapshot_query(delivery_id, allowed_class_ids)
        if class_id:
            if allowed_class_ids is not None and class_id not in allowed_class_ids:
                raise HTTPException(status_code=403, detail='Bạn không được xem lớp này.')
            query = query.filter(UdemyStudentProgress.class_id == class_id)
        search = str(q or '').strip().lower()
        if search:
            like = f'%{search}%'
            query = query.outerjoin(AcademicStudent, AcademicStudent.id == UdemyStudentProgress.student_id).outerjoin(
                AcademicClass, AcademicClass.id == UdemyStudentProgress.class_id,
            ).filter(or_(
                func.lower(UdemyStudentProgress.email).like(like),
                func.lower(func.coalesce(UdemyStudentProgress.display_name, '')).like(like),
                func.lower(func.coalesce(AcademicStudent.student_code, '')).like(like),
                func.lower(func.coalesce(AcademicStudent.username, '')).like(like),
                func.lower(func.coalesce(AcademicClass.class_code, '')).like(like),
            ))
        normalized_status = str(status_filter or 'all').strip().lower()
        if normalized_status == 'late':
            query = query.filter(
                UdemyStudentProgress.match_status == 'matched_roster',
                UdemyStudentProgress.progress_percent < current_required,
            ) if current_required is not None else query.filter(UdemyStudentProgress.id == '__no_late_rows__')
        elif normalized_status == 'on_track':
            query = query.filter(
                UdemyStudentProgress.match_status == 'matched_roster',
                UdemyStudentProgress.progress_percent >= current_required,
            ) if current_required is not None else query.filter(UdemyStudentProgress.id == '__no_on_track_rows__')
        elif normalized_status == 'no_plan':
            query = query.filter(UdemyStudentProgress.match_status == 'matched_roster') if current_required is None else query.filter(UdemyStudentProgress.id == '__no_plan_rows__')
        elif normalized_status == 'unmatched':
            query = query.filter(UdemyStudentProgress.match_status == 'unmatched')
        elif normalized_status == 'ambiguous':
            query = query.filter(UdemyStudentProgress.match_status == 'ambiguous')
        elif normalized_status == 'outside_roster':
            query = query.filter(UdemyStudentProgress.match_status == 'matched_student_outside_roster')
        elif normalized_status == 'alerts':
            predicates = [UdemyStudentProgress.match_status != 'matched_roster']
            if current_required is None:
                predicates.append(UdemyStudentProgress.match_status == 'matched_roster')
            else:
                predicates.append(UdemyStudentProgress.progress_percent < current_required)
            query = query.filter(or_(*predicates))
        elif normalized_status not in {'', 'all'}:
            raise HTTPException(status_code=400, detail='Bộ lọc trạng thái Udemy không hợp lệ.')

        total = query.count()
        direction_desc = str(sort_dir or 'asc').lower() == 'desc'
        sort_columns = {
            'student': func.lower(func.coalesce(UdemyStudentProgress.display_name, UdemyStudentProgress.email)),
            'progress': UdemyStudentProgress.progress_percent,
            'required': UdemyStudentProgress.required_progress_percent,
            'class': UdemyStudentProgress.class_id,
            'updated': UdemyStudentProgress.last_imported_at,
            'status': UdemyStudentProgress.progress_percent,
        }
        sort_column = sort_columns.get(sort_by, sort_columns['student'])
        query = query.order_by(sort_column.desc() if direction_desc else sort_column.asc(), UdemyStudentProgress.id.asc())
        clean_page = max(1, int(page or 1))
        clean_size = max(1, min(200, int(page_size or 50)))
        snapshots = query.offset((clean_page - 1) * clean_size).limit(clean_size).all()

        student_ids = [item.student_id for item in snapshots if item.student_id]
        class_ids = [item.class_id for item in snapshots if item.class_id]
        students = {item.id: item for item in self.db.query(AcademicStudent).filter(AcademicStudent.id.in_(student_ids or ['__none__'])).all()}
        classes = {item.id: item for item in self.db.query(AcademicClass).filter(AcademicClass.id.in_(class_ids or ['__none__'])).all()}
        teachers = self._class_teacher_map(class_ids)
        items: list[dict[str, Any]] = []
        for snapshot in snapshots:
            student = students.get(snapshot.student_id)
            class_row = classes.get(snapshot.class_id)
            status_key, status_label = self._status_for_snapshot(snapshot, current_required_percent=current_required)
            metadata = snapshot.metadata_json if isinstance(snapshot.metadata_json, dict) else {}
            required = current_required
            items.append({
                'id': snapshot.id,
                'student_id': snapshot.student_id,
                'student_code': student.student_code if student else None,
                'student_username': student.username if student else None,
                'display_name': snapshot.display_name or (student.full_name if student else None) or snapshot.email,
                'email': snapshot.email,
                'class_id': snapshot.class_id,
                'class_code': class_row.class_code if class_row else None,
                'class_name': class_row.class_name if class_row else None,
                'campus': class_row.campus if class_row else (student.campus if student else None),
                'teacher_names': teachers.get(snapshot.class_id or '', []),
                'progress_percent': round(float(snapshot.progress_percent or 0), 2),
                'required_progress_percent': required,
                'variance_percent': round(float(snapshot.progress_percent or 0) - required, 2) if required is not None else None,
                'is_late': status_key == 'late' if status_key in {'late', 'on_track'} else None,
                'status': status_key,
                'status_label': status_label,
                'match_status': snapshot.match_status,
                'current_plan_week': milestone.get('week_number') if milestone else None,
                'current_deadline_date': milestone.get('deadline_date') if milestone else None,
                'last_import_batch_id': snapshot.last_import_batch_id,
                'source_format': snapshot.source_format,
                'last_imported_at': snapshot.last_imported_at,
                'diagnostic': metadata.get('diagnostic'),
            })
        total_pages = max(1, math.ceil(total / clean_size)) if total else 0
        return {
            'items': items,
            'total': total,
            'page': clean_page,
            'page_size': clean_size,
            'total_pages': total_pages,
            'has_next': clean_page < total_pages,
        }

    def export_workbook(
        self,
        delivery_id: str,
        *,
        allowed_class_ids: set[str] | None = None,
        scope_label: str = 'Toàn bộ môn',
        q: str | None = None,
        class_id: str | None = None,
        status_filter: str | None = None,
        max_rows: int | None = None,
    ) -> bytes:
        context = self.dashboard(delivery_id, allowed_class_ids=allowed_class_ids, scope_label=scope_label)
        all_rows: list[dict[str, Any]] = []
        page = 1
        while True:
            chunk = self.list_students(
                delivery_id,
                allowed_class_ids=allowed_class_ids,
                q=q,
                class_id=class_id,
                status_filter=status_filter,
                page=page,
                page_size=200,
                sort_by='student',
                sort_dir='asc',
            )
            all_rows.extend(chunk['items'])
            if max_rows is not None and len(all_rows) > int(max_rows):
                raise ValueError('Số dòng export vượt giới hạn đồng bộ; hãy dùng job export nền.')
            if not chunk['has_next']:
                break
            page += 1

        wb = Workbook()
        ws = wb.active
        ws.title = 'TongQuan'
        delivery = context['delivery']
        summary = context['summary']
        overview_rows = [
            ['BÁO CÁO TIẾN ĐỘ UDEMY', ''],
            ['Môn', f"{delivery['subject_code']} - {delivery['subject_name']}"],
            ['Học kỳ / Block', f"{delivery['term_name']} / {delivery['block_name']}"],
            ['Hệ', str(delivery['branch']).upper()],
            ['Phạm vi dữ liệu', summary['scope_label']],
            ['Tổng sinh viên', summary['total_students']],
            ['Đạt tiến độ', summary['on_track_students']],
            ['Chậm tiến độ', summary['late_students']],
            ['Chưa có kế hoạch', summary['no_plan_students']],
            ['Chưa khớp / cần đối chiếu', summary['unmatched_students'] + summary['ambiguous_students'] + summary['outside_roster_students']],
            ['Tiến độ trung bình (%)', summary['average_progress_percent']],
            ['Mốc yêu cầu hiện tại (%)', summary['required_progress_percent']],
            ['Tuần kế hoạch hiện tại', summary['current_plan_week']],
            ['Deadline hiện tại', summary['current_deadline_date']],
            ['Lần import gần nhất', summary['last_imported_at']],
        ]
        for row in overview_rows:
            ws.append(row)
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill('solid', fgColor='17365D')
        ws.merge_cells('A1:B1')
        for cell in ws['A'][1:]:
            cell.font = Font(bold=True)
        ws.column_dimensions['A'].width = 34
        ws.column_dimensions['B'].width = 55

        detail = wb.create_sheet('TienDoSinhVien')
        headers = ['STT', 'Mã SV', 'Username', 'Họ tên', 'Email Udemy', 'Lớp', 'Cơ sở', 'Giảng viên', 'Tiến độ (%)', 'Yêu cầu (%)', 'Chênh lệch', 'Tuần', 'Deadline', 'Trạng thái', 'Đối chiếu AP', 'Cập nhật gần nhất', 'Ghi chú']
        detail.append(headers)
        for index, item in enumerate(all_rows, 1):
            detail.append([
                index, item.get('student_code'), item.get('student_username'), item.get('display_name'), item.get('email'),
                item.get('class_code'), item.get('campus'), ', '.join(item.get('teacher_names') or []),
                item.get('progress_percent'), item.get('required_progress_percent'), item.get('variance_percent'),
                item.get('current_plan_week'), item.get('current_deadline_date'), item.get('status_label'),
                item.get('match_status'), item.get('last_imported_at'), item.get('diagnostic'),
            ])

        warning = wb.create_sheet('CanhBao')
        warning.append(headers)
        warning_index = 0
        for item in all_rows:
            if item.get('status') not in {'late', 'unmatched', 'ambiguous', 'outside_roster', 'no_plan'}:
                continue
            warning_index += 1
            warning.append([
                warning_index, item.get('student_code'), item.get('student_username'), item.get('display_name'), item.get('email'),
                item.get('class_code'), item.get('campus'), ', '.join(item.get('teacher_names') or []),
                item.get('progress_percent'), item.get('required_progress_percent'), item.get('variance_percent'),
                item.get('current_plan_week'), item.get('current_deadline_date'), item.get('status_label'),
                item.get('match_status'), item.get('last_imported_at'), item.get('diagnostic'),
            ])

        history = wb.create_sheet('LichSuImport')
        history_headers = ['Thời điểm', 'File', 'Trạng thái', 'Định dạng', 'Tổng dòng', 'Bản ghi SV', 'Khớp roster', 'Ngoài roster', 'Không khớp', 'Mơ hồ', 'Dòng lỗi', 'Người thực hiện', 'Thông báo']
        history.append(history_headers)
        for item in context['recent_imports']:
            history.append([
                item.get('created_at'), item.get('file_name'), item.get('status'), item.get('parser_format'), item.get('total_rows'),
                item.get('processed_rows'), item.get('matched_rows'), item.get('outside_roster_rows'), item.get('unmatched_rows'),
                item.get('ambiguous_rows'), item.get('failed_rows'), item.get('requested_by'), item.get('error_message'),
            ])

        guide = wb.create_sheet('HuongDan')
        guide['A1'] = 'Hướng dẫn đọc báo cáo Udemy'
        guide['A1'].font = Font(size=15, bold=True)
        notes = [
            'Đạt tiến độ: tiến độ hiện tại lớn hơn hoặc bằng mốc kế hoạch gần nhất đã đến hạn.',
            'Chậm tiến độ: tiến độ hiện tại thấp hơn mốc kế hoạch gần nhất đã đến hạn.',
            'Chưa có kế hoạch: môn chưa có mốc kế hoạch đã đến hạn nên hệ thống chưa kết luận đạt/chậm.',
            'Ngoài roster AP: email có trong danh mục sinh viên nhưng chưa thuộc danh sách lớp AP của môn hiện tại.',
            'Cần đối chiếu: email khớp nhiều lớp hoặc nhiều sinh viên; quản trị cần kiểm tra lại dữ liệu AP/Udemy.',
            f"Báo cáo chỉ chứa dữ liệu trong phạm vi quyền: {scope_label}.",
        ]
        for index, note in enumerate(notes, 3):
            guide.cell(row=index, column=1, value=f'- {note}')
        guide.column_dimensions['A'].width = 110

        header_fill = PatternFill('solid', fgColor='17365D')
        widths = [8, 15, 20, 28, 34, 16, 12, 28, 14, 14, 14, 10, 14, 18, 22, 20, 60]
        for sheet in [detail, warning]:
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = Font(color='FFFFFF', bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for idx, width in enumerate(widths, 1):
                sheet.column_dimensions[chr(64 + idx) if idx <= 26 else 'A'].width = width
            sheet.freeze_panes = 'A2'
            sheet.auto_filter.ref = sheet.dimensions
        for cell in history[1]:
            cell.fill = header_fill
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for idx, width in enumerate([20, 36, 16, 20, 12, 12, 12, 12, 12, 12, 12, 22, 55], 1):
            history.column_dimensions[chr(64 + idx)].width = width
        history.freeze_panes = 'A2'
        history.auto_filter.ref = history.dimensions
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
        raw = BytesIO()
        wb.save(raw)
        wb.close()
        return raw.getvalue()

    def current_summary(self, delivery_id: str) -> dict[str, Any]:
        delivery = self.db.get(AcademicSubjectDelivery, delivery_id)
        if not delivery:
            raise HTTPException(status_code=404, detail='Không tìm thấy môn.')
        query = self.db.query(UdemyStudentProgress).filter(UdemyStudentProgress.subject_delivery_id == delivery_id)
        total = query.count()
        late = query.filter(UdemyStudentProgress.is_late.is_(True)).count()
        matched = query.filter(UdemyStudentProgress.match_status == 'matched_roster').count()
        outside = query.filter(UdemyStudentProgress.match_status == 'matched_student_outside_roster').count()
        ambiguous = query.filter(UdemyStudentProgress.match_status == 'ambiguous').count()
        unmatched = query.filter(UdemyStudentProgress.match_status == 'unmatched').count()
        avg = self.db.query(func.avg(UdemyStudentProgress.progress_percent)).filter(UdemyStudentProgress.subject_delivery_id == delivery_id).scalar()
        latest = self.db.query(func.max(UdemyStudentProgress.last_imported_at)).filter(UdemyStudentProgress.subject_delivery_id == delivery_id).scalar()
        return {
            'subject_delivery_id': delivery_id, 'total_students': total, 'matched_students': matched,
            'outside_roster_students': outside, 'ambiguous_students': ambiguous, 'unmatched_students': unmatched,
            'late_students': late, 'average_progress_percent': round(float(avg), 2) if avg is not None else None,
            'last_imported_at': latest,
        }
