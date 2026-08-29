from __future__ import annotations

import hashlib
import math
import re
import secrets
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from typing import Any

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.core.json_safe import json_safe_value
from app.services.object_storage import StorageError, get_object_storage
from app.models.academic import (
    AcademicBlock,
    AcademicSubject,
    AcademicSubjectDelivery,
    AcademicTerm,
    UdemySubjectPlan,
    UdemySubjectPlanMilestone,
)


@dataclass(frozen=True)
class ParsedMilestone:
    week_number: int
    deadline_date: date
    required_progress_percent: float


class UdemyPlanService:
    PREVIEW_TTL_HOURS = 2
    MAX_UPLOAD_BYTES = 10 * 1024 * 1024
    MAX_ROWS = 2000
    MAX_WEEKS = 52
    MAX_XLSX_ENTRIES = 5000
    MAX_XLSX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
    REQUIRED_HEADERS = {'term_name', 'block_name', 'subject_code', 'item_count'}

    HEADER_ALIASES = {
        'stt': {'stt', 'no', 'so thu tu', 'số thứ tự'},
        'term_name': {'hoc ky', 'học kỳ', 'semester', 'term'},
        'block_name': {'block', 'khoi', 'khối'},
        'subject_code': {'ma mon hoc', 'mã môn học', 'ma mon', 'mã môn', 'subject code', 'subject_code'},
        'subject_name': {'ten mon hoc', 'tên môn học', 'ten mon', 'tên môn', 'subject name', 'subject_name'},
        'item_count': {'so luong item', 'số lượng item', 'items', 'item count', 'item_count'},
    }

    def __init__(self, db: Session):
        self.db = db

    @staticmethod
    def normalize_text(value: Any) -> str:
        import unicodedata
        text = unicodedata.normalize('NFKD', str(value or '').strip().lower())
        text = ''.join(char for char in text if not unicodedata.combining(char))
        # Vietnamese đ/Đ does not decompose under NFKD, so normalize it explicitly.
        text = text.replace('đ', 'd')
        text = re.sub(r'[^a-z0-9]+', ' ', text).strip()
        return text

    @staticmethod
    def normalize_branch(value: str | None) -> str:
        normalized = str(value or 'poly').strip().lower()
        if normalized not in {'poly', 'ptcd'}:
            raise HTTPException(status_code=422, detail='Hệ chỉ nhận poly hoặc ptcd.')
        return normalized

    @classmethod
    def _canonical_header(cls, value: Any) -> tuple[str | None, int | None]:
        normalized = cls.normalize_text(value)
        for canonical, aliases in cls.HEADER_ALIASES.items():
            if normalized in {cls.normalize_text(item) for item in aliases}:
                return canonical, None
        match = re.fullmatch(r'week\s*(\d+)', normalized)
        if match:
            return 'deadline', int(match.group(1))
        match = re.fullmatch(r'(?:tien do|tiendo|progress)\s*week\s*(\d+)', normalized)
        if match:
            return 'progress', int(match.group(1))
        return None, None

    @classmethod
    def _find_header_row(cls, worksheet) -> tuple[int, dict[str, int], dict[int, dict[str, int]]]:
        for row_no in range(1, min(12, worksheet.max_row) + 1):
            base_columns: dict[str, int] = {}
            week_columns: dict[int, dict[str, int]] = {}
            for col_no in range(1, worksheet.max_column + 1):
                canonical, week_no = cls._canonical_header(worksheet.cell(row=row_no, column=col_no).value)
                if not canonical:
                    continue
                if canonical in {'deadline', 'progress'} and week_no:
                    if week_no > cls.MAX_WEEKS:
                        continue
                    week_columns.setdefault(week_no, {})[canonical] = col_no
                else:
                    base_columns[canonical] = col_no
            if cls.REQUIRED_HEADERS.issubset(base_columns) and week_columns:
                return row_no, base_columns, week_columns
        raise ValueError('Không tìm thấy dòng tiêu đề hợp lệ. File phải có Học kỳ, Block, Mã môn học, Số lượng Item và các cột Week/Tiến độ week.')

    @staticmethod
    def _cell_text(cell) -> str:
        value = cell.value
        return '' if value is None else str(value).strip()

    @classmethod
    def _parse_positive_int(cls, cell, *, field_label: str) -> int:
        value = cell.value
        if value is None or str(value).strip() == '':
            raise ValueError(f'{field_label} không được để trống.')
        if isinstance(value, bool):
            raise ValueError(f'{field_label} phải là số nguyên lớn hơn 0.')
        try:
            number = float(value)
        except (TypeError, ValueError) as exc:
            raise ValueError(f'{field_label} phải là số nguyên lớn hơn 0.') from exc
        if not math.isfinite(number) or number <= 0 or not number.is_integer():
            raise ValueError(f'{field_label} phải là số nguyên lớn hơn 0.')
        return int(number)

    @classmethod
    def _parse_percent(cls, cell) -> float:
        value = cell.value
        if value is None or str(value).strip() == '':
            raise ValueError('Tiến độ không được để trống khi đã có deadline.')
        if isinstance(value, str):
            text = value.strip().replace(',', '.')
            is_percent = text.endswith('%')
            text = text[:-1].strip() if is_percent else text
            try:
                number = float(text)
            except ValueError as exc:
                raise ValueError('Tiến độ phải là số từ 0 đến 100.') from exc
            if is_percent:
                number = number
        else:
            try:
                number = float(value)
            except (TypeError, ValueError) as exc:
                raise ValueError('Tiến độ phải là số từ 0 đến 100.') from exc
            if '%' in str(getattr(cell, 'number_format', '') or '') and 0 <= number <= 1:
                number *= 100
        if not math.isfinite(number) or number < 0 or number > 100:
            raise ValueError('Tiến độ phải nằm trong khoảng 0–100.')
        return round(number, 2)

    @classmethod
    def _parse_date(cls, cell) -> date:
        value = cell.value
        if value is None or str(value).strip() == '':
            raise ValueError('Deadline không được để trống khi đã có tiến độ.')
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            try:
                converted = from_excel(value, cell.parent.parent.epoch)
                return converted.date() if isinstance(converted, datetime) else converted
            except Exception as exc:
                raise ValueError('Deadline dạng số Excel không hợp lệ.') from exc
        text = str(value).strip()
        formats = (
            '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%Y/%m/%d',
            '%d/%m/%y', '%d-%m-%y', '%Y-%m-%d %H:%M:%S',
        )
        for fmt in formats:
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        raise ValueError(f'Deadline “{text}” không đúng định dạng ngày. Dùng dd/mm/yyyy hoặc yyyy-mm-dd.')

    def _resolve_delivery(self, *, term_name: str, block_name: str, subject_code: str, branch: str) -> tuple[AcademicSubjectDelivery | None, list[str]]:
        messages: list[str] = []
        term_key = self.normalize_text(term_name)
        block_key = self.normalize_text(block_name)
        subject_key = str(subject_code or '').strip().upper()
        terms = self.db.query(AcademicTerm).filter(func.lower(func.coalesce(AcademicTerm.branch, branch)) == branch).all()
        term = next((item for item in terms if term_key in {self.normalize_text(item.term_name), self.normalize_text(item.term_code), self.normalize_text(item.ap_term_id)}), None)
        if not term:
            return None, [f'Không tìm thấy học kỳ “{term_name}” trong hệ {branch.upper()}.']
        blocks = self.db.query(AcademicBlock).filter(AcademicBlock.term_id == term.id).all()
        block = next((item for item in blocks if block_key in {self.normalize_text(item.block_name), self.normalize_text(item.block_code), self.normalize_text(item.ap_block_id)}), None)
        if not block:
            return None, [f'Không tìm thấy Block “{block_name}” trong học kỳ {term.term_name}.']
        subject = self.db.query(AcademicSubject).filter(
            func.upper(AcademicSubject.subject_code) == subject_key,
            func.lower(func.coalesce(AcademicSubject.branch, branch)) == branch,
        ).first()
        if not subject:
            return None, [f'Không tìm thấy môn {subject_key} trong hệ {branch.upper()}. Hãy lấy danh sách môn từ AP trước.']
        delivery = self.db.query(AcademicSubjectDelivery).filter(
            AcademicSubjectDelivery.subject_id == subject.id,
            AcademicSubjectDelivery.term_id == term.id,
            AcademicSubjectDelivery.block_id == block.id,
            func.lower(AcademicSubjectDelivery.branch) == branch,
            AcademicSubjectDelivery.active.is_(True),
        ).first()
        if not delivery:
            return None, [f'Môn {subject_key} chưa có bản ghi triển khai cho {term.term_name} · {block.block_name}.']
        if delivery.learning_platform != 'udemy':
            current = 'Chưa chọn' if delivery.learning_platform is None else str(delivery.learning_platform).upper()
            messages.append(f'Môn {subject_key} đang ở nền tảng {current}. Hãy chuyển sang Udemy trước khi import kế hoạch.')
        return delivery, messages

    def parse_workbook(self, raw: bytes, *, filename: str, branch: str, requested_by: str | None) -> dict[str, Any]:
        if not raw:
            raise ValueError('File Excel rỗng.')
        if len(raw) > self.MAX_UPLOAD_BYTES:
            raise ValueError('File import vượt giới hạn 10 MB.')
        try:
            with zipfile.ZipFile(BytesIO(raw)) as archive:
                entries = archive.infolist()
                if len(entries) > self.MAX_XLSX_ENTRIES:
                    raise ValueError('File Excel có quá nhiều thành phần, không an toàn để xử lý.')
                uncompressed_bytes = sum(max(0, int(item.file_size or 0)) for item in entries)
                if uncompressed_bytes > self.MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise ValueError('Dữ liệu giải nén của file Excel vượt giới hạn 100 MB.')
            workbook = load_workbook(BytesIO(raw), data_only=True, read_only=True)
        except ValueError:
            raise
        except Exception as exc:
            raise ValueError('Không đọc được file Excel. Hãy dùng file .xlsx hợp lệ.') from exc
        if not workbook.worksheets:
            workbook.close()
            raise ValueError('File Excel không có sheet dữ liệu.')
        worksheet = workbook.worksheets[0]
        try:
            header_row, base_columns, week_columns = self._find_header_row(worksheet)
        except Exception:
            workbook.close()
            raise
        branch_value = self.normalize_branch(branch)
        errors: list[dict[str, Any]] = []
        warnings: list[dict[str, Any]] = []
        rows: list[dict[str, Any]] = []
        seen_delivery_ids: set[str] = set()

        for row_no in range(header_row + 1, min(worksheet.max_row, header_row + self.MAX_ROWS) + 1):
            subject_code = self._cell_text(worksheet.cell(row=row_no, column=base_columns['subject_code'])).upper()
            term_name = self._cell_text(worksheet.cell(row=row_no, column=base_columns['term_name']))
            block_name = self._cell_text(worksheet.cell(row=row_no, column=base_columns['block_name']))
            item_cell = worksheet.cell(row=row_no, column=base_columns['item_count'])
            subject_name = self._cell_text(worksheet.cell(row=row_no, column=base_columns.get('subject_name', base_columns['subject_code'])))
            if not any([subject_code, term_name, block_name, item_cell.value not in {None, ''}]):
                continue
            row_errors: list[str] = []
            row_warnings: list[str] = []
            if not subject_code:
                row_errors.append('Thiếu mã môn học.')
            if not term_name:
                row_errors.append('Thiếu học kỳ.')
            if not block_name:
                row_errors.append('Thiếu Block.')
            try:
                item_count = self._parse_positive_int(item_cell, field_label='Số lượng Item')
            except ValueError as exc:
                item_count = 0
                row_errors.append(str(exc))

            milestones: list[ParsedMilestone] = []
            for week_no in sorted(week_columns):
                columns = week_columns[week_no]
                deadline_cell = worksheet.cell(row=row_no, column=columns.get('deadline', 0)) if columns.get('deadline') else None
                progress_cell = worksheet.cell(row=row_no, column=columns.get('progress', 0)) if columns.get('progress') else None
                deadline_blank = deadline_cell is None or deadline_cell.value is None or str(deadline_cell.value).strip() == ''
                progress_blank = progress_cell is None or progress_cell.value is None or str(progress_cell.value).strip() == ''
                if deadline_blank and progress_blank:
                    continue
                if deadline_blank != progress_blank:
                    row_errors.append(f'Week {week_no}: phải nhập đồng thời deadline và tiến độ.')
                    continue
                try:
                    deadline = self._parse_date(deadline_cell)
                    progress = self._parse_percent(progress_cell)
                    milestones.append(ParsedMilestone(week_no, deadline, progress))
                except ValueError as exc:
                    row_errors.append(f'Week {week_no}: {exc}')
            if not milestones:
                row_errors.append('Phải có ít nhất một mốc Week/tiến độ.')
            for index, milestone in enumerate(milestones):
                if index > 0:
                    previous = milestones[index - 1]
                    if milestone.deadline_date <= previous.deadline_date:
                        row_errors.append(f'Week {milestone.week_number}: deadline phải sau Week {previous.week_number}.')
                    if milestone.required_progress_percent < previous.required_progress_percent:
                        row_errors.append(f'Week {milestone.week_number}: tiến độ không được giảm so với Week {previous.week_number}.')
            if milestones and milestones[-1].required_progress_percent < 100:
                row_warnings.append(f'Mốc cuối hiện là {milestones[-1].required_progress_percent:g}%, chưa đạt 100%.')

            delivery = None
            resolution_errors: list[str] = []
            if subject_code and term_name and block_name:
                delivery, resolution_errors = self._resolve_delivery(term_name=term_name, block_name=block_name, subject_code=subject_code, branch=branch_value)
                row_errors.extend(resolution_errors)
            if delivery and delivery.id in seen_delivery_ids:
                row_errors.append('Môn này xuất hiện nhiều hơn một lần trong file.')
            if delivery:
                seen_delivery_ids.add(delivery.id)

            active_plan = None
            if delivery:
                active_plan = self.db.query(UdemySubjectPlan).filter(
                    UdemySubjectPlan.subject_delivery_id == delivery.id,
                    UdemySubjectPlan.active.is_(True),
                ).order_by(UdemySubjectPlan.version.desc()).first()
            row_payload = {
                'row_no': row_no,
                'delivery_id': delivery.id if delivery else None,
                'term_name': term_name,
                'block_name': block_name,
                'branch': branch_value,
                'subject_code': subject_code,
                'subject_name': subject_name,
                'item_count': item_count,
                'current_version': int(active_plan.version) if active_plan else None,
                'next_version': int(active_plan.version + 1) if active_plan else 1,
                'action': 'new_version' if active_plan else 'create',
                'milestones': [
                    {
                        'week_number': item.week_number,
                        'deadline_date': item.deadline_date.isoformat(),
                        'required_progress_percent': item.required_progress_percent,
                    }
                    for item in milestones
                ],
                'errors': row_errors,
                'warnings': row_warnings,
            }
            rows.append(row_payload)
            for message in row_errors:
                errors.append({'row': row_no, 'subject_code': subject_code or None, 'code': 'INVALID_PLAN_ROW', 'message': message})
            for message in row_warnings:
                warnings.append({'row': row_no, 'subject_code': subject_code or None, 'code': 'PLAN_WARNING', 'message': message})

        if worksheet.max_row > header_row + self.MAX_ROWS:
            errors.append({'row': header_row + self.MAX_ROWS + 1, 'subject_code': None, 'code': 'ROW_LIMIT', 'message': f'File vượt giới hạn {self.MAX_ROWS} dòng dữ liệu.'})
        valid_rows = [item for item in rows if not item['errors']]
        file_hash = hashlib.sha256(raw).hexdigest()
        workbook.close()
        return {
            'schema_version': 'udemy-plan-preview/v1',
            'policy_version': 'udemy-plan-management/batch32',
            'filename': filename,
            'file_sha256': file_hash,
            'branch': branch_value,
            'requested_by': requested_by,
            'created_at': datetime.now(timezone.utc).isoformat(),
            'header_row': header_row,
            'total_rows': len(rows),
            'valid_count': len(valid_rows),
            'error_count': len(errors),
            'warning_count': len(warnings),
            'rows': rows,
            'errors': errors,
            'warnings': warnings,
            'can_commit': bool(valid_rows) and not errors,
        }

    @staticmethod
    def _preview_key(token: str) -> str:
        return f'udemy-plan-previews/{token}.json'

    @classmethod
    def _cleanup_expired_previews(cls) -> None:
        cutoff = datetime.now(timezone.utc) - timedelta(hours=cls.PREVIEW_TTL_HOURS)
        storage = get_object_storage()
        try:
            items = storage.list_objects('udemy-plan-previews')
        except StorageError:
            return
        for item in items:
            try:
                modified = item.last_modified
                if modified is not None and modified.tzinfo is None:
                    modified = modified.replace(tzinfo=timezone.utc)
                if modified is not None and modified < cutoff:
                    storage.delete(item.reference, missing_ok=True)
            except (StorageError, TypeError):
                continue

    def persist_preview(self, preview: dict[str, Any]) -> str:
        self._cleanup_expired_previews()
        token = secrets.token_hex(16)
        stored_preview = dict(preview)
        stored_preview['preview_token'] = token
        get_object_storage().put_json(self._preview_key(token), json_safe_value(stored_preview))
        return token

    def load_preview(self, token: str, *, requested_by: str | None) -> tuple[str, dict[str, Any]]:
        if not re.fullmatch(r'[0-9a-f]{32}', str(token or '').lower()):
            raise HTTPException(status_code=400, detail='Mã preview kế hoạch Udemy không hợp lệ.')
        storage = get_object_storage()
        reference = storage.reference_for_key(self._preview_key(token))
        try:
            preview = storage.read_json(reference)
        except StorageError as exc:
            raise HTTPException(status_code=404, detail='Preview đã hết hạn hoặc không tồn tại. Hãy upload lại file.') from exc
        try:
            created = datetime.fromisoformat(str(preview.get('created_at') or '').replace('Z', '+00:00'))
            if created.tzinfo is None:
                created = created.replace(tzinfo=timezone.utc)
        except (TypeError, ValueError) as exc:
            raise HTTPException(status_code=400, detail='Preview không hợp lệ. Hãy upload lại file.') from exc
        if datetime.now(timezone.utc) - created.astimezone(timezone.utc) > timedelta(hours=self.PREVIEW_TTL_HOURS):
            storage.delete(reference, missing_ok=True)
            raise HTTPException(status_code=410, detail='Preview đã hết hạn sau 2 giờ. Hãy upload lại file.')
        if str(preview.get('requested_by')) != str(requested_by):
            raise HTTPException(status_code=403, detail='Preview không thuộc người dùng hiện tại.')
        return reference, preview

    def _delivery_context(self, delivery: AcademicSubjectDelivery) -> tuple[AcademicSubject, AcademicTerm, AcademicBlock]:
        subject = self.db.get(AcademicSubject, delivery.subject_id)
        term = self.db.get(AcademicTerm, delivery.term_id)
        block = self.db.get(AcademicBlock, delivery.block_id)
        if not subject or not term or not block:
            raise HTTPException(status_code=409, detail='Dữ liệu môn/học kỳ/Block không còn đầy đủ.')
        return subject, term, block

    def serialize_plan(self, plan: UdemySubjectPlan, *, include_delivery: bool = False) -> dict[str, Any]:
        milestones = self.db.query(UdemySubjectPlanMilestone).filter(
            UdemySubjectPlanMilestone.plan_id == plan.id,
        ).order_by(UdemySubjectPlanMilestone.week_number.asc()).all()
        payload: dict[str, Any] = {
            'id': plan.id,
            'subject_delivery_id': plan.subject_delivery_id,
            'version': plan.version,
            'item_count': plan.item_count,
            'active': plan.active,
            'source': plan.source,
            'source_file_name': plan.source_file_name,
            'source_file_hash': plan.source_file_hash,
            'imported_by': plan.imported_by,
            'imported_at': plan.imported_at,
            'note': plan.note,
            'metadata_json': plan.metadata_json if isinstance(plan.metadata_json, dict) else {},
            'created_at': plan.created_at,
            'updated_at': plan.updated_at,
            'milestones': [
                {
                    'id': item.id,
                    'week_number': item.week_number,
                    'deadline_date': item.deadline_date,
                    'required_progress_percent': item.required_progress_percent,
                    'metadata_json': item.metadata_json if isinstance(item.metadata_json, dict) else {},
                }
                for item in milestones
            ],
        }
        if include_delivery:
            delivery = self.db.get(AcademicSubjectDelivery, plan.subject_delivery_id)
            if delivery:
                subject, term, block = self._delivery_context(delivery)
                payload['delivery'] = {
                    'id': delivery.id,
                    'subject_id': subject.id,
                    'subject_code': subject.subject_code,
                    'subject_name': subject.subject_name,
                    'term_id': term.id,
                    'term_name': term.term_name,
                    'block_id': block.id,
                    'block_name': block.block_name,
                    'branch': delivery.branch,
                    'learning_platform': delivery.learning_platform,
                }
        return payload

    def get_plan_detail(self, delivery_id: str) -> dict[str, Any]:
        delivery = self.db.get(AcademicSubjectDelivery, delivery_id)
        if not delivery or not delivery.active:
            raise HTTPException(status_code=404, detail='Không tìm thấy môn trong phạm vi học kỳ/Block.')
        subject, term, block = self._delivery_context(delivery)
        active_plan = self.db.query(UdemySubjectPlan).filter(
            UdemySubjectPlan.subject_delivery_id == delivery.id,
            UdemySubjectPlan.active.is_(True),
        ).order_by(UdemySubjectPlan.version.desc()).first()
        return {
            'delivery': {
                'id': delivery.id,
                'subject_id': subject.id,
                'subject_code': subject.subject_code,
                'subject_name': subject.subject_name,
                'term_id': term.id,
                'term_name': term.term_name,
                'block_id': block.id,
                'block_name': block.block_name,
                'branch': delivery.branch,
                'learning_platform': delivery.learning_platform,
            },
            'active_plan': self.serialize_plan(active_plan) if active_plan else None,
        }

    def list_plan_history(self, delivery_id: str) -> list[dict[str, Any]]:
        delivery = self.db.get(AcademicSubjectDelivery, delivery_id)
        if not delivery:
            raise HTTPException(status_code=404, detail='Không tìm thấy môn.')
        plans = self.db.query(UdemySubjectPlan).filter(
            UdemySubjectPlan.subject_delivery_id == delivery_id,
        ).order_by(UdemySubjectPlan.version.desc()).all()
        return [self.serialize_plan(item) for item in plans]

    @classmethod
    def validate_manual_payload(cls, *, item_count: int, milestones: list[dict[str, Any]]) -> list[ParsedMilestone]:
        if int(item_count or 0) <= 0:
            raise HTTPException(status_code=422, detail='Số lượng Item phải lớn hơn 0.')
        if not milestones:
            raise HTTPException(status_code=422, detail='Phải có ít nhất một mốc kế hoạch.')
        if len(milestones) > cls.MAX_WEEKS:
            raise HTTPException(status_code=422, detail=f'Mỗi kế hoạch tối đa {cls.MAX_WEEKS} mốc.')
        parsed: list[ParsedMilestone] = []
        seen_weeks: set[int] = set()
        for raw in milestones:
            week = int(raw.get('week_number') or 0)
            if week < 1 or week > cls.MAX_WEEKS or week in seen_weeks:
                raise HTTPException(status_code=422, detail='Số tuần phải duy nhất và nằm trong khoảng 1–52.')
            seen_weeks.add(week)
            deadline_raw = raw.get('deadline_date')
            try:
                deadline = deadline_raw if isinstance(deadline_raw, date) else date.fromisoformat(str(deadline_raw))
            except Exception as exc:
                raise HTTPException(status_code=422, detail=f'Week {week}: deadline không hợp lệ.') from exc
            try:
                progress = float(raw.get('required_progress_percent'))
            except Exception as exc:
                raise HTTPException(status_code=422, detail=f'Week {week}: tiến độ không hợp lệ.') from exc
            if not math.isfinite(progress) or progress < 0 or progress > 100:
                raise HTTPException(status_code=422, detail=f'Week {week}: tiến độ phải từ 0 đến 100.')
            parsed.append(ParsedMilestone(week, deadline, round(progress, 2)))
        parsed.sort(key=lambda item: item.week_number)
        for index, item in enumerate(parsed):
            if index == 0:
                continue
            previous = parsed[index - 1]
            if item.deadline_date <= previous.deadline_date:
                raise HTTPException(status_code=422, detail=f'Week {item.week_number}: deadline phải sau Week {previous.week_number}.')
            if item.required_progress_percent < previous.required_progress_percent:
                raise HTTPException(status_code=422, detail=f'Week {item.week_number}: tiến độ không được giảm.')
        return parsed

    def create_version(self, *, delivery_id: str, item_count: int, milestones: list[dict[str, Any]], actor: str | None, source: str, source_file_name: str | None = None, source_file_hash: str | None = None, note: str | None = None, metadata: dict[str, Any] | None = None, commit: bool = True) -> UdemySubjectPlan:
        delivery = self.db.query(AcademicSubjectDelivery).filter(AcademicSubjectDelivery.id == delivery_id).with_for_update().first()
        if not delivery or not delivery.active:
            raise HTTPException(status_code=404, detail='Không tìm thấy môn trong phạm vi học kỳ/Block.')
        if delivery.learning_platform != 'udemy':
            raise HTTPException(status_code=409, detail='Chỉ tạo kế hoạch cho môn đã chọn nền tảng Udemy.')
        parsed = self.validate_manual_payload(item_count=item_count, milestones=milestones)
        current_rows = self.db.query(UdemySubjectPlan).filter(UdemySubjectPlan.subject_delivery_id == delivery.id).order_by(UdemySubjectPlan.version.desc()).all()
        preview_token = str((metadata or {}).get('preview_token') or '').strip()
        if source == 'excel_import' and preview_token:
            existing = next((
                item for item in current_rows
                if isinstance(item.metadata_json, dict) and str(item.metadata_json.get('preview_token') or '') == preview_token
            ), None)
            if existing:
                return existing
        next_version = max([int(item.version or 0) for item in current_rows] or [0]) + 1
        now = datetime.utcnow()
        for item in current_rows:
            if item.active:
                item.active = False
                item.updated_at = now
                self.db.add(item)
        plan = UdemySubjectPlan(
            subject_delivery_id=delivery.id,
            version=next_version,
            item_count=int(item_count),
            active=True,
            source=source,
            source_file_name=source_file_name,
            source_file_hash=source_file_hash,
            imported_by=actor,
            imported_at=now,
            note=(note or '').strip() or None,
            metadata_json=json_safe_value(metadata or {}),
            created_at=now,
            updated_at=now,
        )
        self.db.add(plan)
        self.db.flush()
        for index, item in enumerate(parsed, 1):
            self.db.add(UdemySubjectPlanMilestone(
                plan_id=plan.id,
                week_number=item.week_number,
                deadline_date=item.deadline_date,
                required_progress_percent=item.required_progress_percent,
                sort_order=index,
                metadata_json={},
                created_at=now,
                updated_at=now,
            ))
        delivery.updated_at = now
        meta = dict(delivery.metadata_json or {}) if isinstance(delivery.metadata_json, dict) else {}
        meta['udemy_plan'] = {'active_plan_id': plan.id, 'version': next_version, 'updated_at': now.isoformat(), 'source': source}
        delivery.metadata_json = json_safe_value(meta)
        self.db.add(delivery)
        if commit:
            self.db.commit()
            self.db.refresh(plan)
        return plan

    def commit_preview(self, preview: dict[str, Any], *, actor: str | None) -> list[UdemySubjectPlan]:
        if not preview.get('can_commit') or int(preview.get('error_count') or 0) > 0:
            raise HTTPException(status_code=400, detail='Chỉ có thể import khi file không còn lỗi.')
        created: list[UdemySubjectPlan] = []
        try:
            valid_rows = sorted(
                (row for row in (preview.get('rows') or []) if not row.get('errors')),
                key=lambda row: str(row.get('delivery_id') or ''),
            )
            for row in valid_rows:
                plan = self.create_version(
                    delivery_id=str(row.get('delivery_id') or ''),
                    item_count=int(row.get('item_count') or 0),
                    milestones=list(row.get('milestones') or []),
                    actor=actor,
                    source='excel_import',
                    source_file_name=str(preview.get('filename') or '') or None,
                    source_file_hash=str(preview.get('file_sha256') or '') or None,
                    note='Import kế hoạch Udemy từ Excel',
                    metadata={'preview_schema': preview.get('schema_version'), 'preview_token': preview.get('preview_token'), 'source_row': row.get('row_no'), 'warnings': row.get('warnings') or []},
                    commit=False,
                )
                created.append(plan)
            self.db.commit()
            for item in created:
                self.db.refresh(item)
            return created
        except Exception:
            self.db.rollback()
            raise

    @classmethod
    def build_template(cls) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'KeHoachUdemy'
        sheet.merge_cells('A1:Z1')
        sheet['A1'] = 'Import Setup Môn Học Udemy - AI Server'
        sheet['A1'].font = Font(size=15, bold=True, color='FFFFFF')
        sheet['A1'].fill = PatternFill('solid', fgColor='17365D')
        sheet['A1'].alignment = Alignment(horizontal='center')
        headers = ['STT', 'Học kỳ', 'Block', 'Mã môn học', 'Tên môn học', 'Số lượng Item']
        for week in range(1, 11):
            headers.extend([f'Week {week}', f'Tiến độ week {week}'])
        for col, value in enumerate(headers, 1):
            cell = sheet.cell(row=2, column=col, value=value)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='1F4E78')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sample = [1, 'Fall 2026', 'Block 1', 'SOF3032', 'SOF3032 - Môn Udemy mẫu', 6]
        for week in range(1, 11):
            sample.extend([None, None])
        sample[6] = date(2026, 9, 7)
        sample[7] = 20
        sample[8] = date(2026, 9, 14)
        sample[9] = 40
        sample[10] = date(2026, 9, 21)
        sample[11] = 60
        sample[12] = date(2026, 9, 28)
        sample[13] = 80
        sample[14] = date(2026, 10, 5)
        sample[15] = 100
        for col, value in enumerate(sample, 1):
            sheet.cell(row=3, column=col, value=value)
        for col in range(7, 27, 2):
            sheet.cell(row=3, column=col).number_format = 'dd/mm/yyyy'
        widths = [8, 20, 15, 16, 38, 16] + [15, 18] * 10
        for index, width in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.freeze_panes = 'A3'
        sheet.auto_filter.ref = 'A2:Z3'
        guide = workbook.create_sheet('HuongDan')
        guide['A1'] = 'Hướng dẫn import kế hoạch Udemy'
        guide['A1'].font = Font(size=14, bold=True)
        notes = [
            'Mỗi dòng là một môn trong một Học kỳ + Block + Hệ.',
            'Môn phải đã được lấy từ AP và chọn nền tảng Udemy trên trang Quản lý môn học.',
            'Số lượng Item phải là số nguyên lớn hơn 0.',
            'Deadline nhập dạng ngày Excel, dd/mm/yyyy hoặc yyyy-mm-dd.',
            'Tiến độ từ 0 đến 100, không được giảm; deadline phải tăng dần.',
            'Import lại không ghi đè lịch sử: hệ thống tạo phiên bản kế hoạch mới và giữ phiên bản cũ.',
        ]
        for row, note in enumerate(notes, 3):
            guide.cell(row=row, column=1, value=f'- {note}')
        guide.column_dimensions['A'].width = 110
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    @classmethod
    def build_error_workbook(cls, preview: dict[str, Any]) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'LoiImport'
        headers = ['Dòng', 'Mã môn', 'Mã lỗi', 'Nội dung']
        for col, value in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=value)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='B91C1C')
        for row_no, item in enumerate(preview.get('errors') or [], 2):
            sheet.cell(row=row_no, column=1, value=item.get('row'))
            sheet.cell(row=row_no, column=2, value=item.get('subject_code'))
            sheet.cell(row=row_no, column=3, value=item.get('code'))
            sheet.cell(row=row_no, column=4, value=item.get('message'))
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 18
        sheet.column_dimensions['C'].width = 24
        sheet.column_dimensions['D'].width = 90
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()
