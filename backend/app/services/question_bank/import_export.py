from __future__ import annotations

import csv
import io
import uuid
import zipfile
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.models.question import Question, QuestionReviewLog
from app.models.question_bank import QuestionBankVersion, Subject, SubjectChapter
from app.services.generation_cache import question_fingerprint
from app.services.question_family import normalize_difficulty
from app.services.object_storage import StorageError, get_object_storage


IMPORT_HEADERS = [
    'question_text', 'option_a', 'option_b', 'option_c', 'option_d',
    'correct_answer', 'difficulty', 'concept_title', 'learning_objective',
    'explanation', 'source_ref',
]

HEADER_ALIASES = {
    'cau_hoi': 'question_text',
    'câu_hỏi': 'question_text',
    'noi_dung_cau_hoi': 'question_text',
    'nội_dung_câu_hỏi': 'question_text',
    'dap_an_a': 'option_a',
    'đáp_án_a': 'option_a',
    'dap_an_b': 'option_b',
    'đáp_án_b': 'option_b',
    'dap_an_c': 'option_c',
    'đáp_án_c': 'option_c',
    'dap_an_d': 'option_d',
    'đáp_án_d': 'option_d',
    'dap_an_dung': 'correct_answer',
    'đáp_án_đúng': 'correct_answer',
    'do_kho': 'difficulty',
    'độ_khó': 'difficulty',
    'concept': 'concept_title',
    'muc_tieu_hoc_tap': 'learning_objective',
    'mục_tiêu_học_tập': 'learning_objective',
    'giai_thich': 'explanation',
    'giải_thích': 'explanation',
    'nguon': 'source_ref',
    'nguồn': 'source_ref',
}


def _normalize_header(value: Any) -> str:
    text = str(value or '').strip().lower().replace(' ', '_').replace('-', '_')
    return HEADER_ALIASES.get(text, text)


def build_import_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'questions'
    ws.append(IMPORT_HEADERS)
    ws.freeze_panes = 'A2'
    widths = [42, 28, 28, 28, 28, 16, 14, 24, 32, 36, 28]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width
    guide = wb.create_sheet('instructions')
    guide.append(['Cột', 'Bắt buộc', 'Quy tắc'])
    guide.append(['question_text', 'Có', 'Nội dung câu hỏi trắc nghiệm một đáp án đúng.'])
    guide.append(['option_a..option_d', 'Có', 'Bốn phương án khác nhau, không để trống.'])
    guide.append(['correct_answer', 'Có', 'A, B, C hoặc D.'])
    guide.append(['difficulty', 'Có', 'easy, medium hoặc hard.'])
    guide.append(['concept_title', 'Không', 'Concept/nguyên tắc chính của câu hỏi.'])
    guide.append(['learning_objective', 'Không', 'Mục tiêu học tập.'])
    guide.append(['explanation', 'Không', 'Giải thích đáp án đúng.'])
    guide.append(['source_ref', 'Không', 'Nguồn tham chiếu.'])
    guide.append(['Lưu ý', '', 'Import luôn tạo câu ở trạng thái Chờ duyệt; không tự động publish hoặc đưa vào Release.'])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def validate_import_archive(raw: bytes, *, max_members: int = 500, max_uncompressed_bytes: int = 40 * 1024 * 1024) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            infos = archive.infolist()
            if len(infos) > max_members:
                raise ValueError(f'File Excel có quá nhiều entry nén ({len(infos)} > {max_members}).')
            total_uncompressed = sum(int(info.file_size or 0) for info in infos)
            if total_uncompressed > max_uncompressed_bytes:
                raise ValueError(f'File Excel giải nén quá lớn ({total_uncompressed} bytes > {max_uncompressed_bytes} bytes).')
            for info in infos:
                name = str(info.filename or '').replace('\\', '/')
                if name.startswith('/') or '../' in f'/{name}':
                    raise ValueError('File Excel chứa đường dẫn nén không an toàn.')
    except zipfile.BadZipFile as exc:
        raise ValueError('File không phải Excel .xlsx hợp lệ.') from exc


def parse_import_workbook(raw: bytes, *, max_rows: int = 2000) -> dict[str, Any]:
    if not raw:
        raise ValueError('File import trống')
    validate_import_archive(raw)
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb['questions'] if 'questions' in wb.sheetnames else wb[wb.sheetnames[0]]
    iterator = ws.iter_rows(values_only=True)
    header_row = next(iterator, None)
    if not header_row:
        raise ValueError('File không có dòng tiêu đề')
    headers = [_normalize_header(value) for value in header_row]
    missing = [name for name in ['question_text', 'option_a', 'option_b', 'option_c', 'option_d', 'correct_answer', 'difficulty'] if name not in headers]
    if missing:
        raise ValueError(f"Thiếu cột bắt buộc: {', '.join(missing)}")

    rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    duplicate_keys: set[str] = set()
    for excel_row_no, values in enumerate(iterator, start=2):
        if excel_row_no > max_rows + 1:
            errors.append({'row': excel_row_no, 'field': '*', 'code': 'MAX_ROWS_EXCEEDED', 'message': f'File vượt giới hạn {max_rows} dòng.'})
            break
        mapped = {headers[index]: values[index] if index < len(values) else None for index in range(len(headers)) if headers[index]}
        if not any(str(value or '').strip() for value in mapped.values()):
            continue
        normalized = {
            'row_no': excel_row_no,
            'question_text': str(mapped.get('question_text') or '').strip(),
            'option_a': str(mapped.get('option_a') or '').strip(),
            'option_b': str(mapped.get('option_b') or '').strip(),
            'option_c': str(mapped.get('option_c') or '').strip(),
            'option_d': str(mapped.get('option_d') or '').strip(),
            'correct_answer': str(mapped.get('correct_answer') or '').strip().upper(),
            'difficulty_raw': str(mapped.get('difficulty') or '').strip().lower(),
            'difficulty': normalize_difficulty(str(mapped.get('difficulty') or '').strip()),
            'concept_title': str(mapped.get('concept_title') or '').strip(),
            'learning_objective': str(mapped.get('learning_objective') or '').strip(),
            'explanation': str(mapped.get('explanation') or '').strip(),
            'source_ref': str(mapped.get('source_ref') or '').strip(),
        }
        row_errors: list[dict[str, Any]] = []
        for field in ['question_text', 'option_a', 'option_b', 'option_c', 'option_d']:
            if not normalized[field]:
                row_errors.append({'row': excel_row_no, 'field': field, 'code': 'REQUIRED', 'message': 'Không được để trống.'})
        if normalized['correct_answer'] not in {'A', 'B', 'C', 'D'}:
            row_errors.append({'row': excel_row_no, 'field': 'correct_answer', 'code': 'INVALID_ANSWER', 'message': 'Chỉ chấp nhận A, B, C hoặc D.'})
        if normalized['difficulty_raw'] not in {'easy', 'medium', 'hard'}:
            row_errors.append({'row': excel_row_no, 'field': 'difficulty', 'code': 'INVALID_DIFFICULTY', 'message': 'Chỉ chấp nhận easy, medium hoặc hard.'})
        options = [normalized['option_a'], normalized['option_b'], normalized['option_c'], normalized['option_d']]
        if len({value.casefold() for value in options if value}) != len([value for value in options if value]):
            row_errors.append({'row': excel_row_no, 'field': 'options', 'code': 'DUPLICATE_OPTIONS', 'message': 'Các phương án phải khác nhau.'})
        key = normalized['question_text'].casefold()
        if key and key in duplicate_keys:
            row_errors.append({'row': excel_row_no, 'field': 'question_text', 'code': 'DUPLICATE_IN_FILE', 'message': 'Câu hỏi bị lặp trong file.'})
        duplicate_keys.add(key)
        if row_errors:
            errors.extend(row_errors)
        normalized.pop('difficulty_raw', None)
        rows.append(normalized)

    valid_rows = [row for row in rows if not any(error['row'] == row['row_no'] for error in errors)]
    return {
        'rows': rows,
        'valid_rows': valid_rows,
        'errors': errors,
        'total_rows': len(rows),
        'valid_count': len(valid_rows),
        'error_count': len({error['row'] for error in errors}),
    }


def persist_preview(payload: dict[str, Any]) -> tuple[str, str]:
    token = uuid.uuid4().hex
    payload = {**payload, 'preview_created_at': datetime.now(timezone.utc).isoformat()}
    key = f'question-bank/_pending-operation-files/question-import-{token}.json'
    reference = get_object_storage().put_json(key, payload)
    return token, reference


def load_preview(reference: str) -> dict[str, Any]:
    try:
        return get_object_storage().read_json(reference)
    except StorageError as exc:
        raise ValueError('Không tìm thấy dữ liệu preview import. Hãy upload lại file.') from exc


def build_import_error_workbook(payload: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'import-errors'
    ws.append(['row', 'field', 'code', 'message'])
    for item in payload.get('errors') or []:
        ws.append([item.get('row'), item.get('field'), item.get('code'), item.get('message')])
    ws.freeze_panes = 'A2'
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 70
    source = wb.create_sheet('source-preview')
    source.append(IMPORT_HEADERS)
    for row in payload.get('rows') or []:
        source.append([row.get(name) for name in IMPORT_HEADERS])
    source.freeze_panes = 'A2'
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def import_questions(db: Session, *, bank_version_id: str, preview_path: str, actor: str | None = None) -> dict[str, Any]:
    version = db.get(QuestionBankVersion, bank_version_id)
    if not version:
        raise ValueError('Không tìm thấy Bank Version')
    chapter = db.get(SubjectChapter, version.chapter_id)
    subject = db.get(Subject, version.subject_id)
    if not chapter or not subject:
        raise ValueError('Bank Version thiếu liên kết môn/bài')
    payload = load_preview(preview_path)
    rows = list(payload.get('valid_rows') or [])
    if not rows:
        raise ValueError('Không có dòng hợp lệ để import')

    created_ids: list[str] = []
    skipped: list[dict[str, Any]] = []
    for row in rows:
        q_hash = question_fingerprint(
            row['question_text'],
            course_id=f'bank:{version.id}',
            source_node_id=f'manual-import:{version.id}',
            difficulty=row['difficulty'],
        )
        existing = db.query(Question.id).filter(Question.bank_version_id == version.id, Question.question_hash == q_hash).first()
        if existing:
            skipped.append({'row': row.get('row_no'), 'reason': 'duplicate_question'})
            continue
        question = Question(
            course_id=f'bank:{version.id}',
            department_id=subject.department_id,
            subject_id=version.subject_id,
            subject_chapter_id=version.chapter_id,
            bank_version_id=version.id,
            lesson_title=chapter.title,
            block_id=f'bank-version:{version.id}:manual-import',
            topic=chapter.title,
            concept_title=row.get('concept_title') or None,
            source_evidence=row.get('source_ref') or '',
            difficulty=row['difficulty'],
            cognitive_level='remember',
            learning_objective=row.get('learning_objective') or '',
            pedagogy_json={},
            question_type='single_choice',
            question_text=row['question_text'],
            question_hash=q_hash,
            option_a=row['option_a'],
            option_b=row['option_b'],
            option_c=row['option_c'],
            option_d=row['option_d'],
            correct_answer=row['correct_answer'],
            explanation=row.get('explanation') or '',
            source_ref=row.get('source_ref') or 'question-import',
            source_type='manual_import',
            source_node_id=f'bank-version:{version.id}:manual-import',
            source_node_title=chapter.title,
            chapter_node_id=version.chapter_id,
            chapter_title=chapter.title,
            source_excerpt=row.get('source_ref') or '',
            tags=['manual_import'],
            ai_rationale='',
            quality_score=0.0,
            quality_flags=['manual_import_requires_review'],
            model_provider='manual',
            model_name='manual_import',
            status='pending_review',
            authoring_mode='import',
            created_by=actor,
        )
        db.add(question)
        db.flush()
        db.add(QuestionReviewLog(
            id=str(uuid.uuid4()),
            question_id=question.id,
            old_status='imported',
            new_status='pending_review',
            actor=actor or 'system',
            note=f'Import câu hỏi từ Excel, dòng {row.get("row_no")}.',
        ))
        created_ids.append(question.id)
    db.commit()
    try:
        get_object_storage().delete(preview_path, missing_ok=True)
    except Exception:
        pass
    return {
        'ok': True,
        'bank_version_id': bank_version_id,
        'created_count': len(created_ids),
        'skipped_count': len(skipped),
        'created_question_ids': created_ids,
        'skipped': skipped,
        'message': f'Đã import {len(created_ids)} câu ở trạng thái Chờ duyệt.' + (f' Bỏ qua {len(skipped)} câu trùng.' if skipped else ''),
    }


def export_questions_csv(rows: list[Question]) -> bytes:
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(IMPORT_HEADERS + ['status', 'quality_score', 'created_at'])
    for row in rows:
        writer.writerow([
            row.question_text, row.option_a, row.option_b, row.option_c, row.option_d,
            row.correct_answer, row.difficulty, row.concept_title or '', row.learning_objective or '',
            row.explanation or '', row.source_ref or '', row.status, row.quality_score,
            row.created_at.isoformat() if row.created_at else '',
        ])
    return ('\ufeff' + out.getvalue()).encode('utf-8')


def _openedx_parsed_to_canonical(parsed: Any) -> dict[str, Any]:
    from app.services.question_content import normalize_question_content
    qtype = str(parsed.question_type or '').strip()
    if qtype in {'single_select', 'multi_select'}:
        options = [{'id': f'import-opt-{index + 1}', 'text': choice.text, 'correct': bool(choice.correct), 'feedback': ''} for index, choice in enumerate(parsed.choices)]
        return normalize_question_content(qtype, {'response': {'type': qtype, 'options': options}})
    if qtype == 'text_input':
        if parsed.string_response_type not in {'ci', 'cs'}:
            raise ValueError(f'Open edX stringresponse type={parsed.string_response_type} chưa được hỗ trợ; không import để tránh thay đổi cách chấm.')
        return normalize_question_content(qtype, {'response': {'type': qtype, 'accepted_answers': [{'text': answer, 'case_sensitive': bool(parsed.case_sensitive)} for answer in parsed.accepted_answers]}})
    if qtype == 'numerical_input':
        return normalize_question_content(qtype, {'response': {'type': qtype, 'answer': parsed.numerical_answer, 'tolerance': parsed.numerical_tolerance, 'tolerance_type': parsed.numerical_tolerance_type}})
    raise ValueError(f'Open edX response type {qtype or "unknown"} chưa được hỗ trợ.')


def preview_openedx_problem_import(problem_olx: str, *, max_questions: int = 200) -> dict[str, Any]:
    from app.services.problem_parser import parse_problem_xml
    raw = str(problem_olx or '')
    if not raw.strip(): raise ValueError('OLX import trống.')
    if len(raw.encode('utf-8')) > 2 * 1024 * 1024: raise ValueError('OLX import vượt giới hạn 2 MB.')
    parsed_rows = parse_problem_xml(raw)
    if len(parsed_rows) > max_questions: raise ValueError(f'OLX có quá nhiều câu hỏi ({len(parsed_rows)} > {max_questions}).')
    rows=[]; errors=[]
    for index, parsed in enumerate(parsed_rows,start=1):
        try:
            canonical=_openedx_parsed_to_canonical(parsed)
            rows.append({'index':index,'question_text':parsed.question,'question_type':parsed.question_type,'content':canonical,'solution':parsed.solution,'warnings':list(parsed.warnings or [])})
        except ValueError as exc:
            errors.append({'index':index,'question_text':str(parsed.question or '')[:240],'question_type':parsed.question_type,'error':str(exc)[:500]})
    warnings=[]
    if '<img' in raw.lower(): warnings.append('OLX có ảnh. Bản import giữ response type nhưng không tự sao chép asset Open edX; hãy upload lại ảnh thủ công trong ACMS.')
    if not parsed_rows: errors.append({'index':0,'error':'Không tìm thấy response type Open edX được hỗ trợ trong OLX.'})
    return {'ok':not errors,'total_parsed':len(parsed_rows),'valid_count':len(rows),'invalid_count':len(errors),'questions':rows,'errors':errors,'warnings':warnings,'message':f'Đọc được {len(rows)}/{len(parsed_rows)} câu Open edX hợp lệ.'}


def import_openedx_problem_questions(db: Session, *, bank_version_id: str, problem_olx: str, source_ref: str = 'openedx-olx-import', actor: str | None = None) -> dict[str, Any]:
    from app.services.question_content import apply_canonical_content, question_response_fingerprint
    preview=preview_openedx_problem_import(problem_olx)
    if preview['errors']: raise ValueError('OLX còn response không hợp lệ/không hỗ trợ; hãy sửa trước khi import.')
    version=db.get(QuestionBankVersion,bank_version_id)
    if not version: raise ValueError('Không tìm thấy Bank Version.')
    subject=db.get(Subject,version.subject_id); chapter=db.get(SubjectChapter,version.chapter_id)
    if not subject or not chapter: raise ValueError('Bank Version thiếu Subject hoặc Chapter.')
    created_ids=[]; skipped=[]
    for row in preview['questions']:
        qtype=str(row['question_type']); text=str(row['question_text'] or '').strip()
        q_hash=question_fingerprint(text,course_id=f'bank:{version.id}',source_node_id=f'openedx-import:{source_ref}:{qtype}',difficulty='medium')
        if db.query(Question.id).filter(Question.bank_version_id==version.id,Question.question_hash==q_hash).first():
            skipped.append({'index':row['index'],'reason':'duplicate_question'}); continue
        try:
            with db.begin_nested():
                question=Question(course_id=f'bank:{version.id}',department_id=subject.department_id,subject_id=version.subject_id,subject_chapter_id=version.chapter_id,bank_version_id=version.id,lesson_title=chapter.title,block_id=f'bank-version:{version.id}:openedx-import',topic=chapter.title,source_evidence=source_ref,difficulty='medium',cognitive_level='remember',learning_objective='',pedagogy_json={},question_type=qtype,question_text=text,question_hash=q_hash,option_a='',option_b='',option_c='',option_d='',correct_answer='A',explanation=str(row.get('solution') or ''),source_ref=source_ref or 'openedx-olx-import',source_type='openedx_import',source_node_id=f'bank-version:{version.id}:openedx-import',source_node_title=chapter.title,chapter_node_id=version.chapter_id,chapter_title=chapter.title,source_excerpt=source_ref,tags=['openedx_import',qtype],ai_rationale='',quality_score=0.0,quality_flags=['openedx_import_requires_review'],model_provider='manual',model_name='openedx_import',status='pending_review',authoring_mode='import',created_by=actor)
                apply_canonical_content(question,qtype,row['content']); question.question_hash=question_response_fingerprint(question)
                db.add(question); db.flush()
                db.add(QuestionReviewLog(id=str(uuid.uuid4()),question_id=question.id,old_status='imported',new_status='pending_review',actor=actor or 'system',note=f'Import Open edX OLX response type {qtype}.')); db.flush(); created_ids.append(question.id)
        except Exception as exc:
            skipped.append({'index':row['index'],'reason':'database_row_error','error_type':type(exc).__name__})
    db.commit()
    return {'ok':True,'bank_version_id':bank_version_id,'created_count':len(created_ids),'skipped_count':len(skipped),'created_question_ids':created_ids,'skipped':skipped,'warnings':preview.get('warnings') or [],'message':f'Đã import {len(created_ids)} câu Open edX ở trạng thái Chờ duyệt.'+(f' Bỏ qua {len(skipped)} câu.' if skipped else '')}
