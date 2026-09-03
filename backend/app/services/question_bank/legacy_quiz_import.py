from __future__ import annotations

import hashlib
import io
import json
import re
import unicodedata
import uuid
import zipfile
from collections import Counter, defaultdict
from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path, PurePosixPath
from typing import Any, Callable, Iterable

from openpyxl import load_workbook
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from app.models.question import Question, QuestionMedia, QuestionReviewLog
from app.models.question_bank import (
    LearningMaterialVersion,
    QuestionBankVersion,
    Subject,
    SubjectChapter,
    SubjectOffering,
)
from app.services.bank_dashboard_stats import BankDashboardStatsService
from app.services.generation_cache import question_fingerprint
from app.services.object_storage import ObjectStorage, StorageError, get_object_storage
from app.services.question_bank.helpers import safe_upload_filename
from app.services.question_content import apply_canonical_content, normalize_question_content
from app.services.question_media import validate_question_image


LEGACY_IMPORT_TERM = 'SU26'
LEGACY_PREVIEW_TTL_HOURS = 2
MAX_WORKBOOKS = 20
MAX_QUESTIONS = 10_000
MAX_ROWS_PER_SHEET = 8_000
MAX_ASSETS = 500
MAX_TOTAL_ASSET_BYTES = 100 * 1024 * 1024
MAX_ARCHIVE_MEMBERS = 1_500
MAX_ARCHIVE_UNCOMPRESSED_BYTES = 150 * 1024 * 1024

IMAGE_SUFFIXES = {'.png', '.jpg', '.jpeg', '.webp'}
ASSET_ARCHIVE_SUFFIXES = {'.zip'}
BLANK_TOKEN_RE = re.compile(r'\[_{3,}\]')
IMAGE_MARKER_RE = re.compile(r'\[\s*([^\[\]]+?\.(?:png|jpe?g|webp))\s*\]', re.IGNORECASE)

HEADER_ALIASES = {
    'NO': 'number',
    'STT': 'number',
    'SO_THU_TU': 'number',
    'QUESTION': 'question',
    'CAU_HOI': 'question',
    'ABC': 'option_label',
    'LABEL': 'option_label',
    'NHAN_DAP_AN': 'option_label',
    'ANSWER': 'option_text',
    'DAP_AN': 'option_text',
    'CORRECT': 'correct',
    'DAP_AN_DUNG': 'correct',
    'TYPE': 'question_type',
    'LOAI': 'question_type',
    'LOAI_CAU_HOI': 'question_type',
    'NGUONG': 'threshold',
    'DIFFICULTY': 'difficulty',
    'DO_KHO': 'difficulty',
    'MIX_CHOICE': 'mix_choice',
    'BAI': 'lesson',
    'LESSON': 'lesson',
}
REQUIRED_HEADERS = {'number', 'question', 'option_label', 'option_text', 'correct'}


def _utcnow_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _sha256(raw: bytes) -> str:
    return hashlib.sha256(raw).hexdigest()


def _fold(value: object) -> str:
    text = unicodedata.normalize('NFKD', str(value or '').strip())
    text = ''.join(char for char in text if not unicodedata.combining(char))
    text = text.upper().replace('Đ', 'D')
    return re.sub(r'[^A-Z0-9]+', '_', text).strip('_')


def _normalized_asset_name(value: object) -> str:
    basename = str(value or '').replace('\\', '/').rsplit('/', 1)[-1].strip()
    return unicodedata.normalize('NFC', basename).casefold()


def _normalized_question_key(value: object) -> str:
    return ' '.join(str(value or '').casefold().split())


def _safe_slug(value: object, *, fallback: str = 'item', max_len: int = 80) -> str:
    folded = _fold(value).lower()
    slug = re.sub(r'[^a-z0-9_.-]+', '-', folded).strip('.-')
    return (slug or fallback)[:max_len]


def _error(
    code: str,
    message: str,
    *,
    workbook: str = '',
    sheet: str = '',
    row: int | None = None,
    field: str = '*',
) -> dict[str, Any]:
    return {
        'code': code,
        'message': message,
        'workbook': workbook,
        'sheet': sheet,
        'row': row,
        'field': field,
    }


def _validate_zip_archive(
    raw: bytes,
    *,
    label: str,
    max_members: int = MAX_ARCHIVE_MEMBERS,
    max_uncompressed_bytes: int = MAX_ARCHIVE_UNCOMPRESSED_BYTES,
) -> list[zipfile.ZipInfo]:
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            infos = archive.infolist()
            if len(infos) > max_members:
                raise ValueError(f'{label} có quá nhiều entry nén ({len(infos)} > {max_members}).')
            total = sum(int(info.file_size or 0) for info in infos)
            if total > max_uncompressed_bytes:
                raise ValueError(
                    f'{label} giải nén quá lớn ({total} bytes > {max_uncompressed_bytes} bytes).'
                )
            for info in infos:
                raw_name = str(info.filename or '').replace('\\', '/')
                path = PurePosixPath(raw_name)
                if raw_name.startswith('/') or path.is_absolute() or '..' in path.parts:
                    raise ValueError(f'{label} chứa đường dẫn nén không an toàn.')
            return infos
    except zipfile.BadZipFile as exc:
        raise ValueError(f'{label} không phải file nén hợp lệ.') from exc


def _find_header(rows: list[tuple[Any, ...]]) -> tuple[int, dict[int, str]]:
    best: tuple[int, dict[int, str]] | None = None
    for index, row in enumerate(rows[:12]):
        mapped: dict[int, str] = {}
        for column, value in enumerate(row):
            canonical = HEADER_ALIASES.get(_fold(value))
            if canonical and canonical not in mapped.values():
                mapped[column] = canonical
        if REQUIRED_HEADERS.issubset(set(mapped.values())):
            return index, mapped
        if best is None or len(mapped) > len(best[1]):
            best = (index, mapped)
    found = sorted(set((best or (0, {}))[1].values()))
    missing = sorted(REQUIRED_HEADERS - set(found))
    raise ValueError(f"Thiếu cột bắt buộc: {', '.join(missing)}.")


def _mapped_row(values: tuple[Any, ...], columns: dict[int, str]) -> dict[str, Any]:
    return {
        field: values[index] if index < len(values) else None
        for index, field in columns.items()
    }


def _correct_labels(value: object) -> list[str]:
    raw = _fold(value).replace('_', '')
    return [char for char in raw if 'A' <= char <= 'L']


def _question_type(raw_type: object, correct_labels: list[str], question_text: str) -> str:
    raw = str(raw_type or '').strip()
    if raw:
        try:
            numeric = int(float(raw.replace(',', '.')))
        except ValueError as exc:
            raise ValueError('TYPE chỉ chấp nhận 0, 1 hoặc 2.') from exc
        mapping = {0: 'single_select', 1: 'multi_select', 2: 'dropdown_fill'}
        if numeric not in mapping:
            raise ValueError('TYPE chỉ chấp nhận 0, 1 hoặc 2.')
        return mapping[numeric]
    if BLANK_TOKEN_RE.search(question_text):
        return 'dropdown_fill'
    return 'single_select' if len(correct_labels) == 1 else 'multi_select'


def _difficulty(raw_threshold: object, raw_difficulty: object) -> str:
    threshold = str(raw_threshold or '').strip()
    if threshold:
        try:
            numeric = int(float(threshold.replace(',', '.')))
        except ValueError as exc:
            raise ValueError('NGƯỠNG chỉ chấp nhận 1 (Dễ), 2 (Trung bình), 3 (Khó).') from exc
        mapping = {1: 'easy', 2: 'medium', 3: 'hard'}
        if numeric not in mapping:
            raise ValueError('NGƯỠNG chỉ chấp nhận 1 (Dễ), 2 (Trung bình), 3 (Khó).')
        return mapping[numeric]

    explicit = _fold(raw_difficulty)
    if explicit:
        mapping = {
            '1': 'easy',
            'EASY': 'easy',
            'DE': 'easy',
            '2': 'medium',
            'MEDIUM': 'medium',
            'TRUNG_BINH': 'medium',
            '3': 'hard',
            'HARD': 'hard',
            'KHO': 'hard',
        }
        if explicit not in mapping:
            raise ValueError('Độ khó chỉ chấp nhận easy/medium/hard hoặc Dễ/Trung bình/Khó.')
        return mapping[explicit]
    # Keep a non-null compatibility value for the legacy Question column. The
    # separate difficulty_classified flag remains authoritative for Quiz planning.
    return 'medium'


def _difficulty_is_classified(raw_threshold: object, raw_difficulty: object) -> bool:
    return bool(str(raw_threshold or '').strip() or str(raw_difficulty or '').strip())


def _difficulty_count_key(question: dict[str, Any]) -> str:
    classified = question.get('difficulty_classified')
    if classified is None:
        classified = _difficulty_is_classified(
            question.get('threshold_raw'), question.get('difficulty_raw')
        )
    if not bool(classified):
        return 'unclassified'
    return str(question.get('difficulty') or '')


def _chapter_number(sheet_name: str, sheet_index: int) -> tuple[int, str | None]:
    match = re.search(r'(\d+)\s*$', str(sheet_name or '').strip())
    if match:
        return int(match.group(1)), None
    fallback = sheet_index + 1
    return fallback, f'Sheet {sheet_name!r} không có số bài ở cuối tên; dùng bài {fallback}.'


def _embedded_images(raw: bytes, grouped_ranges: dict[str, list[tuple[int, int, dict[str, Any]]]]) -> list[str]:
    warnings: list[str] = []
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            if not any(str(name).startswith('xl/media/') for name in archive.namelist()):
                return warnings
    except zipfile.BadZipFile:
        return warnings

    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=False, data_only=True)
    except Exception as exc:
        return [f'Không đọc được ảnh nhúng trong workbook: {exc.__class__.__name__}.']
    for worksheet in workbook.worksheets:
        ranges = grouped_ranges.get(worksheet.title, [])
        for image_index, image in enumerate(getattr(worksheet, '_images', []) or [], start=1):
            try:
                anchor_row = int(image.anchor._from.row) + 1
                image_raw = bytes(image._data())
                validated = validate_question_image(image_raw)
            except Exception as exc:
                warnings.append(
                    f'Sheet {worksheet.title}: bỏ qua ảnh nhúng không đọc được ({exc.__class__.__name__}).'
                )
                continue
            question = next(
                (item for start, end, item in ranges if start <= anchor_row <= end),
                None,
            )
            if question is None:
                warnings.append(
                    f'Sheet {worksheet.title}: ảnh nhúng tại dòng {anchor_row} không gắn với câu hỏi nào.'
                )
                continue
            filename = f'{_safe_slug(worksheet.title)}-row-{anchor_row}-image-{image_index}{validated.extension}'
            question.setdefault('embedded_media', []).append({
                'filename': filename,
                'raw': image_raw,
                'mime_type': validated.mime_type,
                'sha256': validated.sha256,
                'width': validated.width,
                'height': validated.height,
                'anchor_row': anchor_row,
            })
    return warnings


def _parse_sheet(
    worksheet: Any,
    *,
    workbook_name: str,
    sheet_index: int,
) -> tuple[dict[str, Any], list[tuple[int, int, dict[str, Any]]]]:
    errors: list[dict[str, Any]] = []
    warnings: list[str] = []
    rows = list(worksheet.iter_rows(values_only=True))
    if len(rows) > MAX_ROWS_PER_SHEET:
        errors.append(_error(
            'MAX_ROWS_PER_SHEET',
            f'Sheet vượt giới hạn {MAX_ROWS_PER_SHEET} dòng.',
            workbook=workbook_name,
            sheet=worksheet.title,
            row=MAX_ROWS_PER_SHEET + 1,
        ))
        rows = rows[:MAX_ROWS_PER_SHEET]
    try:
        header_index, columns = _find_header(rows)
    except ValueError as exc:
        return ({
            'sheet_name': worksheet.title,
            'chapter_no': sheet_index + 1,
            'chapter_title': worksheet.title,
            'questions': [],
            'errors': [_error(
                'MISSING_HEADERS',
                str(exc),
                workbook=workbook_name,
                sheet=worksheet.title,
            )],
            'warnings': [],
            'question_count': 0,
            'type_counts': {},
            'difficulty_counts': {},
        }, [])

    chapter_no, chapter_warning = _chapter_number(worksheet.title, sheet_index)
    if chapter_warning:
        warnings.append(chapter_warning)

    groups: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    for excel_row, values in enumerate(rows[header_index + 1 :], start=header_index + 2):
        mapped = _mapped_row(values, columns)
        has_value = any(str(value or '').strip() for value in mapped.values())
        if not has_value:
            continue
        number = str(mapped.get('number') or '').strip()
        if number:
            if current:
                current['end_row'] = excel_row - 1
                groups.append(current)
            current = {
                'source_row': excel_row,
                'end_row': excel_row,
                'question_no': number,
                'question_text_raw': mapped.get('question'),
                'question_type_raw': mapped.get('question_type'),
                'threshold_raw': mapped.get('threshold'),
                'difficulty_raw': mapped.get('difficulty'),
                'lesson_raw': mapped.get('lesson'),
                'options_raw': [],
                'correct_raw': [],
            }
        elif current is None:
            warnings.append(f'Sheet {worksheet.title}, dòng {excel_row}: bỏ qua dòng không có NO.')
            continue
        if current is None:
            continue
        for field in ('question_text_raw', 'question_type_raw', 'threshold_raw', 'difficulty_raw', 'lesson_raw'):
            source_field = {
                'question_text_raw': 'question',
                'question_type_raw': 'question_type',
                'threshold_raw': 'threshold',
                'difficulty_raw': 'difficulty',
                'lesson_raw': 'lesson',
            }[field]
            if not str(current.get(field) or '').strip() and str(mapped.get(source_field) or '').strip():
                current[field] = mapped.get(source_field)
        label = _fold(mapped.get('option_label')).replace('_', '')
        option_text = str(mapped.get('option_text') or '').strip()
        if label or option_text:
            current['options_raw'].append({'label': label, 'text': option_text, 'row': excel_row})
        if str(mapped.get('correct') or '').strip():
            current['correct_raw'].append(str(mapped.get('correct')).strip())
        current['end_row'] = excel_row
    if current:
        groups.append(current)

    questions: list[dict[str, Any]] = []
    ranges: list[tuple[int, int, dict[str, Any]]] = []
    for group in groups:
        row = int(group['source_row'])
        raw_prompt = str(group.get('question_text_raw') or '').strip()
        image_refs = [match.group(1).strip() for match in IMAGE_MARKER_RE.finditer(raw_prompt)]
        prompt = IMAGE_MARKER_RE.sub(' ', raw_prompt)
        prompt = re.sub(r'[ \t]+', ' ', prompt)
        prompt = re.sub(r' *\n *', '\n', prompt).strip()
        labels = _correct_labels(' '.join(group.get('correct_raw') or []))
        question_errors: list[dict[str, Any]] = []
        if not prompt:
            question_errors.append(_error(
                'QUESTION_REQUIRED',
                'Nội dung câu hỏi không được để trống.',
                workbook=workbook_name,
                sheet=worksheet.title,
                row=row,
                field='QUESTION',
            ))
        try:
            qtype = _question_type(group.get('question_type_raw'), labels, prompt)
        except ValueError as exc:
            qtype = 'single_select'
            question_errors.append(_error(
                'INVALID_TYPE', str(exc), workbook=workbook_name, sheet=worksheet.title, row=row, field='TYPE'
            ))
        difficulty_classified = _difficulty_is_classified(
            group.get('threshold_raw'), group.get('difficulty_raw')
        )
        try:
            difficulty = _difficulty(group.get('threshold_raw'), group.get('difficulty_raw'))
        except ValueError as exc:
            difficulty = 'medium'
            question_errors.append(_error(
                'INVALID_DIFFICULTY',
                str(exc),
                workbook=workbook_name,
                sheet=worksheet.title,
                row=row,
                field='NGƯỠNG',
            ))

        options: list[dict[str, Any]] = []
        seen_labels: set[str] = set()
        for option_index, raw_option in enumerate(group.get('options_raw') or []):
            label = str(raw_option.get('label') or '').strip().upper()
            text = str(raw_option.get('text') or '').strip()
            option_row = int(raw_option.get('row') or row)
            if not re.fullmatch(r'[A-L]', label):
                question_errors.append(_error(
                    'INVALID_OPTION_LABEL',
                    'Nhãn đáp án phải là một chữ cái từ A đến L.',
                    workbook=workbook_name,
                    sheet=worksheet.title,
                    row=option_row,
                    field='ABC',
                ))
                continue
            if label in seen_labels:
                question_errors.append(_error(
                    'DUPLICATE_OPTION_LABEL',
                    f'Nhãn đáp án {label} bị lặp.',
                    workbook=workbook_name,
                    sheet=worksheet.title,
                    row=option_row,
                    field='ABC',
                ))
                continue
            if not text:
                question_errors.append(_error(
                    'OPTION_REQUIRED',
                    f'Đáp án {label} không được để trống.',
                    workbook=workbook_name,
                    sheet=worksheet.title,
                    row=option_row,
                    field='ANSWER',
                ))
            seen_labels.add(label)
            options.append({
                'id': f'opt-{label.lower()}',
                'label': label,
                'text': text,
                'correct': label in labels,
                'feedback': '',
            })
        unknown_labels = [label for label in labels if label not in seen_labels]
        if unknown_labels:
            question_errors.append(_error(
                'INVALID_CORRECT_KEY',
                f"Đáp án đúng tham chiếu nhãn không tồn tại: {', '.join(unknown_labels)}.",
                workbook=workbook_name,
                sheet=worksheet.title,
                row=row,
                field='CORRECT',
            ))

        response: dict[str, Any] = {'type': qtype, 'options': options}
        if qtype == 'dropdown_fill':
            response['correct_option_ids'] = [f'opt-{label.lower()}' for label in labels]
            blank_count = len(BLANK_TOKEN_RE.findall(prompt))
            if blank_count != len(labels):
                question_errors.append(_error(
                    'BLANK_ANSWER_COUNT_MISMATCH',
                    f'Số ô trống ({blank_count}) phải bằng số đáp án đúng theo thứ tự ({len(labels)}).',
                    workbook=workbook_name,
                    sheet=worksheet.title,
                    row=row,
                    field='QUESTION/CORRECT',
                ))
        try:
            content = normalize_question_content(qtype, {'response': response})
        except ValueError as exc:
            content = {'schema_version': 2, 'response': response}
            question_errors.append(_error(
                'INVALID_QUESTION',
                str(exc),
                workbook=workbook_name,
                sheet=worksheet.title,
                row=row,
            ))

        question = {
            'source_row': row,
            'end_row': int(group['end_row']),
            'question_no': str(group.get('question_no') or ''),
            'question_text_raw': raw_prompt,
            'question_text': prompt,
            'question_type': qtype,
            'question_type_raw': str(group.get('question_type_raw') or ''),
            'threshold_raw': str(group.get('threshold_raw') or ''),
            'difficulty_raw': str(group.get('difficulty_raw') or ''),
            'difficulty': difficulty,
            'difficulty_classified': difficulty_classified,
            'correct_key': ''.join(labels),
            'content': content,
            'image_refs': image_refs,
            'embedded_media': [],
            'duplicate_in_source': False,
            'errors': question_errors,
        }
        questions.append(question)
        ranges.append((row, int(group['end_row']), question))
        errors.extend(question_errors)

    type_counts = Counter(item['question_type'] for item in questions)
    difficulty_counts = Counter(_difficulty_count_key(item) for item in questions)
    return ({
        'sheet_name': worksheet.title,
        'chapter_no': chapter_no,
        'chapter_title': worksheet.title,
        'questions': questions,
        'errors': errors,
        'warnings': warnings,
        'question_count': len(questions),
        'type_counts': dict(type_counts),
        'difficulty_counts': dict(difficulty_counts),
    }, ranges)


def parse_legacy_quiz_workbook(raw: bytes, *, filename: str) -> dict[str, Any]:
    data = bytes(raw or b'')
    safe_name = safe_upload_filename(filename)
    if not data:
        raise ValueError('File Excel trống.')
    if Path(safe_name).suffix.lower() != '.xlsx':
        raise ValueError('Chỉ hỗ trợ file Excel .xlsx.')
    _validate_zip_archive(data, label='File Excel')
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError('Không đọc được file Excel .xlsx.') from exc
    if not workbook.sheetnames:
        raise ValueError('File Excel không có sheet.')

    sheets: list[dict[str, Any]] = []
    ranges: dict[str, list[tuple[int, int, dict[str, Any]]]] = {}
    chapter_numbers: dict[int, str] = {}
    errors: list[dict[str, Any]] = []
    warnings: list[str] = []
    for sheet_index, worksheet in enumerate(workbook.worksheets):
        parsed, sheet_ranges = _parse_sheet(
            worksheet,
            workbook_name=safe_name,
            sheet_index=sheet_index,
        )
        chapter_no = int(parsed['chapter_no'])
        if chapter_no in chapter_numbers:
            duplicate_error = _error(
                'DUPLICATE_CHAPTER_NUMBER',
                f'Sheet {worksheet.title!r} và {chapter_numbers[chapter_no]!r} cùng ánh xạ vào bài {chapter_no}.',
                workbook=safe_name,
                sheet=worksheet.title,
            )
            parsed['errors'].append(duplicate_error)
        else:
            chapter_numbers[chapter_no] = worksheet.title
        sheets.append(parsed)
        ranges[worksheet.title] = sheet_ranges
        errors.extend(parsed['errors'])
        warnings.extend(parsed['warnings'])

    warnings.extend(_embedded_images(data, ranges))

    duplicates: dict[str, list[tuple[str, dict[str, Any]]]] = defaultdict(list)
    for sheet in sheets:
        for question in sheet['questions']:
            key = _normalized_question_key(question.get('question_text_raw') or question['question_text'])
            if key:
                duplicates[key].append((sheet['sheet_name'], question))
    for occurrences in duplicates.values():
        if len(occurrences) < 2:
            continue
        refs: list[str] = []
        for sheet_name, question in occurrences:
            question['duplicate_in_source'] = True
            refs.append(f'{sheet_name}!{question["source_row"]}')
        warnings.append(
            f"Câu hỏi trùng trong nguồn cũ tại {', '.join(refs)}; vẫn giữ từng dòng và bắt buộc duyệt."
        )

    all_questions = [question for sheet in sheets for question in sheet['questions']]
    type_counts = Counter(item['question_type'] for item in all_questions)
    difficulty_counts = Counter(_difficulty_count_key(item) for item in all_questions)
    unclassified_difficulty_count = int(difficulty_counts.get('unclassified') or 0)
    if unclassified_difficulty_count:
        warnings.append(
            f'{unclassified_difficulty_count} câu không có NGƯỠNG/độ khó; vẫn được import '
            'và sẽ được phân bổ linh hoạt khi tạo Quiz sau khi duyệt.'
        )
    image_reference_count = sum(len(item['image_refs']) for item in all_questions)
    embedded_image_count = sum(len(item['embedded_media']) for item in all_questions)
    return {
        'filename': safe_name,
        'raw': data,
        'sha256': _sha256(data),
        'sheets': sheets,
        'sheet_count': len(sheets),
        'question_count': len(all_questions),
        'type_counts': dict(type_counts),
        'difficulty_counts': dict(difficulty_counts),
        'image_reference_count': image_reference_count,
        'embedded_image_count': embedded_image_count,
        'warnings': warnings,
        'errors': errors,
        'can_commit': bool(all_questions) and not errors,
    }


def _asset_input(item: Any) -> tuple[str, bytes]:
    if isinstance(item, tuple) and len(item) == 2:
        return str(item[0]), bytes(item[1])
    if isinstance(item, dict):
        return str(item.get('filename') or ''), bytes(item.get('raw') or b'')
    raise ValueError('File ảnh đính kèm không hợp lệ.')


def collect_legacy_assets(items: Iterable[Any] | None) -> dict[str, Any]:
    assets: dict[str, dict[str, Any]] = {}
    warnings: list[str] = []
    errors: list[dict[str, Any]] = []
    total_bytes = 0
    entry_count = 0

    def add_asset(name: str, raw: bytes, *, source: str) -> None:
        nonlocal total_bytes, entry_count
        entry_count += 1
        total_bytes += len(raw)
        if entry_count > MAX_ASSETS:
            raise ValueError(f'Vượt giới hạn {MAX_ASSETS} ảnh đính kèm.')
        if total_bytes > MAX_TOTAL_ASSET_BYTES:
            raise ValueError('Tổng dung lượng ảnh đính kèm vượt 100 MB.')
        safe_name = safe_upload_filename(name)
        key = _normalized_asset_name(safe_name)
        try:
            validated = validate_question_image(raw)
        except ValueError as exc:
            errors.append(_error(
                'INVALID_IMAGE', f'{safe_name}: {exc}', workbook=source, field='asset'
            ))
            return
        existing = assets.get(key)
        if existing:
            if existing['sha256'] != validated.sha256:
                errors.append(_error(
                    'CONFLICTING_ASSET',
                    f'Hai file ảnh cùng tên {safe_name!r} nhưng khác nội dung.',
                    workbook=source,
                    field='asset',
                ))
            return
        assets[key] = {
            'asset_key': key,
            'filename': safe_name,
            'raw': raw,
            'sha256': validated.sha256,
            'mime_type': validated.mime_type,
            'width': validated.width,
            'height': validated.height,
            'size_bytes': len(raw),
            'source': source,
        }

    for item in list(items or []):
        filename, raw = _asset_input(item)
        suffix = Path(filename).suffix.lower()
        if suffix in IMAGE_SUFFIXES:
            try:
                add_asset(filename, raw, source=filename)
            except ValueError as exc:
                errors.append(_error('ASSET_LIMIT', str(exc), workbook=filename, field='asset'))
                break
            continue
        if suffix not in ASSET_ARCHIVE_SUFFIXES:
            errors.append(_error(
                'UNSUPPORTED_ASSET',
                f'Chỉ hỗ trợ ảnh PNG/JPEG/WebP hoặc ZIP, nhận {filename!r}.',
                workbook=filename,
                field='asset',
            ))
            continue
        try:
            infos = _validate_zip_archive(raw, label=f'ZIP {filename}')
            with zipfile.ZipFile(io.BytesIO(raw)) as archive:
                for info in infos:
                    if info.is_dir():
                        continue
                    member_name = str(info.filename or '')
                    if Path(member_name).suffix.lower() not in IMAGE_SUFFIXES:
                        warnings.append(f'Bỏ qua file không phải ảnh trong {filename}: {member_name}.')
                        continue
                    add_asset(member_name, archive.read(info), source=filename)
        except ValueError as exc:
            errors.append(_error('INVALID_ASSET_ARCHIVE', str(exc), workbook=filename, field='asset'))
        if errors and errors[-1].get('code') == 'ASSET_LIMIT':
            break
    return {
        'assets': list(assets.values()),
        'asset_count': len(assets),
        'total_bytes': total_bytes,
        'warnings': warnings,
        'errors': errors,
    }


def _subject_matches_filename(subject: Subject, filename: str) -> bool:
    stem = Path(filename).stem.strip()
    code = str(subject.code or '').strip()
    if not code or not stem.casefold().startswith(code.casefold()):
        return False
    suffix = stem[len(code) :]
    return not suffix or not suffix[0].isalnum()


def build_legacy_quiz_preview(
    db: Session,
    *,
    workbooks: Iterable[Any],
    assets: Iterable[Any] | None = None,
    visible_subject_ids: set[str] | None = None,
    actor: str | None = None,
) -> dict[str, Any]:
    workbook_inputs = list(workbooks or [])
    if not workbook_inputs:
        raise ValueError('Cần chọn ít nhất 1 file Excel.')
    if len(workbook_inputs) > MAX_WORKBOOKS:
        raise ValueError(f'Mỗi lần import tối đa {MAX_WORKBOOKS} file Excel.')

    subject_query = db.query(Subject).filter(func.lower(Subject.status) == 'active')
    if visible_subject_ids is not None:
        if not visible_subject_ids:
            subjects: list[Subject] = []
        else:
            subjects = subject_query.filter(Subject.id.in_(visible_subject_ids)).all()
    else:
        subjects = subject_query.all()

    asset_result = collect_legacy_assets(assets)
    asset_lookup = {item['asset_key']: item for item in asset_result['assets']}
    preview_assets = list(asset_result['assets'])
    errors = list(asset_result['errors'])
    warnings = list(asset_result['warnings'])
    parsed_workbooks: list[dict[str, Any]] = []
    total_questions = 0

    for raw_item in workbook_inputs:
        filename, raw = _asset_input(raw_item)
        try:
            parsed = parse_legacy_quiz_workbook(raw, filename=filename)
        except ValueError as exc:
            safe_name = safe_upload_filename(filename)
            failure = _error('INVALID_WORKBOOK', str(exc), workbook=safe_name)
            parsed_workbooks.append({
                'filename': safe_name,
                'raw': bytes(raw),
                'sha256': _sha256(bytes(raw)),
                'subject_id': None,
                'subject_code': None,
                'subject_name': None,
                'sheets': [],
                'sheet_count': 0,
                'question_count': 0,
                'type_counts': {},
                'difficulty_counts': {},
                'image_reference_count': 0,
                'embedded_image_count': 0,
                'warnings': [],
                'errors': [failure],
            })
            errors.append(failure)
            continue

        matches = [subject for subject in subjects if _subject_matches_filename(subject, parsed['filename'])]
        if not matches:
            match_error = _error(
                'SUBJECT_NOT_FOUND',
                'Không tìm thấy môn active, có quyền truy cập và có mã trùng chính xác đầu tên file.',
                workbook=parsed['filename'],
                field='filename',
            )
            parsed['errors'].append(match_error)
        elif len(matches) > 1:
            match_error = _error(
                'AMBIGUOUS_SUBJECT',
                f"Tên file khớp nhiều môn: {', '.join(sorted(item.code for item in matches))}.",
                workbook=parsed['filename'],
                field='filename',
            )
            parsed['errors'].append(match_error)
        else:
            subject = matches[0]
            parsed.update({
                'subject_id': subject.id,
                'department_id': subject.department_id,
                'subject_code': subject.code,
                'subject_name': subject.name,
            })

        errors.extend(parsed['errors'])
        warnings.extend(parsed['warnings'])
        total_questions += int(parsed['question_count'])

        for sheet in parsed['sheets']:
            for question in sheet['questions']:
                media_keys: list[str] = []
                embedded = list(question.pop('embedded_media', []) or [])
                for embedded_index, media in enumerate(embedded, start=1):
                    asset_key = (
                        f'embedded:{parsed["sha256"]}:{sheet["sheet_name"]}:'
                        f'{question["source_row"]}:{embedded_index}'
                    )
                    media['asset_key'] = asset_key
                    media['size_bytes'] = len(media.get('raw') or b'')
                    preview_assets.append(media)
                    asset_lookup[asset_key] = media
                    media_keys.append(asset_key)

                unresolved_markers: list[str] = []
                for marker in question.get('image_refs') or []:
                    key = _normalized_asset_name(marker)
                    if key in asset_lookup:
                        media_keys.append(key)
                    else:
                        unresolved_markers.append(marker)
                if unresolved_markers and len(embedded) >= len(unresolved_markers):
                    unresolved_markers = []
                if unresolved_markers:
                    missing_error = _error(
                        'MISSING_IMAGE',
                        f"Thiếu ảnh: {', '.join(unresolved_markers)}. Hãy tải ảnh trực tiếp hoặc trong ZIP.",
                        workbook=parsed['filename'],
                        sheet=sheet['sheet_name'],
                        row=question['source_row'],
                        field='QUESTION',
                    )
                    question['errors'].append(missing_error)
                    sheet['errors'].append(missing_error)
                    parsed['errors'].append(missing_error)
                    errors.append(missing_error)
                question['media_asset_keys'] = list(dict.fromkeys(media_keys))
                if len(question['media_asset_keys']) > 4:
                    media_error = _error(
                        'TOO_MANY_IMAGES',
                        'Mỗi câu hỏi tối đa 4 ảnh.',
                        workbook=parsed['filename'],
                        sheet=sheet['sheet_name'],
                        row=question['source_row'],
                        field='QUESTION',
                    )
                    question['errors'].append(media_error)
                    sheet['errors'].append(media_error)
                    parsed['errors'].append(media_error)
                    errors.append(media_error)
        parsed_workbooks.append(parsed)

    if len(preview_assets) > MAX_ASSETS:
        errors.append(_error(
            'MAX_ASSETS',
            f'Mỗi lần import tối đa {MAX_ASSETS} ảnh, kể cả ảnh nhúng trong Excel.',
            field='asset',
        ))
    preview_asset_bytes = sum(len(item.get('raw') or b'') for item in preview_assets)
    if preview_asset_bytes > MAX_TOTAL_ASSET_BYTES:
        errors.append(_error(
            'MAX_ASSET_BYTES',
            'Tổng dung lượng ảnh, kể cả ảnh nhúng trong Excel, vượt 100 MB.',
            field='asset',
        ))
    if total_questions > MAX_QUESTIONS:
        errors.append(_error(
            'MAX_QUESTIONS', f'Mỗi lần import tối đa {MAX_QUESTIONS} câu hỏi.'
        ))

    type_counts: Counter[str] = Counter()
    difficulty_counts: Counter[str] = Counter()
    sheet_count = 0
    image_count = 0
    for workbook in parsed_workbooks:
        sheet_count += int(workbook.get('sheet_count') or 0)
        type_counts.update(workbook.get('type_counts') or {})
        difficulty_counts.update(workbook.get('difficulty_counts') or {})
        for sheet in workbook.get('sheets') or []:
            for question in sheet.get('questions') or []:
                image_count += len(question.get('media_asset_keys') or [])

    return {
        'preview_version': 2,
        'created_at': _utcnow_iso(),
        'requested_by': actor,
        'target_term': LEGACY_IMPORT_TERM,
        'workbooks': parsed_workbooks,
        'assets': preview_assets,
        'workbook_count': len(parsed_workbooks),
        'sheet_count': sheet_count,
        'question_count': total_questions,
        'original_question_count': total_questions,
        'skipped_invalid_question_count': 0,
        'skipped_invalid_questions': [],
        'type_counts': dict(type_counts),
        'difficulty_counts': dict(difficulty_counts),
        'image_count': image_count,
        'warnings': list(dict.fromkeys(warnings)),
        'errors': errors,
        'can_commit': bool(total_questions) and not errors,
    }


def _legacy_error_identity(error: dict[str, Any]) -> tuple[str, ...]:
    return tuple(
        str(error.get(field) or '')
        for field in ('code', 'message', 'workbook', 'sheet', 'row', 'field')
    )


def _legacy_question_stats(preview: dict[str, Any]) -> dict[str, int]:
    invalid_count = 0
    missing_image_count = 0
    for workbook in preview.get('workbooks') or []:
        for sheet in workbook.get('sheets') or []:
            for question in sheet.get('questions') or []:
                question_errors = list(question.get('errors') or [])
                if not question_errors:
                    continue
                invalid_count += 1
                if any(str(error.get('code') or '') == 'MISSING_IMAGE' for error in question_errors):
                    missing_image_count += 1
    return {
        'invalid_question_count': invalid_count,
        'missing_image_question_count': missing_image_count,
    }


def _legacy_question_counts(questions: Iterable[dict[str, Any]]) -> tuple[dict[str, int], dict[str, int]]:
    question_list = list(questions)
    return (
        dict(Counter(str(item.get('question_type') or '') for item in question_list)),
        dict(Counter(_difficulty_count_key(item) for item in question_list)),
    )


def discard_invalid_legacy_quiz_questions(preview: dict[str, Any]) -> dict[str, Any]:
    """Remove question-scoped failures while preserving every non-question blocker."""
    payload = deepcopy(preview)
    discarded = list(payload.get('skipped_invalid_questions') or [])
    discarded_error_ids: set[tuple[str, ...]] = set()
    discarded_now: list[dict[str, Any]] = []

    for workbook in payload.get('workbooks') or []:
        retained_sheets: list[dict[str, Any]] = []
        workbook_questions: list[dict[str, Any]] = []
        for sheet in workbook.get('sheets') or []:
            retained_questions: list[dict[str, Any]] = []
            sheet_error_ids: set[tuple[str, ...]] = set()
            for question in sheet.get('questions') or []:
                question_errors = list(question.get('errors') or [])
                if not question_errors:
                    retained_questions.append(question)
                    continue
                identities = {_legacy_error_identity(error) for error in question_errors}
                sheet_error_ids.update(identities)
                discarded_error_ids.update(identities)
                discarded_now.append({
                    'workbook': workbook.get('filename'),
                    'sheet': sheet.get('sheet_name'),
                    'row': question.get('source_row'),
                    'question_no': question.get('question_no'),
                    'error_codes': list(dict.fromkeys(
                        str(error.get('code') or 'INVALID_QUESTION')
                        for error in question_errors
                    )),
                    'image_refs': list(question.get('image_refs') or []),
                })

            sheet['questions'] = retained_questions
            sheet['errors'] = [
                error
                for error in sheet.get('errors') or []
                if _legacy_error_identity(error) not in sheet_error_ids
            ]
            sheet['question_count'] = len(retained_questions)
            sheet['type_counts'], sheet['difficulty_counts'] = _legacy_question_counts(
                retained_questions
            )
            retained_sheets.append(sheet)
            workbook_questions.extend(retained_questions)

        workbook['sheets'] = retained_sheets
        workbook['sheet_count'] = len(retained_sheets)
        workbook['question_count'] = len(workbook_questions)
        workbook['type_counts'], workbook['difficulty_counts'] = _legacy_question_counts(
            workbook_questions
        )
        workbook['image_reference_count'] = sum(
            len(question.get('image_refs') or []) for question in workbook_questions
        )
        workbook['errors'] = [
            error
            for error in workbook.get('errors') or []
            if _legacy_error_identity(error) not in discarded_error_ids
        ]

    payload['errors'] = [
        error
        for error in payload.get('errors') or []
        if _legacy_error_identity(error) not in discarded_error_ids
    ]
    discarded.extend(discarded_now)
    payload['skipped_invalid_questions'] = discarded
    payload['skipped_invalid_question_count'] = len(discarded)

    remaining_questions = [
        question
        for workbook in payload.get('workbooks') or []
        for sheet in workbook.get('sheets') or []
        for question in sheet.get('questions') or []
    ]
    payload['question_count'] = len(remaining_questions)
    payload['sheet_count'] = sum(
        int(workbook.get('sheet_count') or 0) for workbook in payload.get('workbooks') or []
    )
    payload['type_counts'], payload['difficulty_counts'] = _legacy_question_counts(
        remaining_questions
    )
    payload['image_count'] = sum(
        len(question.get('media_asset_keys') or []) for question in remaining_questions
    )
    payload['can_commit'] = bool(remaining_questions) and not payload['errors']
    payload['preview_version'] = 2
    payload['updated_at'] = _utcnow_iso()
    if discarded_now:
        payload['warnings'] = list(dict.fromkeys([
            *(payload.get('warnings') or []),
            (
                f'Đã bỏ qua {len(discarded_now)} câu lỗi theo xác nhận của người dùng; '
                'các câu này không được tạo trong ngân hàng đề.'
            ),
        ]))
    return payload


def _legacy_preview_key(token: str) -> str:
    if not re.fullmatch(r'[0-9a-f]{32}', str(token or '')):
        raise ValueError('Preview token không hợp lệ.')
    return f'question-bank/_pending-operation-files/legacy-quiz-{token}/preview.json'


def legacy_preview_reference(token: str, storage: ObjectStorage | None = None) -> str:
    target = storage or get_object_storage()
    return target.reference_for_key(_legacy_preview_key(token))


def replace_legacy_quiz_preview(
    token: str,
    preview: dict[str, Any],
    *,
    storage: ObjectStorage | None = None,
) -> str:
    target = storage or get_object_storage()
    payload = deepcopy(preview)
    payload['preview_token'] = token
    payload['updated_at'] = _utcnow_iso()
    return target.put_json(_legacy_preview_key(token), payload)


def persist_legacy_quiz_preview(
    preview: dict[str, Any],
    *,
    storage: ObjectStorage | None = None,
) -> tuple[str, str]:
    target = storage or get_object_storage()
    token = uuid.uuid4().hex
    prefix = f'question-bank/_pending-operation-files/legacy-quiz-{token}'
    payload = deepcopy(preview)
    payload['preview_token'] = token
    payload['created_at'] = payload.get('created_at') or _utcnow_iso()
    written: list[str] = []
    try:
        for workbook_index, workbook in enumerate(payload.get('workbooks') or [], start=1):
            raw = bytes(workbook.pop('raw', b'') or b'')
            key = f'{prefix}/workbooks/{workbook_index:02d}-{safe_upload_filename(workbook["filename"])}'
            reference = target.put_bytes(
                key,
                raw,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            written.append(reference)
            workbook['pending_reference'] = reference
        for asset_index, asset in enumerate(payload.get('assets') or [], start=1):
            raw = bytes(asset.pop('raw', b'') or b'')
            suffix = Path(str(asset.get('filename') or '')).suffix.lower() or '.bin'
            key = f'{prefix}/assets/{asset_index:04d}-{asset["sha256"][:12]}{suffix}'
            reference = target.put_bytes(key, raw, content_type=asset.get('mime_type'))
            written.append(reference)
            asset['pending_reference'] = reference
        preview_key = f'{prefix}/preview.json'
        reference = target.put_json(preview_key, payload)
        written.append(reference)
        return token, reference
    except Exception:
        for reference in reversed(written):
            try:
                target.delete(reference, missing_ok=True)
            except StorageError:
                pass
        raise


def load_legacy_quiz_preview(
    token: str,
    *,
    storage: ObjectStorage | None = None,
) -> dict[str, Any]:
    if not re.fullmatch(r'[0-9a-f]{32}', str(token or '')):
        raise ValueError('Preview token không hợp lệ.')
    target = storage or get_object_storage()
    try:
        return target.read_json(legacy_preview_reference(token, target))
    except StorageError as exc:
        raise ValueError('Không tìm thấy preview import. Hãy tải file lên lại.') from exc


def public_legacy_quiz_preview(preview: dict[str, Any]) -> dict[str, Any]:
    question_stats = _legacy_question_stats(preview)
    workbooks: list[dict[str, Any]] = []
    for workbook in preview.get('workbooks') or []:
        sheets = []
        for sheet in workbook.get('sheets') or []:
            sheets.append({
                'sheet_name': sheet.get('sheet_name'),
                'chapter_no': sheet.get('chapter_no'),
                'chapter_title': sheet.get('chapter_title'),
                'question_count': int(sheet.get('question_count') or 0),
                'type_counts': sheet.get('type_counts') or {},
                'difficulty_counts': sheet.get('difficulty_counts') or {},
                'warning_count': len(sheet.get('warnings') or []),
                'error_count': len(sheet.get('errors') or []),
            })
        workbooks.append({
            'filename': workbook.get('filename'),
            'subject_id': workbook.get('subject_id'),
            'subject_code': workbook.get('subject_code'),
            'subject_name': workbook.get('subject_name'),
            'sheet_count': int(workbook.get('sheet_count') or 0),
            'question_count': int(workbook.get('question_count') or 0),
            'type_counts': workbook.get('type_counts') or {},
            'difficulty_counts': workbook.get('difficulty_counts') or {},
            'image_reference_count': int(workbook.get('image_reference_count') or 0),
            'embedded_image_count': int(workbook.get('embedded_image_count') or 0),
            'warning_count': len(workbook.get('warnings') or []),
            'error_count': len(workbook.get('errors') or []),
            'sheets': sheets,
        })
    can_commit = bool(preview.get('can_commit'))
    skipped_invalid_count = int(preview.get('skipped_invalid_question_count') or 0)
    if can_commit and skipped_invalid_count:
        message = (
            f'Đã loại {skipped_invalid_count} câu lỗi khỏi lần import. '
            'Các câu còn lại có thể được import vào SU26.'
        )
    elif can_commit:
        message = 'Preview hợp lệ. Có thể tạo job import vào SU26.'
    elif question_stats['invalid_question_count']:
        message = (
            'Có câu hỏi lỗi. Hãy bổ sung ảnh/điều chỉnh file rồi kiểm tra lại, '
            'hoặc bỏ qua toàn bộ câu lỗi.'
        )
    else:
        message = 'Preview còn lỗi cấp file, môn hoặc sheet và chưa thể import.'
    return {
        'ok': True,
        'preview_token': str(preview.get('preview_token') or ''),
        'target_term': str(preview.get('target_term') or LEGACY_IMPORT_TERM),
        'workbook_count': int(preview.get('workbook_count') or 0),
        'sheet_count': int(preview.get('sheet_count') or 0),
        'question_count': int(preview.get('question_count') or 0),
        'original_question_count': int(
            preview.get('original_question_count') or preview.get('question_count') or 0
        ),
        **question_stats,
        'skipped_invalid_question_count': skipped_invalid_count,
        'skipped_invalid_questions': list(preview.get('skipped_invalid_questions') or []),
        'can_skip_invalid_questions': question_stats['invalid_question_count'] > 0,
        'type_counts': preview.get('type_counts') or {},
        'difficulty_counts': preview.get('difficulty_counts') or {},
        'image_count': int(preview.get('image_count') or 0),
        'workbooks': workbooks,
        'warnings': list(preview.get('warnings') or []),
        'errors': list(preview.get('errors') or []),
        'can_commit': can_commit,
        'message': message,
    }


def _editable_bank_version(
    db: Session,
    *,
    subject: Subject,
    offering: SubjectOffering,
    chapter: SubjectChapter,
    actor: str | None,
) -> tuple[QuestionBankVersion, bool]:
    existing = (
        db.query(QuestionBankVersion)
        .filter(
            QuestionBankVersion.subject_id == subject.id,
            QuestionBankVersion.subject_offering_id == offering.id,
            QuestionBankVersion.chapter_id == chapter.id,
            func.lower(QuestionBankVersion.status).in_(['draft', 'reviewing']),
        )
        .order_by(QuestionBankVersion.version_no.desc(), QuestionBankVersion.created_at.desc())
        .with_for_update()
        .first()
    )
    if existing:
        return existing, False
    max_version = db.query(func.max(QuestionBankVersion.version_no)).filter(
        QuestionBankVersion.subject_id == subject.id,
        QuestionBankVersion.chapter_id == chapter.id,
    ).scalar()
    version_no = int(max_version or 0) + 1
    version = QuestionBankVersion(
        subject_id=subject.id,
        chapter_id=chapter.id,
        subject_offering_id=offering.id,
        version_no=version_no,
        version_code=f'v{version_no}.0',
        title=f'{subject.code} {LEGACY_IMPORT_TERM} - {chapter.title}',
        change_note='Tạo tự động từ import quiz CMS cũ.',
        status='draft',
        created_by=actor,
        metadata_json={'created_by_legacy_quiz_import': True, 'target_term': LEGACY_IMPORT_TERM},
    )
    db.add(version)
    db.flush()
    return version, True


def _ensure_offering(
    db: Session,
    *,
    subject: Subject,
    actor: str | None,
) -> tuple[SubjectOffering, bool]:
    code = f'{subject.code}_{LEGACY_IMPORT_TERM}'
    offering = (
        db.query(SubjectOffering)
        .filter(
            SubjectOffering.subject_id == subject.id,
            or_(
                func.upper(SubjectOffering.code) == code.upper(),
                func.upper(SubjectOffering.term) == LEGACY_IMPORT_TERM,
            ),
        )
        .order_by(
            (func.upper(SubjectOffering.code) == code.upper()).desc(),
            SubjectOffering.created_at.asc(),
        )
        .with_for_update()
        .first()
    )
    if offering:
        return offering, False
    offering = SubjectOffering(
        department_id=subject.department_id,
        subject_id=subject.id,
        code=code,
        name=f'{subject.name} - {LEGACY_IMPORT_TERM}',
        term=LEGACY_IMPORT_TERM,
        version_code='v1.0',
        status='draft',
        metadata_json={'created_by_legacy_quiz_import': True, 'target_term': LEGACY_IMPORT_TERM},
        created_by=actor,
    )
    db.add(offering)
    db.flush()
    return offering, True


def _ensure_chapter(
    db: Session,
    *,
    subject: Subject,
    offering: SubjectOffering,
    sheet: dict[str, Any],
) -> tuple[SubjectChapter, bool]:
    chapter_no = int(sheet['chapter_no'])
    chapter = (
        db.query(SubjectChapter)
        .filter(
            SubjectChapter.subject_id == subject.id,
            SubjectChapter.subject_offering_id == offering.id,
            SubjectChapter.chapter_no == chapter_no,
        )
        .with_for_update()
        .first()
    )
    if chapter:
        return chapter, False
    chapter = SubjectChapter(
        subject_id=subject.id,
        subject_offering_id=offering.id,
        chapter_no=chapter_no,
        title=f'Bài {chapter_no}'[:255],
        description=f'Khởi tạo từ sheet {sheet.get("sheet_name") or chapter_no} của quiz CMS cũ.',
        sort_order=chapter_no,
        status='active',
    )
    db.add(chapter)
    db.flush()
    return chapter, True


def _cleanup_preview_objects(preview: dict[str, Any], storage: ObjectStorage) -> None:
    references = [
        workbook.get('pending_reference')
        for workbook in preview.get('workbooks') or []
    ]
    references.extend(asset.get('pending_reference') for asset in preview.get('assets') or [])
    token = str(preview.get('preview_token') or '')
    if token:
        references.append(legacy_preview_reference(token, storage))
    for reference in references:
        if not reference:
            continue
        try:
            storage.delete(str(reference), missing_ok=True)
        except StorageError:
            pass


def import_legacy_quiz_preview(
    db: Session,
    *,
    preview: dict[str, Any] | None = None,
    preview_token: str | None = None,
    actor: str | None = None,
    storage: ObjectStorage | None = None,
    progress_callback: Callable[[int, int, str], None] | None = None,
    cleanup_preview: bool = True,
) -> dict[str, Any]:
    target = storage or get_object_storage()
    payload = preview or load_legacy_quiz_preview(str(preview_token or ''), storage=target)
    token = str(payload.get('preview_token') or preview_token or '')
    requested_by = str(payload.get('requested_by') or '')
    if requested_by and actor and requested_by != actor:
        raise ValueError('Preview này thuộc về người dùng khác.')
    if not payload.get('can_commit') or payload.get('errors'):
        raise ValueError('Preview có lỗi chặn, không thể import.')
    if str(payload.get('target_term') or '') != LEGACY_IMPORT_TERM:
        raise ValueError(f'Import quiz CMS cũ chỉ được phép vào phiên bản {LEGACY_IMPORT_TERM}.')

    assets = {str(item['asset_key']): item for item in payload.get('assets') or []}
    total_sheets = max(1, sum(len(item.get('sheets') or []) for item in payload.get('workbooks') or []))
    completed_sheets = 0
    created_questions = 0
    skipped_questions = 0
    created_offerings = 0
    created_chapters = 0
    created_bank_versions = 0
    created_materials = 0
    skipped_invalid_questions = int(payload.get('skipped_invalid_question_count') or 0)
    bank_version_ids: set[str] = set()
    permanent_workbooks: dict[str, str] = {}

    for workbook in payload.get('workbooks') or []:
        subject_id = str(workbook.get('subject_id') or '')
        subject = (
            db.query(Subject)
            .filter(Subject.id == subject_id, func.lower(Subject.status) == 'active')
            .with_for_update()
            .first()
        )
        if not subject:
            raise ValueError(f'Môn của file {workbook.get("filename")} không còn active hoặc không tồn tại.')
        workbook_raw = target.read_bytes(str(workbook['pending_reference']))
        workbook_hash = str(workbook.get('sha256') or _sha256(workbook_raw))
        permanent_key = (
            f'question-bank/legacy-quiz-imports/{_safe_slug(subject.code)}/{LEGACY_IMPORT_TERM}/'
            f'{token}/{workbook_hash[:12]}-{safe_upload_filename(str(workbook["filename"]))}'
        )
        permanent_reference = target.put_bytes(
            permanent_key,
            workbook_raw,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        permanent_workbooks[workbook_hash] = permanent_reference

        offering, offering_created = _ensure_offering(db, subject=subject, actor=actor)
        created_offerings += int(offering_created)
        db.commit()

        for sheet in workbook.get('sheets') or []:
            written_media: list[str] = []
            try:
                subject = db.query(Subject).filter(Subject.id == subject.id).with_for_update().one()
                offering = db.query(SubjectOffering).filter(SubjectOffering.id == offering.id).with_for_update().one()
                chapter, chapter_created = _ensure_chapter(
                    db,
                    subject=subject,
                    offering=offering,
                    sheet=sheet,
                )
                created_chapters += int(chapter_created)
                bank_version, bank_created = _editable_bank_version(
                    db,
                    subject=subject,
                    offering=offering,
                    chapter=chapter,
                    actor=actor,
                )
                created_bank_versions += int(bank_created)
                bank_version_ids.add(bank_version.id)

                material_title = f'Quiz CMS cũ - {workbook["filename"]} - {sheet["sheet_name"]}'[:255]
                material = (
                    db.query(LearningMaterialVersion)
                    .filter(
                        LearningMaterialVersion.bank_version_id == bank_version.id,
                        LearningMaterialVersion.content_hash == workbook_hash,
                        LearningMaterialVersion.title == material_title,
                        LearningMaterialVersion.change_type == 'legacy_quiz_import',
                    )
                    .first()
                )
                if not material:
                    max_material_version = db.query(func.max(LearningMaterialVersion.version_no)).filter(
                        LearningMaterialVersion.bank_version_id == bank_version.id,
                    ).scalar()
                    material = LearningMaterialVersion(
                        subject_id=subject.id,
                        chapter_id=chapter.id,
                        subject_offering_id=offering.id,
                        bank_version_id=bank_version.id,
                        title=material_title,
                        file_name=str(workbook['filename']),
                        file_type='xlsx',
                        storage_path=permanent_reference,
                        content_hash=workbook_hash,
                        version_no=int(max_material_version or 0) + 1,
                        change_type='legacy_quiz_import',
                        uploaded_by=actor,
                        status='active',
                    )
                    db.add(material)
                    db.flush()
                    created_materials += 1

                for question_data in sheet.get('questions') or []:
                    source_node_id = (
                        f'legacy-quiz:{token}:{workbook_hash[:16]}:'
                        f'{_safe_slug(sheet["sheet_name"], max_len=48)}:{question_data["source_row"]}'
                    )
                    source_ref = (
                        f'{workbook["filename"]}#{sheet["sheet_name"]}!row={question_data["source_row"]}'
                    )[:1024]
                    existing = db.query(Question).filter(
                        Question.bank_version_id == bank_version.id,
                        Question.source_node_id == source_node_id,
                        Question.source_ref == source_ref,
                    ).first()
                    if existing:
                        skipped_questions += 1
                        continue
                    evidence = {
                        'import_batch': token,
                        'workbook': workbook['filename'],
                        'workbook_sha256': workbook_hash,
                        'sheet': sheet['sheet_name'],
                        'source_row': question_data['source_row'],
                        'question_no': question_data.get('question_no'),
                        'question_text_raw': question_data.get('question_text_raw'),
                        'image_refs': question_data.get('image_refs') or [],
                        'type_raw': question_data.get('question_type_raw'),
                        'threshold_raw': question_data.get('threshold_raw'),
                        'difficulty_raw': question_data.get('difficulty_raw'),
                        'difficulty_classified': bool(question_data.get('difficulty_classified')),
                        'concept_classified': False,
                        'correct_key': question_data.get('correct_key'),
                        'imported_by': actor,
                        'duplicate_in_legacy_source': bool(question_data.get('duplicate_in_source')),
                    }
                    family_digest = hashlib.sha256(source_node_id.encode('utf-8')).hexdigest()[:32]
                    quality_flags = ['legacy_import_requires_review', 'legacy_import_unclassified_concept']
                    if not bool(question_data.get('difficulty_classified')):
                        quality_flags.append('legacy_import_unclassified_difficulty')
                    if question_data.get('duplicate_in_source'):
                        quality_flags.append('duplicate_in_legacy_source')
                    question = Question(
                        course_id=f'bank:{bank_version.id}',
                        department_id=subject.department_id,
                        subject_id=subject.id,
                        subject_chapter_id=chapter.id,
                        bank_version_id=bank_version.id,
                        material_version_id=material.id,
                        lesson_id=f'{LEGACY_IMPORT_TERM}:chapter:{chapter.chapter_no}',
                        lesson_title=chapter.title,
                        block_id=f'bank-version:{bank_version.id}:legacy-quiz-import',
                        topic=chapter.title,
                        question_family_id=f'legacy-{family_digest}',
                        variant_no=1,
                        source_evidence=json.dumps(evidence, ensure_ascii=False, separators=(',', ':')),
                        difficulty=question_data['difficulty'],
                        cognitive_level='remember',
                        learning_objective='',
                        pedagogy_json={},
                        authoring_mode='import',
                        created_by=actor,
                        question_text=question_data['question_text'],
                        question_hash=question_fingerprint(
                            question_data['question_text'],
                            course_id=f'bank:{bank_version.id}',
                            source_node_id=source_node_id,
                            difficulty=question_data['difficulty'],
                        ),
                        option_a='',
                        option_b='',
                        option_c='',
                        option_d='',
                        correct_answer='A',
                        explanation='',
                        source_ref=source_ref,
                        source_type='legacy_quiz_excel',
                        source_node_id=source_node_id,
                        source_node_title=str(sheet['sheet_name'])[:512],
                        chapter_node_id=chapter.id,
                        chapter_title=chapter.title,
                        source_excerpt=f'Excel row {question_data["source_row"]}',
                        tags=['legacy_quiz_import', LEGACY_IMPORT_TERM.lower()],
                        ai_rationale='',
                        quality_score=0.0,
                        quality_flags=quality_flags,
                        is_duplicate=bool(question_data.get('duplicate_in_source')),
                        model_provider='manual',
                        model_name='legacy_quiz_import',
                        status='pending_review',
                    )
                    apply_canonical_content(
                        question,
                        question_data['question_type'],
                        question_data['content'],
                    )
                    db.add(question)
                    db.flush()
                    db.add(QuestionReviewLog(
                        id=str(uuid.uuid4()),
                        question_id=question.id,
                        old_status='imported',
                        new_status='pending_review',
                        actor=actor or 'system',
                        note=(
                            f'Import từ quiz CMS cũ vào {LEGACY_IMPORT_TERM}; '
                            f'file {workbook["filename"]}, sheet {sheet["sheet_name"]}, '
                            f'dòng {question_data["source_row"]}. Câu hỏi cần được duyệt.'
                        ),
                    ))

                    for sort_order, asset_key in enumerate(question_data.get('media_asset_keys') or []):
                        asset = assets.get(str(asset_key))
                        if not asset:
                            raise ValueError(f'Không tìm thấy asset {asset_key!r} trong preview.')
                        raw = target.read_bytes(str(asset['pending_reference']))
                        validated = validate_question_image(
                            raw,
                            declared_content_type=str(asset.get('mime_type') or ''),
                        )
                        media_id = str(uuid.uuid4())
                        filename = safe_upload_filename(str(asset.get('filename') or f'image-{sort_order + 1}'))
                        extension = validated.extension
                        media_key = (
                            f'question-media/{bank_version.id}/{question.id}/'
                            f'{media_id}-{_safe_slug(Path(filename).stem)}-{validated.sha256[:12]}{extension}'
                        )
                        media_reference = target.put_bytes(
                            media_key,
                            raw,
                            content_type=validated.mime_type,
                        )
                        written_media.append(media_reference)
                        db.add(QuestionMedia(
                            id=media_id,
                            question_id=question.id,
                            bank_version_id=bank_version.id,
                            media_role='prompt_image',
                            storage_reference=media_reference,
                            file_name=f'{Path(filename).stem}{extension}',
                            mime_type=validated.mime_type,
                            size_bytes=len(raw),
                            sha256=validated.sha256,
                            width=validated.width,
                            height=validated.height,
                            alt_text=f'Hình minh họa câu {question_data.get("question_no") or question_data["source_row"]}',
                            sort_order=sort_order,
                            created_by=actor,
                        ))
                    created_questions += 1
                db.commit()
            except Exception:
                db.rollback()
                for reference in written_media:
                    try:
                        target.delete(reference, missing_ok=True)
                    except StorageError:
                        pass
                raise

            completed_sheets += 1
            if progress_callback:
                progress_callback(
                    completed_sheets,
                    total_sheets,
                    f'Đã import {completed_sheets}/{total_sheets} sheet',
                )

    for bank_version_id in sorted(bank_version_ids):
        try:
            BankDashboardStatsService(db).refresh_for_bank_version(bank_version_id)
        except Exception:
            db.rollback()

    if cleanup_preview:
        _cleanup_preview_objects(payload, target)
    return {
        'ok': True,
        'preview_token': token,
        'target_term': LEGACY_IMPORT_TERM,
        'requested_by': actor,
        'created_question_count': created_questions,
        'skipped_question_count': skipped_questions,
        'skipped_invalid_question_count': skipped_invalid_questions,
        'pending_review_count': created_questions,
        'created_offering_count': created_offerings,
        'created_chapter_count': created_chapters,
        'created_bank_version_count': created_bank_versions,
        'created_material_count': created_materials,
        'bank_version_ids': sorted(bank_version_ids),
        'workbook_references': permanent_workbooks,
        'message': (
            (
                f'Đã import {created_questions} câu vào {LEGACY_IMPORT_TERM}; '
                f'đã loại {skipped_invalid_questions} câu lỗi; '
                'tất cả câu mới đang chờ duyệt.'
            )
            if skipped_invalid_questions
            else (
                f'Đã import {created_questions} câu vào {LEGACY_IMPORT_TERM}; '
                'tất cả câu mới đang chờ duyệt.'
            )
        ),
    }
